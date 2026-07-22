# -*- coding: utf-8 -*-
# report_builder 20260615_v1：新版日报工作簿构建(由参考生成器v9重构为带参数模块)
import json,re,sys
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment,Font,PatternFill,Border,Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict,Counter

def build_workbook(template, raw, screening_json, rejudge_json, vip_json, out, run_date="", src_start="", src_end=""):
    # -*- coding: utf-8 -*-
    def bid(u):
        m=re.search(r'bid-(\d+)',str(u or '')); return m.group(1) if m else None
    # ---- agent 数据 ----
    sc=json.load(open(screening_json,encoding="utf-8-sig"))
    screen={}
    for k in ("excluded","low_score","considered","recommended","results"):
        for it in (sc.get(k) or []):
            b=bid(it.get("url"))
            if b:
                screen[b]=it
    rjd=json.load(open(rejudge_json,encoding="utf-8-sig"))
    rjd=rjd.get("results") if isinstance(rjd,dict) else rjd
    rejudge={bid(it.get("url")):it for it in rjd if bid(it.get("url"))}
    # VIP 可选：合法全排除时 pipeline 可能不传 vip，或传入空路径
    vip={}
    if vip_json:
        try:
            from pathlib import Path as _Path
            if _Path(str(vip_json)).is_file():
                vp=json.load(open(vip_json,encoding="utf-8-sig"))
                vip={bid(it.get("url")):it for it in (vp.get("projects") or []) if bid(it.get("url"))}
        except Exception:
            vip={}
    # ---- 原始导出: bid_id -> 行字段 ----
    rawwb=load_workbook(raw); rws=rawwb.worksheets[0]
    RAWMAP={}
    for r in range(3, rws.max_row+1):
        if rws.cell(r,1).value in (None,""): continue
        ac=rws.cell(r,29); url=(ac.hyperlink.target if ac.hyperlink else None) or ac.value
        b=bid(url)
        if not b: continue
        g=lambda c: rws.cell(r,c).value
        RAWMAP[b]=dict(title=g(1),publish=g(2),bid_no=g(3),region=g(4),t1=g(5),t2=g(6),
            bid_deadline=g(10),budget=g(11),win_amount=g(12),buyer=g(13),buyer_contact=g(14),buyer_tel=g(15),
            agency=g(16),agency_contact=g(17),agency_tel=g(18),winner=g(19),win_contact=g(20),win_tel=g(21),
            kw=g(24),purchase=g(27),expert=g(28),url=f"https://www.qianlima.com/bid-{b}.html")
    def norm(d):
        d=str(d or "").strip()
        if d in ("keep","推荐","推荐候选","推荐项目","重点推荐"):
            return "keep"
        if d in ("needs_review","待人工复核","待复核","待定","人工复核"):
            return "needs_review"
        if d in ("reject","排除","不推荐","放弃"):
            return "reject"
        return d
    def dec(b):return norm((rejudge.get(b) or {}).get("final_decision") or (rejudge.get(b) or {}).get("decision"))
    def review_label(b):return {"keep":"推荐候选","needs_review":"待人工复核","reject":"排除"}.get(dec(b),"")
    def score_detail(b):
        s=(screen.get(b) or {}).get("scores") or {}
        if not s:return""
        return "金额%s+行业%s+地区%s+资质%s+时间%s+黑名单%s=%s"%(s.get("金额",0),s.get("行业",0),s.get("地区",0),s.get("资质",0),s.get("时间",0),s.get("黑名单",0),s.get("总分",s.get("原始总分","")))
    UI_NOISE = (
        "摘要信息", "招标详情", "进度跟踪", "商机推荐", "用手机查看此详情", "企业详情",
        "相似采购商", "立即查看", "可引荐人脉", "立即引荐", "历史招中标信息", "立即监控",
        "标书代写", "多份优惠", "查看原文", "收藏", "导出", "打印", "返回顶部",
        "在线咨询", "咨询热线", "下载APP", "服务号", "帮助中心",
    )
    BODY_END_MARKERS = ("附件（", "附件下载", "企业商情分析", "招标进度跟踪", "相关招标", "热门推荐")

    def text_value(value):
        return re.sub(r"\s+", " ", str(value or "")).strip()

    def format_budget(raw_value=None, amount_wan=None):
        raw=text_value(raw_value)
        if raw:
            if "万元" in raw:
                return raw
            if "元" in raw:
                raw=raw.replace("元","").replace(",","")
            try:
                value=float(raw.replace(",",""))
                return f"{value/10000:.2f}万元"
            except ValueError:
                return text_value(raw_value)
        if amount_wan not in (None, ""):
            try:
                return f"{float(amount_wan):.2f}万元"
            except (TypeError, ValueError):
                value=text_value(amount_wan)
                return value if "万" in value else value+"万元"
        return ""

    def clean_fragment(value):
        text=text_value(value)
        for token in UI_NOISE:
            text=text.replace(token," ")
        text=re.sub(r"(?:标书代写)?多份优惠", " ", text)
        return text_value(text)

    def evidence_snippets(b, limit=3):
        evidence=(rejudge.get(b) or {}).get("evidence") or []
        output=[]
        for item in evidence:
            value=(item.get("snippet") or item.get("text") or "") if isinstance(item,dict) else str(item)
            value=clean_fragment(value)
            if not value or "附件证据门槛" in value:
                continue
            if len(value)>180:
                value=value[:179]+"…"
            if value not in output:
                output.append(value)
            if len(output)>=limit:
                break
        return output

    def clean_body_text(b):
        item=vip.get(b) or {}
        raw=item.get("body_clean") or item.get("body_text") or item.get("content") or ""
        text=text_value(raw)
        title=text_value((RAWMAP.get(b) or {}).get("title") or item.get("title"))
        if title and title in text:
            text=text[text.find(title):]
        for marker in BODY_END_MARKERS:
            pos=text.find(marker)
            if pos>120:
                text=text[:pos]
        for token in UI_NOISE:
            text=text.replace(token," ")
        text=re.sub(r"\b\d+个\s*联系人\b"," ",text)
        text=re.sub(r"\b\d+人\s*可引荐人脉\b"," ",text)
        text=re.sub(r"\b\d+条\s*历史招中标信息\b"," ",text)
        return text_value(text)

    def field_from_body(b, labels, stop_labels=(), maximum=360):
        text=clean_body_text(b)
        if not text:
            return ""
        label_pattern="|".join(re.escape(x) for x in labels)
        stops=list(stop_labels)+[
            "质量要求","工期","施工期","服务期","交货期","合同履行期",
            "供应商资格","投标人资格","承包人资格","获取采购文件",
            "服务要求","用途","技术要求","价格构成","评定标准","付款方式","支付方式",
            "其它说明","拟采购的材料设备清单","报价相关要求","品目号",
            "采购项目编号","包件编号","项目编号","采购资金来源","采购方式","采购地址",
            "采购人信息","采购代理信息","时间信息","联系方式","二、","三、","四、",
        ]
        stop_pattern="|".join(re.escape(x) for x in stops)
        match=re.search(
            rf"(?:{label_pattern})\s*[：:]?\s*(.+?)(?=(?:{stop_pattern})\s*[：:]?|$)",
            text,
        )
        if not match:
            return ""
        value=text_value(match.group(1))
        return value if len(value)<=maximum else value[:maximum-1]+"…"

    def proj_type(b):
        it=rejudge.get(b) or {};return it.get("project_type") or it.get("business_direction") or ""

    def duration(b):
        item=vip.get(b) or {}
        if item.get("duration"):
            return text_value(item["duration"])
        text=clean_body_text(b)
        match=re.search(r"(?:工期|施工期|服务期|交货期|完成期限)\s*[：:]?\s*([^。；]{0,80}?\d+\s*(?:个月|日历天|个日历天|天|年|周))",text)
        return text_value(match.group(1)) if match else ""

    def payment(b):
        item=vip.get(b) or {}
        if item.get("payment"):
            return text_value(item["payment"])
        return field_from_body(b,("付款方式","支付方式"),maximum=120)

    def funding(b):
        item=vip.get(b) or {}
        if item.get("funding"):
            return text_value(item["funding"])
        return field_from_body(b,("资金来源","资金性质"),maximum=120)

    def qualification(b):
        return field_from_body(
            b,
            ("投标人资格","供应商资格","承包人资格","资格要求","特定资格要求"),
            ("服务要求","用途","技术要求","价格构成","评定标准","付款方式",
             "其它说明","拟采购的材料设备清单","获取采购文件","采购文件"),
            maximum=180,
        )

    def project_scope(b):
        primary=field_from_body(
            b,
            ("承包内容及规模","采购内容","项目内容","建设内容","采购需求","招标范围","工程概况"),
            ("质量要求","工期","施工期","服务期","合同履行期","品目号"),
            maximum=300,
        )
        named=field_from_body(
            b,
            ("采购项目名称","包件名称","项目名称","采购标的"),
            ("采购项目编号","包件编号","项目编号","采购资金来源","采购方式","品目号"),
            maximum=220,
        )
        generic={"详见招标文件","详见采购文件","详见询价文件","详见采购需求"}
        primary=clean_fragment(primary)
        named=clean_fragment(named)
        if primary in generic:
            primary=""
        if named in generic:
            named=""
        if primary and named and (primary in named or named in primary):
            return primary if len(primary)>=len(named) else named
        return primary or named

    def project_attachments(b):
        rows=(rejudge.get(b) or {}).get("page_attachments") or []
        output=[];seen=set()
        for row in rows:
            name=clean_fragment(row.get("name") or "附件")
            source=text_value(row.get("source_url"))
            real=text_value(row.get("real_url"))
            key=real or source or name
            if not key or key in seen:
                continue
            seen.add(key)
            output.append({
                "name":name,
                "url":source or real,
                "source_url":source,
                "real_url":real,
                "file_type":text_value(row.get("file_type")),
                "text_read_ok":bool(row.get("text_read_ok")),
                "error":text_value(row.get("error")),
            })
        return output

    def progress_timeline(b):
        rows=(rejudge.get(b) or {}).get("progress_timeline") or (vip.get(b) or {}).get("progress_timeline") or []
        output=[];seen=set()
        for row in rows:
            date=text_value(row.get("date"))
            node=text_value(row.get("node"))
            title=text_value(row.get("title"))
            key=(date,node,title)
            if (date or node or title) and key not in seen:
                seen.add(key);output.append({"date":date,"node":node,"title":title})
        if not output:
            rm=RAWMAP.get(b,{})
            output=[{
                "date":text_value(rm.get("publish")) or "未获取",
                "node":ntype(b),
                "title":text_value(rm.get("title") or (rejudge.get(b) or {}).get("title")),
            }]
        return output

    def score_total(b):
        scores=(screen.get(b) or {}).get("scores") or {}
        return scores.get("总分",scores.get("原始总分",(rejudge.get(b) or {}).get("score","未获取")))

    def analysis(b):
        item=rejudge.get(b) or {}
        grade=text_value(item.get("recommendation_level") or item.get("final_grade") or "C")
        decision=review_label(b) or "未分类"
        hits=item.get("positive_hits") or item.get("scope_hits") or item.get("product_hits") or []
        match_text="、".join(dict.fromkeys(text_value(x) for x in hits if text_value(x))) or proj_type(b) or "未形成明确业务命中"
        scope=project_scope(b)
        facts=[scope] if scope else evidence_snippets(b,1)
        attachments=project_attachments(b)
        if attachments:
            read_count=sum(1 for row in attachments if row.get("text_read_ok"))
            facts.append(f"页面识别{len(attachments)}个附件，已读取{read_count}个")
        title_scope=" ".join([
            text_value(item.get("title") or (RAWMAP.get(b) or {}).get("title")),
            scope,
        ])
        context=" ".join([title_scope," ".join(evidence_snippets(b,2))])
        domain_tokens=("光伏","充电桩","储能")
        boundary_tokens=("检测","监测","咨询","设计","维护","运维","维修","更换","改造")
        risks=[]
        def add_relevant_risk(value):
            for candidate in re.split(r"[；;]+",text_value(value)):
                candidate=text_value(candidate)
                mentioned_domains=[token for token in domain_tokens if token in candidate]
                mentioned_boundaries=[token for token in boundary_tokens if token in candidate]
                if mentioned_domains and not any(token in title_scope for token in mentioned_domains):
                    continue
                if mentioned_boundaries and not any(token in title_scope for token in mentioned_boundaries):
                    continue
                if candidate and candidate not in risks:
                    risks.append(candidate)
        for key in ("risk_points","ambiguous_hits"):
            value=item.get(key) or []
            if isinstance(value,str):
                value=[value]
            for candidate in value:
                add_relevant_risk(candidate)
        if any(token in scope for token in ("材料采购","设备采购","货物采购")) and not any(
            token in context for token in ("施工","安装","调试","工程承包")
        ):
            add_relevant_risk("项目以材料/设备供货为主，需确认是否接受纯供货边界")
        for key in ("evidence_gap_reason","manual_review_reason","attachment_invalid_reason"):
            add_relevant_risk(item.get(key))
        risks=list(dict.fromkeys(risks))
        action={
            "keep":"建议业务人员核对截止时间、投标资质和工作量后决定是否跟进。",
            "needs_review":"建议优先核对未闭合的正文或附件证据，再决定是否跟进。",
            "reject":"建议不跟进；如业务边界发生变化，再按原始链接复核。",
        }.get(dec(b),"建议人工确认。")
        return "\n".join([
            f"评分与结论：初筛{score_total(b)}分，{grade}级{decision}。",
            f"业务匹配：{match_text}。",
            "事实依据："+("；".join(facts) if facts else "正文未提取到足以支撑结论的具体事实。"),
            "风险/缺口："+("；".join(risks[:4]) if risks else "未发现需要额外说明的证据缺口。"),
            "建议动作："+action,
        ])

    def body_summary(b):
        rm=RAWMAP.get(b,{})
        item=rejudge.get(b) or {}
        attachments=project_attachments(b)
        scope=project_scope(b)
        if not scope:
            snippets=evidence_snippets(b,2)
            scope="；".join(snippets)
        amount=format_budget(rm.get("budget"),item.get("amount_wan"))
        timeline=f"发布日期{text_value(rm.get('publish')) or '未获取'}；截止日期{text_value(rm.get('bid_deadline')) or '未获取'}"
        qual=qualification(b)
        contact="；".join(x for x in [
            f"采购人：{text_value(rm.get('buyer'))}" if rm.get("buyer") else "",
            f"代理机构：{text_value(rm.get('agency'))}" if rm.get("agency") else "",
        ] if x)
        attachment_text=(
            f"页面共{len(attachments)}个附件；已读取{sum(1 for x in attachments if x.get('text_read_ok'))}个。"
            if attachments else "页面未识别到附件。"
        )
        lines=[
            "项目范围："+(scope or "正文未提取到明确采购/施工范围。"),
            "金额与工期："+f"预算{amount or '未获取'}；工期{duration(b) or '未获取'}。",
            "时间节点："+timeline+"。",
            "资质要求："+(qual or "未获取。"),
            "采购主体："+(contact or "未获取。"),
            "附件证据："+attachment_text,
        ]
        return "\n".join(lines)

    def exclude_reason(b):
        it=rejudge.get(b) or {}
        return (it.get("reason") or "；".join(it.get("exclude_hits") or [])) if dec(b)=="reject" else ""

    def review_reason(b):
        item=rejudge.get(b) or {}
        title_scope=" ".join([
            text_value(item.get("title") or (RAWMAP.get(b) or {}).get("title")),
            project_scope(b),
        ])
        domain_tokens=("\u5149\u4f0f","\u5145\u7535\u6869","\u50a8\u80fd")
        clauses=[]
        for source in (item.get("reason"),item.get("judgment_reason")):
            for clause in re.split("[\uFF1B;]+",text_value(source)):
                clause=text_value(clause)
                generic_pattern="\u6807\u9898\u542b[^()\uff08\uff09]{1,100}\u8bcd[\uff08(]([^\uff09)]+)[\uff09)]"
                clause=re.sub(generic_pattern,lambda match:"\u6807\u9898\u547d\u4e2d"+match.group(1),clause)
                mentioned=[token for token in domain_tokens if token in clause]
                if mentioned and not any(token in title_scope for token in mentioned):
                    continue
                if clause and clause not in clauses:
                    clauses.append(clause)
        if clauses:
            return "\uFF1B".join(clauses)
        if dec(b)=="keep":
            return "\u6b63\u6587/\u6807\u9898\u8bc1\u636e\u663e\u793a\u5c5e\u4e8e\u53ef\u505a\u4e1a\u52a1\u8303\u56f4"
        if dec(b)=="needs_review":
            for key in ("evidence_gap_reason","manual_review_reason","attachment_invalid_reason"):
                value=text_value(item.get(key))
                if value:
                    return value
            return "\u73b0\u6709\u6b63\u6587\u6216\u9644\u4ef6\u8bc1\u636e\u672a\u95ed\u5408\uff0c\u9700\u4eba\u5de5\u590d\u6838"
        return exclude_reason(b) or "\u73b0\u6709\u8bc1\u636e\u4e0d\u652f\u6301\u8fdb\u5165\u63a8\u8350\u5019\u9009"

    # ===== 样式 =====
    HEAD=PatternFill("solid",fgColor="4472C4"); HF=Font(name="微软雅黑",size=9,color="FFFFFF",bold=True)
    FONT=Font(name="微软雅黑",size=9,color="333333"); TITLEF=Font(name="微软雅黑",size=14,bold=True,color="1F4E79"); SECF=Font(name="微软雅黑",size=10,bold=True,color="1F4E79")
    Z0=PatternFill("solid",fgColor="FFFFFF"); Z1=PatternFill("solid",fgColor="F2F7FB")
    ORANGE=PatternFill("solid",fgColor="FCE4D6"); YELLOW=PatternFill("solid",fgColor="FFF2CC")
    THIN=Side(style="thin",color="D9D9D9"); BORD=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
    WRAP=Alignment(wrap_text=True,vertical="center")
    THICK=Side(style="medium",color="808080")
    def outer_border(ws,r1,c1,r2,c2):
        for c in range(c1,c2+1):
            ws.cell(r1,c).border=Border(top=THICK,left=ws.cell(r1,c).border.left,right=ws.cell(r1,c).border.right,bottom=ws.cell(r1,c).border.bottom)
            ws.cell(r2,c).border=Border(bottom=THICK,top=ws.cell(r2,c).border.top,left=ws.cell(r2,c).border.left,right=ws.cell(r2,c).border.right)
        for r in range(r1,r2+1):
            ws.cell(r,c1).border=Border(left=THICK,top=ws.cell(r,c1).border.top,right=ws.cell(r,c1).border.right,bottom=ws.cell(r,c1).border.bottom)
            ws.cell(r,c2).border=Border(right=THICK,top=ws.cell(r,c2).border.top,left=ws.cell(r,c2).border.left,bottom=ws.cell(r,c2).border.bottom)
    def style_header(ws,headers,widths):
        ws.sheet_view.showGridLines=False
        for j,h in enumerate(headers,1):
            c=ws.cell(1,j);c.value=h;c.fill=HEAD;c.font=HF;c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center");c.border=BORD
        for j,w in enumerate(widths,1):ws.column_dimensions[get_column_letter(j)].width=w
        ws.row_dimensions[1].height=30; ws.freeze_panes="A2"

    def ntype(b):
        rm=RAWMAP.get(b,{}); title=str(rm.get("title") or "")
        if any(k in title for k in ("中标","成交","结果公示","候选人公示")) or rm.get("win_amount"):
            return "中标通知"
        return "招标公告"

    wb=load_workbook(template)
    # ---------- 全部标讯 (28列) ----------
    name="📋 全部标讯"
    if name in wb.sheetnames: del wb[name]
    ws=wb.create_sheet(name,1)
    H_ALL=["序号","复核","评分明细","公告类型","公告标题","省份","城市","招标单位","项目类型","预算金额","发布日期","投标截止","招标编号","命中关键词","千里马链接","分析意见","排除原因","二筛决定","二筛原因","招标联系人","招标联系电话","招标代理机构","代理联系人","代理联系电话","中标单位","中标联系人","中标联系电话","评审专家"]
    W_ALL=[5,11,30,11,38,6,7,20,14,11,11,11,15,14,16,40,30,11,16,12,13,18,12,13,18,12,13,14]
    style_header(ws,H_ALL,W_ALL)
    # 数据：以原始导出全量为底，关联 agent
    order=list(RAWMAP.keys())
    ri=2
    for seq,b in enumerate(order,1):
        rm=RAWMAP[b]; reg=str(rm.get("region") or "广东-—"); prov,_,city=reg.partition("-")
        rl=review_label(b)
        vals=[seq,rl,score_detail(b),ntype(b),rm.get("title") or "",prov,city,
            rm.get("buyer") or "",proj_type(b),(format_budget(rm.get("budget")) or "—"),
            rm.get("publish") or "—",rm.get("bid_deadline") or "—",rm.get("bid_no") or "",rm.get("kw") or "",
            rm.get("url"),analysis(b),exclude_reason(b),"","",
            rm.get("buyer_contact") or "",rm.get("buyer_tel") or "",rm.get("agency") or "",rm.get("agency_contact") or "",rm.get("agency_tel") or "",
            rm.get("winner") or "",rm.get("win_contact") or "",rm.get("win_tel") or "",rm.get("expert") or ""]
        zfill=Z1 if seq%2==0 else Z0
        for j,v in enumerate(vals,1):
            c=ws.cell(ri,j);c.value=v;c.alignment=WRAP;c.border=BORD;c.fill=zfill;c.font=FONT
        # 复核底色
        if rl=="推荐候选":ws.cell(ri,2).fill=ORANGE
        elif rl=="待人工复核":ws.cell(ri,2).fill=YELLOW
        elif rl=="排除":ws.cell(ri,2).fill=PatternFill("solid",fgColor="EFEFEF")
        # 联系人列(20-28)有值时黄高亮
        for cc in range(20,29):
            if ws.cell(ri,cc).value:ws.cell(ri,cc).fill=YELLOW
        lk=ws.cell(ri,15)
        if lk.value:lk.hyperlink=lk.value;lk.font=Font(color="0563C1",underline="single")
        ws.row_dimensions[ri].height=40; ri+=1

    outer_border(ws,1,1,ri-1,len(H_ALL))
    # ---------- 今日推荐 (31列: +分类 +工期 +付款) ----------
    name="📍 今日推荐标讯"
    if name in wb.sheetnames: del wb[name]
    ws2=wb.create_sheet(name,2)
    H_T=["序号","分类","复核","评分明细","公告类型","公告标题","省份","城市","招标单位","项目类型","预算金额","发布日期","投标截止","工期","付款方式","招标编号","命中关键词","千里马链接","分析意见","排除原因","二筛决定","二筛原因","招标联系人","招标联系电话","招标代理机构","代理联系人","代理联系电话","中标单位","中标联系人","中标联系电话","评审专家"]
    W_T=[5,10,11,30,11,38,6,7,20,14,11,11,11,12,16,15,14,16,40,30,11,11,12,13,18,12,13,18,12,13,14]
    style_header(ws2,H_T,W_T)
    def lvl(b):return {"A":0,"B":1,"C":2}.get(str((rejudge.get(b) or {}).get("recommendation_level","C"))[:1],9)
    def tot(b):
        s=(screen.get(b) or {}).get("scores") or {};return s.get("总分",s.get("原始总分",0)) or 0
    keeps=sorted([b for b in rejudge if dec(b)=="keep"],key=lambda b:(lvl(b),-tot(b)))
    reviews=sorted([b for b in rejudge if dec(b)=="needs_review"],key=lambda b:(lvl(b),-tot(b)))
    ri=2; seq=1
    for grp,bs,fill in [("推荐候选",keeps,ORANGE),("人工复核",reviews,YELLOW)]:
        for b in bs:
            rm=RAWMAP.get(b,{}); reg=str(rm.get("region") or (rejudge.get(b) or {}).get("region") or "广东-—");prov,_,city=reg.partition("-")
            vals=[seq,grp,review_label(b),score_detail(b),ntype(b),rm.get("title") or (rejudge.get(b) or {}).get("title") or "",prov,city,
                rm.get("buyer") or "",proj_type(b),(format_budget(rm.get("budget")) or "—"),
                rm.get("publish") or "—",rm.get("bid_deadline") or "—",duration(b),payment(b),rm.get("bid_no") or "",rm.get("kw") or "",
                rm.get("url"),analysis(b),exclude_reason(b),"","",
                rm.get("buyer_contact") or "",rm.get("buyer_tel") or "",rm.get("agency") or "",rm.get("agency_contact") or "",rm.get("agency_tel") or "",
                rm.get("winner") or "",rm.get("win_contact") or "",rm.get("win_tel") or "",rm.get("expert") or ""]
            for j,v in enumerate(vals,1):
                c=ws2.cell(ri,j);c.value=v;c.alignment=WRAP;c.border=BORD;c.fill=fill;c.font=FONT
            lk=ws2.cell(ri,18)
            if lk.value:lk.hyperlink=lk.value;lk.font=Font(color="0563C1",underline="single")
            ws2.row_dimensions[ri].height=44; ri+=1; seq+=1

    outer_border(ws2,1,1,ri-1,len(H_T))
    # ---------- 统计 概览+两张统计表（按截图，范围=推荐+待复核）----------
    st=wb["📊 统计"]; st.sheet_view.showGridLines=False
    for _m in list(st.merged_cells.ranges): st.unmerge_cells(str(_m))
    for row in st.iter_rows(min_row=1,max_row=st.max_row):
        for c in row:
            try: c.value=None; c.fill=PatternFill(); c.border=Border()
            except: pass
    BAND=PatternFill("solid",fgColor="DDEBF7"); SEC=Font(name="微软雅黑",size=10,bold=True,color="1F4E79")
    HFC=Font(name="微软雅黑",size=9,bold=True,color="FFFFFF"); CEN=Alignment(horizontal="center",vertical="center")
    def band(r,txt):
        st.merge_cells(f"A{r}:H{r}")
        c=st.cell(r,1); c.value=txt; c.fill=BAND; c.font=SEC; c.alignment=Alignment(horizontal="left",vertical="center"); st.row_dimensions[r].height=22
    def hdr(r,labels,fills=None):
        for j,t in enumerate(labels,1):
            c=st.cell(r,j); c.value=t; c.fill=PatternFill("solid",fgColor=(fills[j-1] if fills else "4472C4")); c.font=HFC; c.alignment=CEN; c.border=BORD
    def CAT(b):
        t=(proj_type(b)+" "+((rejudge.get(b) or {}).get("doable_scope") or "")+" "+(RAWMAP.get(b,{}).get("title") or ""))
        if "亮化" in t or "景观" in t: return "亮化"
        if "路灯" in t or "道路照明" in t: return "路灯"
        if "光伏" in t: return "光伏"
        if "充电" in t: return "充电桩"
        if "储能" in t: return "储能"
        if "照明" in t: return "照明"
        return "其他"
    def cat4(b):
        c=CAT(b); return {"路灯":"路灯工程","照明":"建筑照明","亮化":"景观亮化","光伏":"光伏工程"}.get(c,"其他")
    actionable=keeps+reviews
    # 标题
    st.merge_cells("A1:H1"); st["A1"]="招投标日报 · 统计汇总"; st["A1"].font=TITLEF; st["A1"].alignment=Alignment(horizontal="center",vertical="center"); st.row_dimensions[1].height=30
    st.merge_cells("A2:B2"); st["A2"]="📅 生成日期"; st["A2"].font=Font(name="微软雅黑",size=9,bold=True)
    st.merge_cells("C2:D2"); st["C2"]=run_date; st["C2"].font=FONT
    st["E2"]="🔍 筛选范围"; st["E2"].font=Font(name="微软雅黑",size=9,bold=True)
    st.merge_cells("F2:H2"); st["F2"]=f"广东 · 千里马招标网 · {src_start} 至 {src_end}"; st["F2"].font=FONT
    # ① 今日概览
    band(4,"① 今日概览")
    ov=[("今日标讯","4472C4",len(RAWMAP)),("推荐候选","ED7D31",len(keeps)),("待人工复核","FFC000",len(reviews)),("排除","A6A6A6",sum(1 for x in RAWMAP if dec(x)=="reject"))]
    for j,(lab,col,val) in enumerate(ov,1):
        c=st.cell(5,j); c.value=lab; c.fill=PatternFill("solid",fgColor=col); c.font=HFC; c.alignment=CEN; c.border=BORD
        v=st.cell(6,j); v.value=val; v.font=Font(name="微软雅黑",size=12,bold=True); v.alignment=CEN; v.border=BORD
    for j in range(5,9):
        st.cell(5,j).fill=PatternFill("solid",fgColor="4472C4"); st.cell(5,j).border=BORD; st.cell(6,j).border=BORD
    # ② 按地区与项目类型（推荐+待复核）
    band(8,"📍 （推荐标讯和待人工复核）按地区与项目类型统计")
    geo_cols=["城市","标讯数","路灯工程","建筑照明","景观亮化","光伏工程","其他","中标通知"]
    hdr(9,geo_cols)
    cities=OrderedDict()
    for b in actionable:
        city=(RAWMAP.get(b,{}).get("region") or "—").split("-")[-1]
        cities.setdefault(city,[]).append(b)
    r=10
    for city,bs in cities.items():
        cc=Counter(cat4(x) for x in bs); zhongbiao=sum(1 for x in bs if "中标" in ntype(x))
        vals=[city,len(bs),cc.get("路灯工程",0),cc.get("建筑照明",0),cc.get("景观亮化",0),cc.get("光伏工程",0),cc.get("其他",0),zhongbiao]
        for j,v in enumerate(vals,1):
            c=st.cell(r,j); c.value=v; c.font=FONT; c.alignment=CEN; c.border=BORD; c.fill=(Z1 if (r-10)%2 else Z0)
        r+=1
    geo_end=r-1
    # ③ 按公告类型（推荐+待复核）
    tr=r+1
    band(tr,"📊 （推荐标讯和待人工复核）按公告类型统计")
    hdr(tr+1,["公告类型","数量","占比","路灯工程","建筑照明","景观亮化","光伏工程","其他"])
    nt=OrderedDict()
    for b in actionable: nt.setdefault(ntype(b),[]).append(b)
    rr=tr+2; total=len(actionable)
    for typ,bs in nt.items():
        cc=Counter(cat4(x) for x in bs)
        vals=[typ,len(bs),f"{len(bs)/total*100:.1f}%" if total else "0%",cc.get("路灯工程",0),cc.get("建筑照明",0),cc.get("景观亮化",0),cc.get("光伏工程",0),cc.get("其他",0)]
        for j,v in enumerate(vals,1):
            c=st.cell(rr,j); c.value=v; c.font=FONT; c.alignment=CEN; c.border=BORD; c.fill=(Z1 if (rr-(tr+2))%2 else Z0)
        rr+=1
    # 合计
    cc=Counter(cat4(x) for x in actionable)
    tot=["合计",total,"100%",cc.get("路灯工程",0),cc.get("建筑照明",0),cc.get("景观亮化",0),cc.get("光伏工程",0),cc.get("其他",0)]
    for j,v in enumerate(tot,1):
        c=st.cell(rr,j); c.value=v; c.font=Font(name="微软雅黑",size=9,bold=True); c.alignment=CEN; c.border=BORD; c.fill=PatternFill("solid",fgColor="DDEBF7")
    # ④ 筛选规则
    g=rr+2
    band(g,"⚙️ 筛选规则 & 三级分类")
    rules=[("正选词","路灯、照明、亮化、LED、光伏、储能、充电桩"),("排除词","检验监测、环评、纯设备、租赁、设计、监理、检测、中标结果"),("三级分类","推荐候选 / 待人工复核 / 排除")]
    for i,(k,v) in enumerate(rules):
        rk=g+1+i; st.cell(rk,1).value=k; st.cell(rk,1).font=Font(name="微软雅黑",size=9,bold=True); st.cell(rk,1).border=BORD
        st.merge_cells(f"B{rk}:H{rk}"); st.cell(rk,2).value=v; st.cell(rk,2).font=FONT; st.cell(rk,2).alignment=Alignment(vertical="center"); st.cell(rk,2).border=BORD
    # ⑤ 二筛状态
    e=g+1+len(rules)+1
    band(e,"🔬 二筛状态")
    n_pending=sum(1 for b in actionable if True)  # 二筛决定空=全部待定
    sec=[("✅ 重点跟进","E2EFDA",0),("⏸️ 待定","FFF2CC",n_pending),("❌ 放弃投标","FCE4D6",0)]
    for i,(lab,col,val) in enumerate(sec):
        rk=e+1+i
        st.merge_cells(f"A{rk}:D{rk}"); c=st.cell(rk,1); c.value=lab; c.fill=PatternFill("solid",fgColor=col); c.font=Font(name="微软雅黑",size=9,bold=True); c.alignment=CEN; c.border=BORD
        v=st.cell(rk,5); v.value=val; v.font=Font(name="微软雅黑",size=12,bold=True,color="C00000"); v.alignment=CEN; v.border=BORD
    # 列宽
    st.column_dimensions["A"].width=16
    for col in ["B","C","D","E","F","G","H"]: st.column_dimensions[col].width=11

    # ---------- 项目详情：全部字段必须覆盖模板样例 ----------
    proj=[w for w in wb.worksheets if w.title.startswith("项目")]; tp=proj[-1]
    for w in list(proj):
        if w is not tp: del wb[w.title]
    tp.title="项目1"; detail=keeps+reviews; sheets=[tp]
    for i in range(2,max(1,len(detail))+1):
        sheet=wb.copy_worksheet(tp);sheet.title=f"项目{i}";sheets.append(sheet)
    for sheet in sheets: sheet.sheet_view.showGridLines=False

    def find_row(sheet,label):
        for row in range(1,sheet.max_row+1):
            value=text_value(sheet.cell(row,1).value)
            if value==label or value.startswith(label):
                return row
        return 0

    def copy_row_format(sheet,source,target,max_col=5):
        sheet.row_dimensions[target].height=sheet.row_dimensions[source].height
        for col in range(1,max_col+1):
            src=sheet.cell(source,col);dst=sheet.cell(target,col)
            if src.has_style: dst._style=copy(src._style)
            dst.number_format=src.number_format
            dst.alignment=copy(src.alignment)

    def insert_rows_preserving_merges(sheet,index,amount):
        if amount<=0:return
        moved=[]
        for merged in list(sheet.merged_cells.ranges):
            if merged.min_row>=index:
                moved.append((merged.min_row+amount,merged.min_col,merged.max_row+amount,merged.max_col))
                sheet.unmerge_cells(str(merged))
            elif merged.max_row>=index:
                raise ValueError(f"Cannot insert through merged range {merged}")
        sheet.insert_rows(index,amount)
        for r1,c1,r2,c2 in moved:
            sheet.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

    def clear_row_values(sheet,row,max_col=5):
        for col in range(1,max_col+1):
            cell=sheet.cell(row,col)
            if isinstance(cell,MergedCell):
                continue
            cell.value=None;cell.hyperlink=None

    for idx,b in enumerate(detail):
        w=sheets[idx];rm=RAWMAP.get(b,{})
        w["A1"]=f"{review_label(b)} - {rm.get('title') or (rejudge.get(b) or {}).get('title') or ''}"
        w["B4"]=ntype(b)
        w["B5"]=text_value(rm.get("region")) or text_value((rejudge.get(b) or {}).get("region")) or "未获取"
        w["B6"]=text_value(rm.get("publish")) or "未获取"
        w["B7"]=text_value(rm.get("bid_deadline")) or "未获取"
        amount=format_budget(rm.get("budget"),(rejudge.get(b) or {}).get("amount_wan"))
        w["B8"]=amount or "未获取"
        w["B9"]=text_value(rm.get("bid_no")) or "未获取"
        w["B10"]=f"{ntype(b)} / {proj_type(b) or '待确认'}"
        w["B11"]=score_detail(b) or f"总分{score_total(b)}"
        hits=(rejudge.get(b) or {}).get("positive_hits") or (rejudge.get(b) or {}).get("scope_hits") or []
        w["B12"]="、".join(dict.fromkeys(text_value(x) for x in hits if text_value(x))) or text_value(rm.get("kw")) or "未获取"
        w["B13"]=funding(b) or "未获取"
        w["B14"]=duration(b) or "未获取"
        w["B15"]=qualification(b) or "未获取"
        w["A17"]=rm.get("url") or (rejudge.get(b) or {}).get("url") or ""
        if w["A17"].value:
            w["A17"].hyperlink=w["A17"].value;w["A17"].font=Font(color="0563C1",underline="single")
        w["A20"]=analysis(b);w["A20"].alignment=Alignment(wrap_text=True,vertical="top")
        w.row_dimensions[20].height=120
        w["C23"]=review_label(b)
        w["C24"]=review_reason(b)

        events=progress_timeline(b)
        timeline_header=find_row(w,"📅 项目进度跟踪")
        timeline_start=timeline_header+2
        body_header=find_row(w,"📄 招标正文摘要")
        capacity=max(1,body_header-timeline_start)
        if len(events)>capacity:
            insert_rows_preserving_merges(w,body_header,len(events)-capacity)
            for row in range(body_header,body_header+len(events)-capacity):
                copy_row_format(w,timeline_start,row)
        body_header=find_row(w,"📄 招标正文摘要")
        for row in range(timeline_start,body_header):
            clear_row_values(w,row)
        for offset,event in enumerate(events):
            row=timeline_start+offset
            w.cell(row,1).value=event.get("date") or "未获取"
            w.cell(row,2).value=event.get("node") or "未获取"
            w.cell(row,3).value=event.get("title") or "未获取"
            for col in range(1,4): w.cell(row,col).alignment=Alignment(wrap_text=True,vertical="center")
            w.row_dimensions[row].height=34

        body_header=find_row(w,"📄 招标正文摘要")
        body_row=body_header+1
        w.cell(body_row,1).value=body_summary(b)
        w.cell(body_row,1).alignment=Alignment(wrap_text=True,vertical="top")
        w.row_dimensions[body_row].height=150
        if body_row+1<=w.max_row: w.row_dimensions[body_row+1].height=18

        attachment_header=find_row(w,"📎 附件下载")
        attachment_row=attachment_header+1
        attachment_rows=project_attachments(b)
        if len(attachment_rows)>1:
            insert_rows_preserving_merges(w,attachment_row+1,len(attachment_rows)-1)
            for row in range(attachment_row+1,attachment_row+len(attachment_rows)):
                copy_row_format(w,attachment_row,row)
                w.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
                w.merge_cells(start_row=row,start_column=3,end_row=row,end_column=5)
        clear_row_values(w,attachment_row)
        if not attachment_rows:
            w.cell(attachment_row,1).value="无附件"
            w.cell(attachment_row,3).value=""
        else:
            for offset,item in enumerate(attachment_rows):
                row=attachment_row+offset
                w.cell(row,1).value=item["name"]
                w.cell(row,3).value=item["url"]
                if item["url"]:
                    w.cell(row,3).hyperlink=item["url"]
                    w.cell(row,3).font=Font(color="0563C1",underline="single")
                for col in (1,3): w.cell(row,col).alignment=Alignment(wrap_text=True,vertical="center")
                w.row_dimensions[row].height=34

        contact_header=find_row(w,"📞 联系方式")
        if contact_header:
            w.cell(contact_header+1,2).value=text_value(rm.get("buyer_contact")) or "未获取"
            w.cell(contact_header+2,2).value=text_value(rm.get("buyer_tel")) or "未获取"
        
        # 插入正文截图（在联系方式下方）
        screenshot_path = vip.get(b, {}).get("screenshot_path", "")
        if screenshot_path and contact_header:
            from openpyxl.drawing.image import Image as XLImage
            import os
            # 构建完整路径
            full_screenshot_path = os.path.join(os.path.dirname(vip_json), screenshot_path)
            if os.path.exists(full_screenshot_path):
                try:
                    img = XLImage(full_screenshot_path)
                    # 调整图片大小适配单元格宽度
                    img.width = 800  # 固定宽度
                    img.height = 600  # 自适应高度
                    # 插入位置：联系方式下方第二行
                    insert_row = contact_header + 3
                    cell_ref = f"A{insert_row}"
                    w.add_image(img, cell_ref)
                    # 调整行高以适配图片
                    w.row_dimensions[insert_row].height = 450
                except Exception as e:
                    print(f"  ⚠️ 插入截图失败: {e}")

    if not detail:
        # 全排除场景：清空模板示例数据但保留 sheet 结构
        tp["A1"]="今日无推荐候选或待人工复核项目"
        for row in range(4,tp.max_row+1):
            for col in range(1,min(tp.max_column,5)+1):
                if row not in (3,16,19,22,26,32,35,38):
                    c = tp.cell(row,col)
                    if not isinstance(c, MergedCell):
                        c.value = None

    wb.save(out)
    return out

