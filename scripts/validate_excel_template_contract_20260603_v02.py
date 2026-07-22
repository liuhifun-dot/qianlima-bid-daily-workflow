import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def merged_ranges(ws):
    return sorted(str(rng) for rng in ws.merged_cells.ranges)


def cell_value(ws, addr):
    value = ws[addr].value
    return "" if value is None else str(value).strip()


def normalize_decision_for_contract(value):
    value = str(value or "").strip()
    if value in {"keep", "推荐", "推荐候选", "推荐项目", "重点推荐"}:
        return "keep"
    if value in {"needs_review", "待人工复核", "待复核", "待定", "人工复核"}:
        return "needs_review"
    if value in {"reject", "排除", "不推荐", "放弃"}:
        return "reject"
    return value


def flatten_business_rows(data):
    if isinstance(data, list):
        return data
    if not isinstance(data, dict):
        return []
    if isinstance(data.get("results"), list):
        return data["results"]
    rows = []
    for key in ("recommended", "considered", "low_score", "excluded"):
        rows.extend(data.get(key) or [])
    return rows


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(
        description="Validate tender daily report workbook against the fixed Excel template contract."
    )
    parser.add_argument("--template", required=True, help="Template xlsx path")
    parser.add_argument("--workbook", required=True, help="Generated report xlsx path")
    parser.add_argument("--business-json", help="Phase 2 final business rejudge JSON path")
    args = parser.parse_args()

    template_path = Path(args.template)
    workbook_path = Path(args.workbook)
    twb = load_workbook(template_path, data_only=False)
    wwb = load_workbook(workbook_path, data_only=False)

    errors = []
    warnings = []

    business_rows = []
    business_by_title = {}
    if args.business_json:
        business_path = Path(args.business_json)
        if business_path.exists():
            business_payload = json.loads(business_path.read_text(encoding="utf-8-sig"))
            business_rows = flatten_business_rows(business_payload)
            business_by_title = {
                re.sub(r"\s+", "", str(item.get("title") or "")): item
                for item in business_rows if item.get("title")
            }

    expected_sheets_prefix = [
        "\U0001f4ca \u7edf\u8ba1",
        "\U0001f4cb \u5168\u90e8\u6807\u8baf",
        "\U0001f4cd \u4eca\u65e5\u63a8\u8350\u6807\u8baf",
    ]
    for idx, expected in enumerate(expected_sheets_prefix):
        if len(wwb.worksheets) <= idx:
            errors.append(f"Missing sheet index {idx}: {expected}")
        elif wwb.worksheets[idx].title != expected:
            errors.append(
                f"Sheet {idx + 1} name mismatch: expected {expected}, got {wwb.worksheets[idx].title}"
            )

    # The last project sheet in the template is the canonical project-detail form.
    project_template = twb.worksheets[-1]
    expected_project_merges = merged_ranges(project_template)
    project_sheets = [ws for ws in wwb.worksheets if ws.title.startswith("\u9879\u76ee")]
    # 全排除场景：business_rows 全为 reject 时，允许 0 个项目 sheet
    all_reject = bool(business_rows) and all(
        normalize_decision_for_contract(item.get("decision") or item.get("final_decision")) == "reject"
        for item in business_rows
    )
    if not project_sheets and not all_reject:
        errors.append("No project detail sheets found. Expected sheets named 项目1/项目2/...")
    elif not project_sheets and all_reject:
        warnings.append("全排除场景：无推荐候选或待人工复核项目，项目 sheet 为空属正常。")

    fixed_labels = {
        "A3": "\U0001f4cc \u57fa\u672c\u4fe1\u606f",
        "A4": "\u516c\u544a\u7c7b\u578b",
        "A5": "\u7701\u4efd/\u57ce\u5e02",
        "A6": "\u53d1\u5e03\u65e5\u671f",
        "A7": "\u622a\u6b62\u65e5\u671f",
        "A8": "\u9884\u7b97\u91d1\u989d",
        "A9": "\u62db\u6807\u7f16\u53f7",
        "A10": "\u6240\u5c5e\u884c\u4e1a",
        "A11": "\u8bc4\u5206",
        "A12": "\u547d\u4e2d\u5173\u952e\u8bcd",
        "A13": "\u8d44\u91d1\u6765\u6e90",
        "A14": "\u5de5\u671f",
        "A15": "\u8d44\u8d28\u8981\u6c42",
        "A16": "\U0001f517 \u5343\u91cc\u9a6c\u62db\u6807\u7f51\u94fe\u63a5\uff08\u53ef\u70b9\u51fb\u590d\u5236\uff09",
        "A19": "\U0001f4a1 \u6807\u795e\u5206\u6790\u610f\u89c1",
        "A22": "\U0001f52c \u4e8c\u7b5b\u51b3\u5b9a",
        "A23": "\u4e8c\u7b5b\u51b3\u5b9a",
        "A24": "\u539f\u56e0",
        "A26": "\U0001f4c5 \u9879\u76ee\u8fdb\u5ea6\u8ddf\u8e2a",
        "A27": "\u65e5\u671f",
        "B27": "\u8282\u70b9",
        "C27": "\u6807\u9898",
        "A32": "\U0001f4c4 \u62db\u6807\u6b63\u6587\u6458\u8981\uff08\u5343\u91cc\u9a6cVIP\u539f\u6587\uff09",
        "A35": "\U0001f4ce \u9644\u4ef6\u4e0b\u8f7d",
        "A38": "\U0001f4de \u8054\u7cfb\u65b9\u5f0f",
        "A39": "\u8054\u7cfb\u4eba",
        "A40": "\u8054\u7cfb\u7535\u8bdd",
    }

    for ws in project_sheets:
        # 全排除场景：跳过项目 sheet 的字段校验（空 sheet 属正常）
        if all_reject:
            warnings.append(f"{ws.title}: 全排除场景，跳过字段校验。")
            continue
        # 20260615：项目表支持正文动态插行(行号/合并会变)，标签改为全表查找
        sheet_texts = set()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6):
            for c in row:
                if c.value not in (None, ""):
                    sheet_texts.add(str(c.value).strip())
        for expected in fixed_labels.values():
            if expected not in sheet_texts:
                errors.append(f"{ws.title}: 缺少必备标签 {expected!r}")

        # Common wrong mapping: putting second-screening decision into B12/B13/B14/B15.
        for addr in ["B12", "B13", "B14", "B15"]:
            val = cell_value(ws, addr)
            if any(token in val for token in ["\u786e\u8ba4\u8ddf\u8fdb", "\u786e\u8ba4\u6295\u6807", "\u5f85\u4eba\u5de5\u590d\u6838", "\u653e\u5f03\u6295\u6807", "\u4e8c\u7b5b"]):
                errors.append(
                    f"{ws.title}!{addr}: appears to contain screening decision text, "
                    "but B12:B15 must remain keyword/funding/duration/qualification fields."
                )

        link = cell_value(ws, "A17")
        if link and "qianlima.com" in link and not ws["A17"].hyperlink:
            warnings.append(f"{ws.title}!A17: qianlima URL is text only; make it a clickable hyperlink.")

        def find_label_row(label):
            for row_number in range(1, ws.max_row + 1):
                value = cell_value(ws, f"A{row_number}")
                if value == label or value.startswith(label):
                    return row_number
            return 0

        project_title = cell_value(ws, "A1").split(" - ", 1)[-1].strip()
        item = business_by_title.get(re.sub(r"\s+", "", project_title), {})
        all_text = "\n".join(
            str(cell.value) for row in ws.iter_rows() for cell in row
            if cell.value not in (None, "")
        )
        budget = cell_value(ws, "B8")
        unavailable = "\u672a\u83b7\u53d6"
        wan_yuan = "\u4e07\u5143"
        budget_pattern = r"\d+(?:\.\d{1,2})?" + wan_yuan
        if budget and budget != unavailable and not re.fullmatch(budget_pattern, budget):
            errors.append(f"{ws.title}: invalid project budget unit {budget!r}")

        sample_tokens = [
            "example-zbtzs.pdf", "中标通知书.pdf", "东莞松山湖景观亮化工程",
            "0769-12345678", "王主任", "bid-594305346.html",
        ]
        for token in sample_tokens:
            if token in all_text:
                errors.append(f"{ws.title}: template sample leaked into generated project sheet: {token}")

        analysis_row = find_label_row("💡 标神分析意见")
        analysis = cell_value(ws, f"A{analysis_row + 1}") if analysis_row else ""
        if len(analysis) > 700:
            errors.append(f"{ws.title}: analysis exceeds 700 characters")
        for required in ("评分与结论：", "业务匹配：", "事实依据：", "风险/缺口：", "建议动作："):
            if required not in analysis:
                errors.append(f"{ws.title}: analysis is missing required section {required}")

        body_header = find_label_row("📄 招标正文摘要")
        body_summary = cell_value(ws, f"A{body_header + 1}") if body_header else ""
        if len(body_summary) > 900:
            errors.append(f"{ws.title}: body summary exceeds 900 characters")
        for required in ("项目范围：", "金额与工期：", "时间节点：", "资质要求：", "采购主体：", "附件证据："):
            if required not in body_summary:
                errors.append(f"{ws.title}: body summary is missing required fact section {required}")
        for noise in (
            "摘要信息 招标详情", "商机推荐", "用手机查看此详情", "立即查看", "立即引荐",
            "历史招中标信息", "标书代写", "拟采购的材料设备清单", "序号 材料设备", "报价相关要求",
        ):
            if noise in body_summary:
                errors.append(f"{ws.title}: body summary contains page noise {noise!r}")

        decision_reason = cell_value(ws, "C24")
        reason_context = project_title + " " + analysis + " " + body_summary
        # 判断说明语境词：领域词若作为“边界/维护/反例/不排除”等判断话术出现，属合法业务复判说明，
        # 不是把项目误定性成该领域，降级为 warning 放行；否则仍按 error 处理（防幻觉泄漏）。
        judgement_context_markers = (
            "\u8fb9\u754c",       # 边界
            "\u7ef4\u62a4",       # 维护
            "\u4e0d\u80fd\u56e0", # 不能因
            "\u4e0d\u662f\u7eaf", # 不是纯
            "\u662f\u5426\u4e3a", # 是否为
            "\u76f4\u63a5\u6392\u9664", # 直接排除
            "\u5426\u5219",       # 否则
            "\u5982\u679c",       # 如果
        )
        for token in ("\u5149\u4f0f", "\u5145\u7535\u6869", "\u50a8\u80fd"):
            if token in decision_reason and token not in reason_context:
                if any(marker in decision_reason for marker in judgement_context_markers):
                    warnings.append(
                        f"{ws.title}: second-screening reason mentions domain token {token!r} "
                        f"as judgement context (not project domain); pass with review."
                    )
                else:
                    errors.append(
                        f"{ws.title}: second-screening reason leaks unrelated domain token {token!r}"
                    )

        timeline_header = find_label_row("📅 项目进度跟踪")
        if timeline_header and body_header:
            timeline_rows = []
            for row_number in range(timeline_header + 2, body_header):
                date = cell_value(ws, f"A{row_number}")
                node = cell_value(ws, f"B{row_number}")
                title = cell_value(ws, f"C{row_number}")
                if date or node or title:
                    timeline_rows.append((date, node, title))
            if not timeline_rows:
                errors.append(f"{ws.title}: project timeline is empty")
            allowed_nodes = {
                "招标公告", "采购公告", "询价公告", "竞争性磋商公告", "招标预告",
                "变更公告", "答疑公告", "候选人公示", "中标通知", "中标结果", "成交公告", "合同公告",
            }
            for date, node, title in timeline_rows:
                if node not in allowed_nodes:
                    errors.append(f"{ws.title}: invalid timeline node {node!r}")
                if not date or not title:
                    errors.append(f"{ws.title}: incomplete timeline row date={date!r}, title={title!r}")

        attachment_header = find_label_row("📎 附件下载")
        contact_header = find_label_row("📞 联系方式")
        if attachment_header and contact_header:
            actual_attachments = []
            for row_number in range(attachment_header + 1, contact_header):
                name = cell_value(ws, f"A{row_number}")
                url = cell_value(ws, f"C{row_number}")
                if name and name != "无附件":
                    actual_attachments.append((name, url, bool(ws[f"C{row_number}"].hyperlink)))
            expected = []
            seen = set()
            for row in item.get("page_attachments") or []:
                name = str(row.get("name") or "附件").strip()
                url = str(row.get("source_url") or row.get("real_url") or "").strip()
                key = str(row.get("real_url") or row.get("source_url") or name).strip()
                # download_ok 默认 True：旧 JSON 无字段时仍按硬校验
                dl_ok = bool(row.get("download_ok", True))
                if key and key not in seen:
                    seen.add(key)
                    expected.append((name, url, dl_ok))
            if len(actual_attachments) != len(expected):
                errors.append(
                    f"{ws.title}: attachment row count mismatch; expected {len(expected)}, got {len(actual_attachments)}"
                )
            for index, expected_row in enumerate(expected):
                if index >= len(actual_attachments):
                    break
                actual_name, actual_url, clickable = actual_attachments[index]
                expected_name, expected_url, expected_download_ok = expected_row
                if actual_name != expected_name:
                    errors.append(
                        f"{ws.title}: attachment {index + 1} name mismatch; expected {expected_name!r}, got {actual_name!r}"
                    )
                # download_ok=False（如 zip 容器校验失败）时 URL 问题降为 warning，不阻断日报；
                # 若业务侧已带回 source_url，仍希望 Excel 尽量可点，不匹配仅 warning。
                if not expected_download_ok:
                    if not clickable or actual_url != expected_url:
                        warnings.append(
                            f"{ws.title}: attachment {index + 1} URL not clickable "
                            f"(download_ok=False, acceptable)"
                        )
                elif actual_url != expected_url or not clickable:
                    errors.append(
                        f"{ws.title}: attachment {index + 1} must keep the real clickable Qianlima/source URL"
                    )

    # === #6 新增：统计一致性 / 项目页编号连续 / 原始导出表存在（2026-06-11） ===
    # #6d 原始导出全部数据表存在
    if not any(ws.title.startswith("\U0001f4e6") or "原始导出" in ws.title for ws in wwb.worksheets):
        warnings.append("无独立原始导出表；新版『全部标讯』已含千里马原始全量列，视为通过。")

    # #6b 项目页编号连续：项目1/项目2/...
    pnums = sorted(int(m.group(1)) for ws in project_sheets for m in [re.fullmatch(r"项目(\d+)", ws.title)] if m)
    if pnums and pnums != list(range(1, len(pnums) + 1)):
        errors.append(f"项目页编号不连续：{['项目'+str(n) for n in pnums]}；应为 项目1..项目{len(pnums)}。")

    # #6a 统计页占比与计数一致（抓“推荐=0 却显示百分比”等写入错误）
    stats_ws = next((ws for ws in wwb.worksheets if ws.title.startswith("\U0001f4ca")), None)
    if stats_ws is not None:
        label_row = {}
        for row in stats_ws.iter_rows(min_col=1, max_col=1):
            t = "" if row[0].value is None else str(row[0].value).strip()
            if t in ("总复判", "推荐候选", "待人工复核", "排除"):
                label_row[t] = row[0].row
        def _f(v):
            try:
                return float(str(v).replace("%", "").strip())
            except Exception:
                return None
        tr = label_row.get("总复判")
        total = _f(stats_ws.cell(tr, 2).value) if tr else None
        if total and total > 0:
            cat_sum = 0.0
            for label in ("推荐候选", "待人工复核", "排除"):
                r = label_row.get(label)
                if not r:
                    continue
                cnt = _f(stats_ws.cell(r, 2).value)
                pct = _f(stats_ws.cell(r, 3).value)
                if cnt is None:
                    continue
                cat_sum += cnt
                if pct is not None:
                    exp = round(cnt / total * 100, 1)
                    if cnt == 0 and pct != 0:
                        errors.append(f"统计页『{label}』计数为 0，占比却为 {pct}%。")
                    elif abs(pct - exp) > 0.6:
                        errors.append(f"统计页『{label}』占比 {pct}% 与计数 {int(cnt)}/{int(total)}（应 {exp}%）不一致。")
            if abs(cat_sum - total) > 0.5:
                errors.append(f"统计页计数不闭合：推荐+待复核+排除={int(cat_sum)} ≠ 总复判 {int(total)}。")



    # 20260616 Phase3 hard gates: business wording and notice-type contract.
    forbidden_tokens = ["确认投标", "确认跟进"]
    for ws in wwb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = "" if cell.value is None else str(cell.value)
                for token in forbidden_tokens:
                    if token in value:
                        errors.append(f"{ws.title}!{cell.coordinate}: contains forbidden wording {token!r}.")

    allowed_notice_types = {"招标公告", "中标通知", "合计"}
    for ws in wwb.worksheets:
        for row in ws.iter_rows(min_col=1, max_col=1):
            label = "" if row[0].value is None else str(row[0].value).strip()
            if "按公告类型统计" not in label:
                continue
            header_row = row[0].row + 1
            if cell_value(ws, f"A{header_row}") != "公告类型":
                errors.append(f"{ws.title}!A{header_row}: expected 公告类型 header below notice-type section.")
                continue
            data_row = header_row + 1
            while data_row <= ws.max_row:
                typ = cell_value(ws, f"A{data_row}")
                if not typ:
                    break
                if typ not in allowed_notice_types:
                    errors.append(
                        f"{ws.title}!A{data_row}: invalid 公告类型 {typ!r}; allowed only 招标公告/中标通知/合计."
                    )
                if typ == "合计":
                    break
                data_row += 1

    result = {
        "ok": not errors,
        "template": str(template_path),
        "workbook": str(workbook_path),
        "project_sheet_count": len(project_sheets),
        "errors": errors,
        "warnings": warnings,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if not errors else 2


if __name__ == "__main__":
    raise SystemExit(main())
