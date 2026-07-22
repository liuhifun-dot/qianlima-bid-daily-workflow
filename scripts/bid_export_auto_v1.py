# -*- coding: utf-8 -*-
r"""
千里马 VIP 标讯主线导出 - 版本化修复副本

版本: 20260623_v05（正式运行版，Phase1可恢复状态机修复）
用途: 复用已登录的 CDP Chrome，将千里马 VIP 标讯订阅页设置为指定发布日期范围，
      执行批量导出，并从“我的导出记录”下载 xlsx 文件。

20260715_v08c 修改（用户确认的完整状态序列 + 150 条窗口语义）:
  完整序列：校验中(前置/先刷新) → 待生成 → 生成中 → 校验中 → 已生成(→下载)。
  首见「校验中/刷新」：先点目标行「刷新」转入待生成；再点一次「生成」。
  额度弹窗若出现：点「继续导出」（本账号多数不弹）；剩余 0 也不得 fail-fast 停链。
  生成中/校验中：每 10–20 秒点目标行「刷新」重读；只操作目标行，不点历史行下载。
  150 = 导出文件条数上限/滚动窗口（筛选 400 也只落盘约 150），不是「不能生成」。

20260715_v08b 修改（下载认文件加固，防 Downloads 噪声假成功）:
  必须 ctx.download_clicked 后才认新 xlsx；用户 Downloads 需文件名像千里马导出；
  validate_xlsx 校验表头含「标题」+ 基本信息/发布时间等；拒绝 structured_data 等噪声。
  pipeline：Phase1 仅登录类失败才 login_retry。

20260715_v08 修改（导出记录「哲理」：状态栏理解 + 操作栏照做）:
  「刷新」= 目标行操作列 cu-btn「刷新」，不是 Page.reload。
  只锁本次导出目标行，不点历史行下载。

20260617_v04 修改（Phase 1 无人值守下载状态机修复）:
  1. 导出记录页动作改为 DOM 行内点击，避免坐标/视口导致生成、刷新、下载点不中。
  2. 状态机固定：待生成->生成->继续导出->刷新；校验中/生成中->刷新；已生成->下载。
  3. 下载等待超时 900 -> 1200 秒，适配 350+ 条后台生成。

20260615_v02 修改（Phase 1 稳定性修复）:
  1. wait_for_download_file 超时 240 -> 900 秒（350 条生成常超 4 分钟）。
  2. wait_for_export_record_page 超时 180 -> 300 秒。
  3. handle_quantity_prompt_during_export_record "继续导出"重试上限 2 -> 4。
  4. 下载等待循环新增"账号已在其他设备登录"识别 -> fail-fast 明确报错。
  5. 点"下载"前先 scrollIntoView 再重取坐标点击（修下载按钮 Visible=False / 视口外点不中）。
  6. set_date_filter_by_vuex 日期提交+校验最多重试3次（降低偶发日期校验失败）。
  7. 登录超时弹窗文本归一化，支持“确 定”等带空格按钮；登录检测不再用“订阅中心”弱判断放行。
  8. 导出记录页按“本次点击时间 + 导出条数接近度”选目标行，不再只处理第一行。
  9. 导出提交后若页面未自动跳转，允许进入已验证的 myFocus_getExport.html 导出记录入口；仍按本次时间和条数校验目标行，禁止下载旧记录。

使用前提:
1. Chrome 已用 CDP 启动，例如:
   chrome.exe --remote-debugging-port=9222 --remote-allow-origins=* --user-data-dir="%LOCALAPPDATA%\Google\Chrome\User Data CDP"
2. 千里马 VIP 账号已在该 Chrome Profile 中登录。
3. Python 环境已安装 websocket-client。

安全边界:
- 本脚本不包含账号、密码、Cookie、Token。
- 遇到“账号风险提醒”只点击“忽略”，不会点击“立即修改”，不会修改密码。
- 日期筛选以页面真实 Vuex 状态为成功标准，不以 UI 是否关闭为准。
- 最高规则：不清空用户浏览器数据，不干扰用户正在使用的浏览器。
- 不点击列表“全选”；导出范围必须在导出配置弹窗中选择“全部搜索结果”。
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import re
import subprocess
import sys
import time
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

import websocket


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


DEFAULT_SUBSCRIBE_URL = "https://vip.qianlima.com/subscribe-center/subscribe-info/bidding-info?id=159970"
EXPORT_RECORD_ENTRY_URL = "https://vip.qianlima.com/myFocus_getExport.html"
DEFAULT_OUTPUT_ROOT = Path(__file__).resolve().parent


class RunContext:
    def __init__(self, output_root: Path, start_date: str, end_date: str, cdp_download_root: Path | None = None):
        self.run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_root = output_root
        self.start_date = start_date
        self.end_date = end_date
        self.screenshot_dir = output_root / f"qianlima_mainline_screenshots_{self.run_id}"
        self.download_dir = output_root / f"qianlima_mainline_downloads_{self.run_id}"
        # CDP 下载目录使用英文路径（Page.setDownloadBehavior 不支持中文路径）
        self.cdp_download_dir = (cdp_download_root or Path("C:/temp/cdp_download")) / self.run_id
        self.log_json = output_root / f"qianlima_mainline_log_{self.run_id}.json"
        self.log_md = output_root / f"qianlima_mainline_log_{self.run_id}.md"
        self.live_log_jsonl = output_root / f"qianlima_mainline_live_{self.run_id}.jsonl"
        self.checkpoint_path = output_root / f"qianlima_phase1_checkpoint_{self.run_id}.json"
        self.events: list[dict] = []
        self.export_confirm_clicked_at: datetime | None = None
        self.expected_export_count: int | None = None
        self.filter_result_count: int | None = None
        self.quota_export_count: int | None = None
        self.quota_remaining: int | None = None  # 弹窗解析的剩余额度；0 不 fail-fast（150 窗口仍可生成→下载）
        self.target_record_time: datetime | None = None
        self.target_record_count: int | None = None
        self.generation_submitted = False
        self.generation_attempts = 0
        self.last_generation_clicked_at: datetime | None = None
        self.generation_wait_started_at: datetime | None = None
        self.quantity_prompt_confirm_attempts = 0
        self.initial_quota_confirmed = False
        self.generation_quota_confirmed = False
        self.download_clicked = False
        self.allow_existing_export_download = False
        self.export_source = "new_export"  # or fallback_existing_export
        self.screenshot_dir.mkdir(parents=True, exist_ok=True)
        self.download_dir.mkdir(parents=True, exist_ok=True)
        self.cdp_download_dir.mkdir(parents=True, exist_ok=True)

    def checkpoint_payload(self) -> dict:
        def iso(value):
            return value.isoformat(sep=" ") if value else None

        return {
            "schema_version": 1,
            "run_id": self.run_id,
            "start_date": self.start_date,
            "end_date": self.end_date,
            "export_confirm_clicked_at": iso(self.export_confirm_clicked_at),
            "expected_export_count": self.expected_export_count,
            "filter_result_count": self.filter_result_count,
            "quota_export_count": self.quota_export_count,
            "quota_remaining": self.quota_remaining,
            "target_record_time": iso(self.target_record_time),
            "target_record_count": self.target_record_count,
            "generation_submitted": self.generation_submitted,
            "export_source": self.export_source,
            "allow_existing_export_download": self.allow_existing_export_download,
            "generation_attempts": self.generation_attempts,
            "last_generation_clicked_at": iso(self.last_generation_clicked_at),
            "generation_wait_started_at": iso(self.generation_wait_started_at),
            "quantity_prompt_confirm_attempts": self.quantity_prompt_confirm_attempts,
            "initial_quota_confirmed": self.initial_quota_confirmed,
            "generation_quota_confirmed": self.generation_quota_confirmed,
            "download_clicked": self.download_clicked,
            "event_count": len(self.events),
            "updated_at": datetime.now().isoformat(sep=" ", timespec="seconds"),
        }

    def write_checkpoint(self) -> None:
        self.checkpoint_path.write_text(
            json.dumps(self.checkpoint_payload(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def record(self, step: str, status: str, detail=None):
        entry = {
            "time": datetime.now().strftime("%H:%M:%S"),
            "step": step,
            "status": status,
            "detail": detail or {},
        }
        self.events.append(entry)
        with self.live_log_jsonl.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(entry, ensure_ascii=False) + "\n")
        self.write_checkpoint()
        print(f"[{entry['time']}] {step}: {status} {json.dumps(detail or {}, ensure_ascii=False)}", flush=True)
    def save_logs(self):
        self.log_json.write_text(json.dumps(self.events, ensure_ascii=False, indent=2), encoding="utf-8")
        lines = [
            "# 千里马 VIP 主线导出日志",
            "",
            f"- 运行时间：{self.run_id}",
            f"- 目标日期：{self.start_date} 至 {self.end_date}",
            f"- 截图目录：`{self.screenshot_dir}`",
            f"- 下载目录：`{self.download_dir}`",
            "",
            "## 过程记录",
            "",
        ]
        for e in self.events:
            lines.append(f"- {e['time']} | {e['step']} | {e['status']} | `{json.dumps(e['detail'], ensure_ascii=False)}`")
        self.log_md.write_text("\n".join(lines), encoding="utf-8")
        print(f"[log] json={self.log_json}", flush=True)
        print(f"[log] md={self.log_md}", flush=True)


class CDP:
    def __init__(self, tab: dict, port: int):
        self.port = port
        self.tab = tab
        # Chrome 150+ DevTools 只接受 Host: localhost，拒绝 127.0.0.1（返回 404）
        self.ws = websocket.create_connection(
            tab["webSocketDebuggerUrl"],
            suppress_origin=True,
            host="localhost",
            origin=f"http://localhost:{port}",
            timeout=15,
        )
        self._id = 1

    def close(self):
        try:
            self.ws.close()
        except Exception:
            pass

    def send(self, method: str, params=None):
        msg_id = self._id
        self._id += 1
        payload = {"id": msg_id, "method": method}
        if params is not None:
            payload["params"] = params
        self.ws.send(json.dumps(payload))
        return msg_id

    def recv(self, msg_id: int, timeout=15):
        start = time.time()
        old_timeout = self.ws.gettimeout()
        self.ws.settimeout(1)
        try:
            while time.time() - start < timeout:
                try:
                    msg = json.loads(self.ws.recv())
                except Exception:
                    continue
                if msg.get("id") == msg_id:
                    return msg
        finally:
            self.ws.settimeout(old_timeout)
        return None

    def call(self, method: str, params=None, timeout=15):
        return self.recv(self.send(method, params), timeout=timeout)

    def enable(self):
        for domain in ["Runtime", "Page", "DOM", "Network"]:
            self.call(f"{domain}.enable", timeout=8)

    def eval(self, expression: str, timeout=15):
        r = self.call(
            "Runtime.evaluate",
            {"expression": expression, "returnByValue": True, "awaitPromise": True},
            timeout=timeout,
        )
        if not r:
            return None
        result = r.get("result", {}).get("result", {})
        if result.get("subtype") == "error":
            return {"__error__": result.get("description") or result.get("className")}
        if "value" in result:
            return result["value"]
        if result.get("type") == "undefined":
            return None
        return result

    def navigate(self, url: str, wait=5):
        self.call("Page.navigate", {"url": url}, timeout=10)
        time.sleep(wait)

    def click(self, x: int, y: int, wait=0.5):
        for params in [
            {"type": "mouseMoved", "x": x, "y": y},
            {"type": "mousePressed", "x": x, "y": y, "button": "left", "clickCount": 1},
            {"type": "mouseReleased", "x": x, "y": y, "button": "left", "clickCount": 1},
        ]:
            self.call("Input.dispatchMouseEvent", params, timeout=5)
        time.sleep(wait)

    def press_escape(self, wait=0.5):
        for params in [
            {"type": "rawKeyDown", "key": "Escape", "code": "Escape", "windowsVirtualKeyCode": 27, "nativeVirtualKeyCode": 27},
            {"type": "keyUp", "key": "Escape", "code": "Escape", "windowsVirtualKeyCode": 27, "nativeVirtualKeyCode": 27},
        ]:
            self.call("Input.dispatchKeyEvent", params, timeout=5)
        time.sleep(wait)

    def screenshot(self, ctx: RunContext, name: str):
        path = ctx.screenshot_dir / name
        r = self.call("Page.captureScreenshot", {"format": "png"}, timeout=15)
        data = r and r.get("result", {}).get("data")
        if data:
            path.write_bytes(base64.b64decode(data))
            ctx.record("screenshot", "saved", {"path": str(path), "kb": round(path.stat().st_size / 1024, 1)})
        return path


def http_json(port: int, path: str):
    # Chrome 150+ DevTools 只接受 Host: localhost（拒绝 127.0.0.1）；
    # 同时绕过系统代理（如 127.0.0.1:7897），否则本地请求会被代理拦截返回 404。
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    with opener.open(f"http://localhost:{port}{path}", timeout=8) as resp:
        return json.loads(resp.read().decode("utf-8"))


def visible_pages(port: int):
    tabs = http_json(port, "/json")
    return [t for t in tabs if t.get("type") == "page" and t.get("webSocketDebuggerUrl")]


def attach_qianlima(port: int, subscribe_url: str, ctx: RunContext) -> CDP:
    pages = visible_pages(port)
    qlm_pages = [t for t in pages if "vip.qianlima.com" in t.get("url", "")]
    if qlm_pages:
        tab = qlm_pages[0]
    elif pages:
        tab = pages[0]
    else:
        raise RuntimeError("未找到可连接的 Chrome 页面。请确认 Chrome 已用 --remote-debugging-port 启动。")
    agent = CDP(tab, port)
    agent.enable()
    agent.call("Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": str(ctx.cdp_download_dir)}, timeout=5)
    url = tab.get("url", "")
    ctx.record("tab_attach", "selected", {"title": tab.get("title"), "url": url})
    if "vip.qianlima.com" not in url:
        agent.navigate(subscribe_url, wait=6)
    return agent


def dismiss_popups(agent: CDP, ctx: RunContext):
    hits = agent.eval(
        r"""
        (() => {
          const normalize = s => (s || '').replace(/\s+/g, '').trim();
          // 登录超时弹窗：优先点「确定」，不要点「取消」
          const closeTexts = ['关闭','不记录','否','以后再说','知道了','我知道了','忽略','确定','×','x'];
          const body = document.body.innerText || '';
          const loginTimeout = body.includes('登录状态超时') || body.includes('请重新登录') || body.includes('登录已过期');
          const hits = [];
          for (const el of document.querySelectorAll('button, a, span, div, i')) {
            const raw = el.textContent || '';
            const t = normalize(raw);
            const displayText = raw.replace(/\s+/g, ' ').trim();
            const cls = String(el.className || '').toLowerCase();
            const r = el.getBoundingClientRect();
            if (r.height <= 0 || r.width <= 0 || r.height > 80 || r.width > 220) continue;
            if (loginTimeout && t === '确定') {
              hits.push({
                x: Math.round(r.x + r.width/2),
                y: Math.round(r.y + r.height/2),
                text: displayText || cls.slice(0,30),
                normalizedText: t,
                accountRisk: false,
                loginTimeoutConfirm: true,
                priority: 0
              });
              continue;
            }
            if (closeTexts.includes(t) || cls.includes('close')) {
              hits.push({
                x: Math.round(r.x + r.width/2),
                y: Math.round(r.y + r.height/2),
                text: displayText || cls.slice(0,30),
                normalizedText: t,
                accountRisk: body.includes('账号风险提醒') && t === '忽略',
                loginTimeoutConfirm: false,
                priority: 1
              });
            }
          }
          hits.sort((a, b) => (a.priority || 0) - (b.priority || 0));
          return hits.slice(0, 5);
        })()
        """
    ) or []
    for h in hits:
        if h.get("x", 0) > 0 and h.get("y", 0) > 0:
            agent.click(h["x"], h["y"])
            ctx.record("dismiss_popup", "clicked", h)
            if h.get("accountRisk"):
                ctx.record("account_risk_prompt", "ignored", h)
            if h.get("loginTimeoutConfirm"):
                ctx.record("login_timeout_prompt", "confirmed", h)
                # 登录超时点确定后通常会进登录页，立即走自动登录，不要停在表单前干等
                break
    return hits


def _login_script_path() -> Path:
    return Path(__file__).with_name("qianlima_auto_login_20260622_v02.py")


def recover_login_via_auto_login(ctx: RunContext, cdp_port: int, config_path: str | None) -> dict:
    """无人值守：调用方案A自动登录（关超时弹窗 → 密码登录 → 填账密 → 点登录）。

    禁止要求人工登录；仅验证码/账号风险二次验证时失败并写明原因。
    """
    script = _login_script_path()
    if not script.exists():
        raise RuntimeError(f"自动登录脚本缺失：{script}")
    cfg = config_path or os.environ.get("QLM_BID_CONFIG") or ""
    if not cfg or not Path(cfg).exists():
        raise RuntimeError(
            "登录态失效且无法自动登录：未找到本机敏感配置（--config / QLM_BID_CONFIG）。"
            "自动化流程必须能读到账号密码，禁止停在登录框前等待人工。"
        )
    cmd = [
        sys.executable,
        "-X",
        "utf8",
        str(script),
        "--mode",
        "cdp",
        "--cdp-port",
        str(cdp_port),
        "--config",
        cfg,
    ]
    ctx.record("auto_login", "start", {"cmd": " ".join(cmd[:6]) + " --config <redacted>"})
    proc = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=120,
    )
    blob = ((proc.stdout or "") + "\n" + (proc.stderr or "")).strip()
    payload = {}
    # 脚本末行是 JSON
    for line in reversed(blob.splitlines()):
        line = line.strip()
        if line.startswith("{") and line.endswith("}"):
            try:
                payload = json.loads(line)
                break
            except Exception:
                pass
    ctx.record(
        "auto_login",
        "finished",
        {
            "returncode": proc.returncode,
            "ok": bool(payload.get("ok")),
            "reason": (payload.get("reason") or blob[:300]),
        },
    )
    if proc.returncode != 0 or not payload.get("ok"):
        reason = payload.get("reason") or blob[:400] or f"rc={proc.returncode}"
        raise RuntimeError(
            f"自动登录失败（无人值守应可恢复）：{reason}。"
            "若为验证码/账号风险二次验证，需人工一次处理后再跑；否则检查配置账密与页面结构。"
        )
    return payload


def read_login_state(agent: CDP) -> dict:
    return agent.eval(
        r"""
        (() => {
          const body = document.body.innerText || '';
          const url = location.href || '';
          const loginTimeout = body.includes('登录状态超时') || body.includes('请重新登录') || body.includes('登录已过期');
          const onLoginPage = /\/login\b/i.test(url) || body.includes('密码登录') || body.includes('微信扫码登录');
          const hasPassword = !!document.querySelector('input[type=password]');
          const hasAccount = [...document.querySelectorAll('input')].some(i => {
            const t = (i.getAttribute('type')||'').toLowerCase();
            if (t === 'password' || t === 'hidden') return false;
            const ph = i.getAttribute('placeholder') || '';
            return /用户名|账号|手机|手机号/.test(ph) || t === 'text' || t === 'tel';
          });
          const strongLogged = body.includes('高级会员') || body.includes('退出') ||
            !!document.querySelector('.user-avatar, .user-info, [class*=member], [class*=nickname]');
          const subscribeReady = body.includes('标讯订阅') && (body.includes('结果筛选') || body.includes('批量导出') || body.includes('广东全量'));
          const exportRecordReady = body.includes('导出记录') && body.includes('导出日期') && /20\d{2}[-/]\d{1,2}[-/]\d{1,2}\s+\d{1,2}:\d{2}/.test(body);
          const logged = !loginTimeout && !onLoginPage && (strongLogged || subscribeReady || exportRecordReady);
          return {
            url, logged, loginGate: loginTimeout, onLoginPage, hasPassword, hasAccount,
            title: document.title, sample: body.slice(0, 500)
          };
        })()
        """
    ) or {}


def ensure_on_subscribe_page(
    agent: CDP,
    ctx: RunContext,
    subscribe_url: str,
    *,
    cdp_port: int = 9222,
    config_path: str | None = None,
) -> CDP:
    """确保在订阅页且已登录。登录超时/掉线 → 自动登录恢复，禁止要求人工登录。

    返回可能为重新 attach 后的 agent（调用方必须使用返回值）。
    """
    dismissed = dismiss_popups(agent, ctx)
    if any(h.get("loginTimeoutConfirm") for h in dismissed):
        time.sleep(2)
    state = read_login_state(agent)
    ctx.record("login_state", "check", state)

    need_recover = (
        bool(state.get("loginGate"))
        or bool(state.get("onLoginPage"))
        or not state.get("logged")
        or any(h.get("loginTimeoutConfirm") for h in dismissed)
    )
    if need_recover and not state.get("logged"):
        ctx.record(
            "login_state",
            "need_auto_recover",
            {
                "after_timeout_confirm": any(h.get("loginTimeoutConfirm") for h in dismissed),
                "onLoginPage": state.get("onLoginPage"),
                "loginGate": state.get("loginGate"),
            },
        )
        # 关闭当前 CDP 连接，避免与 auto_login 抢同一 tab
        try:
            agent.close()
        except Exception:
            pass
        recover_login_via_auto_login(ctx, cdp_port, config_path)
        agent = attach_qianlima(cdp_port, subscribe_url, ctx)
        time.sleep(1)
        dismiss_popups(agent, ctx)
        state = read_login_state(agent)
        ctx.record("login_state", "after_auto_recover", state)
        if not state.get("logged"):
            raise RuntimeError(
                f"自动登录已执行但仍无登录态。url={state.get('url')} sample={(state.get('sample') or '')[:120]}"
            )

    if "subscribe-center/subscribe-info/bidding-info" not in (state.get("url") or ""):
        agent.navigate(subscribe_url, wait=6)
        dismiss_popups(agent, ctx)
        state = read_login_state(agent)
        if not state.get("logged"):
            # 导航后偶发又掉登录：再恢复一次
            try:
                agent.close()
            except Exception:
                pass
            recover_login_via_auto_login(ctx, cdp_port, config_path)
            agent = attach_qianlima(cdp_port, subscribe_url, ctx)
    return agent


def current_date_filter(agent: CDP):
    return agent.eval(
        r"""
        (() => {
          const info = document.querySelector('.info-filter')?.__vue__;
          const dropdown = Array.from(document.querySelectorAll('._dropdown'))
            .find(d => d.__vue__?.$props?.placeholder === '发布时间')?.__vue__;
          return {
            url: location.href,
            currentFilterParams: info ? info.currentFilterParams : null,
            upListNumber: info?.$store?.state?.subscribe?.upListNumber,
            dropdownCurrent: dropdown ? dropdown.$props.current : null,
            bodyHasStart: document.body.innerText.includes('2026-05-26'),
            bodyHasEnd: document.body.innerText.includes('2026-05-27')
          };
        })()
        """
    )


def date_filter_matches(state: dict, start_date: str, end_date: str):
    params = (state or {}).get("currentFilterParams") or {}
    dropdown = (state or {}).get("dropdownCurrent") or {}
    return (
        params.get("dateType") in (6, 99)
        and params.get("startTime") == f"{start_date} 00:00:00"
        and params.get("endTime") == f"{end_date} 23:59:59"
        and dropdown.get("id") == 99
    )

def set_date_filter_by_vuex(agent: CDP, ctx: RunContext, start_date: str, end_date: str):
    before = current_date_filter(agent)
    ctx.record("date_filter_before", "state", before)
    if date_filter_matches(before, start_date, end_date):
        ctx.record("date_filter", "already_target", before)
        return before

    component_js = f"""
        (() => {{
          const dropdowns = Array.from(document.querySelectorAll('._dropdown'))
            .filter(el => el.offsetParent !== null);
          const target = dropdowns[0];
          if (!target) return {{ok:false, reason:'no visible date dropdown'}};
          let vm = target.__vue__;
          let depth = 0;
          while (vm && typeof vm.datePickerChange !== 'function' && depth < 8) {{
            vm = vm.$parent;
            depth += 1;
          }}
          if (!vm || typeof vm.datePickerChange !== 'function') {{
            return {{ok:false, reason:'DropdownDateType.datePickerChange not found', dropdownText:target.innerText}};
          }}
          vm.datePickerChange(['{start_date}', '{end_date}']);
          if (typeof vm.$forceUpdate === 'function') vm.$forceUpdate();
          return {{
            ok:true,
            depth,
            componentName:vm.$options?.name || '',
            dropdownText:(target.innerText || '').trim(),
            currentSelect:vm.currentSelect || '',
            beginTimeAndEndTime:vm.beginTimeAndEndTime || null
          }};
        }})()
        """
    after = before
    for attempt in range(1, 4):
        result = agent.eval(component_js)
        ctx.record("date_filter_component", "result", {"attempt": attempt, "result": result})
        if not result or not result.get("ok"):
            time.sleep(2)
            continue
        time.sleep(6)
        after = current_date_filter(agent)
        ctx.record("date_filter_after", "state", {"attempt": attempt, "state": after})
        if date_filter_matches(after, start_date, end_date):
            agent.screenshot(ctx, "date_filter_verified.png")
            return after
        ctx.record("date_filter", "retry", {"attempt": attempt, "component": result})
        time.sleep(3)
    agent.screenshot(ctx, "date_filter_verify_failed.png")
    raise RuntimeError("Date component did not persist the requested custom date range.")

def batch_export_button_state(agent: CDP):
    return agent.eval(
        r"""
        (() => {
          const wanted = '\u6279\u91cf\u5bfc\u51fa';
          const buttons = Array.from(document.querySelectorAll('button')).map(el => {
            const text = (el.textContent || '').replace(/\s+/g, ' ').trim();
            const r = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            const disabled = !!el.disabled || el.classList.contains('is-disabled') ||
              el.classList.contains('disable-btn') || style.pointerEvents === 'none';
            return {text, x:Math.round(r.x+r.width/2), y:Math.round(r.y+r.height/2),
              w:Math.round(r.width), h:Math.round(r.height), disabled,
              cls:String(el.className || '')};
          }).filter(o => o.text === wanted && o.w > 0 && o.h > 0);
          const enabled = buttons.find(o => !o.disabled) || null;
          return {found:buttons.length > 0, enabled:!!enabled, button:enabled, candidates:buttons};
        })()
        """
    )


def wait_for_export_ready(agent: CDP, ctx: RunContext, timeout: int = 120):
    """Require a stable custom date, positive result count and enabled export button."""
    deadline = time.time() + timeout
    stable_hits = 0
    reapply_attempts = 0
    while time.time() < deadline:
        date_state = current_date_filter(agent) or {}
        button_state = batch_export_button_state(agent) or {}
        count = date_state.get("upListNumber")
        matches = date_filter_matches(date_state, ctx.start_date, ctx.end_date)
        ready = matches and isinstance(count, int) and count > 0 and button_state.get("enabled")
        ctx.record(
            "export_ready",
            "poll",
            {
                "date_matches": matches,
                "count": count,
                "button_enabled": bool(button_state.get("enabled")),
                "stable_hits": stable_hits,
                "reapply_attempts": reapply_attempts,
            },
        )
        if ready:
            stable_hits += 1
            if stable_hits >= 2:
                ctx.filter_result_count = count
                ctx.expected_export_count = count
                ctx.record("export_ready", "stable", {"count": count})
                return date_state
        else:
            stable_hits = 0
            if not matches:
                if reapply_attempts >= 3:
                    break
                reapply_attempts += 1
                ctx.record("date_filter", "drift_reapply", {"attempt": reapply_attempts, "state": date_state})
                time.sleep(4)
                set_date_filter_by_vuex(agent, ctx, ctx.start_date, ctx.end_date)
        time.sleep(5)
    agent.screenshot(ctx, "export_ready_timeout.png")
    raise RuntimeError("Date filter or result list did not stabilize; batch export is unavailable.")

def click_candidate(agent: CDP, ctx: RunContext, js: str, label: str, wait=1.5):
    res = agent.eval(js)
    ctx.record(f"{label}_find", "result", res)
    if not res or not res.get("found"):
        agent.screenshot(ctx, f"{label}_not_found.png")
        return False
    c = res["candidates"][0]
    agent.click(c["x"], c["y"])
    ctx.record(f"{label}_click", "clicked", c)
    time.sleep(wait)
    return True


def click_batch_export(agent: CDP, ctx: RunContext):
    return click_candidate(
        agent,
        ctx,
        r"""
        (() => {
          const wanted = '\u6279\u91cf\u5bfc\u51fa';
          const cands = [];
          for (const el of document.querySelectorAll('button')) {
            const text = (el.textContent || '').replace(/\s+/g,' ').trim();
            const r = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            const disabled = !!el.disabled || el.classList.contains('is-disabled') ||
              el.classList.contains('disable-btn') || style.pointerEvents === 'none';
            if (text !== wanted || r.width <= 0 || r.height <= 0 || disabled) continue;
            cands.push({x:Math.round(r.x+r.width/2), y:Math.round(r.y+r.height/2), left:Math.round(r.x), top:Math.round(r.y), text, tag:el.tagName, cls:String(el.className||'')});
          }
          cands.sort((a,b) => b.left-a.left || a.top-b.top);
          return {found:cands.length>0, candidates:cands};
        })()
        """,
        "batch_export",
        wait=2,
    )

def ensure_export_dialog_options(agent: CDP, ctx: RunContext):
    state = agent.eval(
        r"""
        (() => {
          const body = document.body.innerText || '';
          return {
            hasDialog: body.includes('导出配置') && body.includes('导出范围'),
            sample: body.slice(-1000)
          };
        })()
        """
    )
    ctx.record("export_dialog", "state", state)
    if not state or not state.get("hasDialog"):
        agent.screenshot(ctx, "export_dialog_not_visible.png")
        raise RuntimeError("未检测到导出配置弹窗。")

    # 尽量选择全部搜索结果、Excel、拓展字段。若页面默认已选中，点击同项通常无副作用。
    for text, label in [("全部搜索结果", "dialog_all_results"), ("Excel", "dialog_excel"), ("拓展字段", "dialog_extended_fields")]:
        click_candidate(
            agent,
            ctx,
            f"""
            (() => {{
              const cands = [];
              for (const el of document.querySelectorAll('.el-dialog__wrapper span, .el-dialog__wrapper label, .el-dialog__wrapper div')) {{
                const t = (el.textContent || '').replace(/\\s+/g,' ').trim();
                const r = el.getBoundingClientRect();
                if (t !== {json.dumps(text, ensure_ascii=False)} || r.width <= 0 || r.height <= 0) continue;
                cands.push({{x:Math.round(r.x+r.width/2), y:Math.round(r.y+r.height/2), left:Math.round(r.x), top:Math.round(r.y), text:t, cls:String(el.className||'').slice(0,80), tag:el.tagName}});
              }}
              cands.sort((a,b) => a.top-b.top || a.left-b.left);
              return {{found:cands.length>0, candidates:cands.slice(0,5)}};
            }})()
            """,
            label,
            wait=0.5,
        )


def click_visible_export_button(agent: CDP, ctx: RunContext, label: str):
    return click_candidate(
        agent,
        ctx,
        r"""
        (() => {
          const cands = [];
          for (const el of document.querySelectorAll('.sure-btn, .el-message-box__btns button, .el-dialog__footer button, button, div')) {
            const text = (el.textContent || '').replace(/\s+/g,' ').trim();
            const normalizedText = text.replace(/\s+/g, '');
            const r = el.getBoundingClientRect();
            const cls = String(el.className || '');
            if (!['导出', '导 出'].includes(text) && normalizedText !== '导出') continue;
            const style = getComputedStyle(el);
            if (style.visibility === 'hidden' || style.display === 'none') continue;
            const inVisibleDialog = !!el.closest('.el-dialog__wrapper:not([style*="display: none"]), .el-message-box__wrapper:not([style*="display: none"])');
            cands.push({x:Math.round(r.x+r.width/2), y:Math.round(r.y+r.height/2), left:Math.round(r.x), top:Math.round(r.y), text, tag:el.tagName, cls:cls.slice(0,80), inVisibleDialog});
          }
          cands.sort((a,b) => Number(b.inVisibleDialog)-Number(a.inVisibleDialog) || b.top-a.top || b.left-a.left);
          return {found:cands.length>0, candidates:cands.slice(0,8)};
        })()
        """,
        label,
        wait=3,
    )


def click_quantity_confirm_export(agent: CDP, ctx: RunContext):
    return click_candidate(
        agent,
        ctx,
        r"""
        (() => {
          const normalize = s => (s || '').replace(/\s+/g, '').trim();
          const pageText = document.body.innerText || '';
          const isConfirmText = text => ['导出', '继续导出', '确认导出', '导 出'].includes(text);
          const hasQuantityPrompt =
            pageText.includes('导出数量提示') ||
            pageText.includes('是否继续导出') ||
            pageText.includes('访问额度不足') ||
            pageText.includes('剩余额度') ||
            pageText.includes('可导出部分');
          const wrappers = Array.from(document.querySelectorAll('.el-message-box__wrapper, .el-dialog__wrapper, .el-message-box, .el-dialog, body'))
            .filter(w => {
              const r = w.getBoundingClientRect();
              const style = getComputedStyle(w);
              const text = (w.textContent || '').replace(/\s+/g, ' ').trim();
              return r.width > 0 && r.height > 0 &&
                style.visibility !== 'hidden' && style.display !== 'none' &&
                (text.includes('导出数量提示') ||
                 text.includes('是否继续导出') ||
                 text.includes('访问额度不足') ||
                 text.includes('剩余额度') ||
                 text.includes('可导出部分'));
            });
          const cands = [];
          for (const wrapper of wrappers) {
            const wrapperRect = wrapper.getBoundingClientRect();
            const z = Number(getComputedStyle(wrapper).zIndex) || 0;
            for (const el of wrapper.querySelectorAll('button, span, div, a')) {
              const text = normalize(el.textContent);
              if (!isConfirmText(text)) continue;
              const r = el.getBoundingClientRect();
              const style = getComputedStyle(el);
              if (r.width <= 0 || r.height <= 0 || style.visibility === 'hidden' || style.display === 'none') continue;
              cands.push({
                x: Math.round(r.x + r.width / 2),
                y: Math.round(r.y + r.height / 2),
                left: Math.round(r.x),
                top: Math.round(r.y),
                text: (el.textContent || '').replace(/\s+/g, ' ').trim(),
                tag: el.tagName,
                cls: String(el.className || '').slice(0, 80),
                wrapperText: (wrapper.textContent || '').replace(/\s+/g, ' ').trim().slice(0, 160),
                wrapperTop: Math.round(wrapperRect.y),
                z
              });
            }
          }
          if (hasQuantityPrompt) {
            for (const el of document.querySelectorAll('button, span, div, a')) {
              const text = normalize(el.textContent);
              if (!isConfirmText(text)) continue;
              const r = el.getBoundingClientRect();
              const style = getComputedStyle(el);
              if (r.width <= 0 || r.height <= 0 || style.visibility === 'hidden' || style.display === 'none') continue;
              const x = Math.round(r.x + r.width / 2);
              const y = Math.round(r.y + r.height / 2);
              const topEl = document.elementFromPoint(x, y);
              const topMost = !!topEl && (el === topEl || el.contains(topEl) || topEl.contains(el));
              cands.push({
                x, y,
                left: Math.round(r.x),
                top: Math.round(r.y),
                text: (el.textContent || '').replace(/\s+/g, ' ').trim(),
                tag: el.tagName,
                cls: String(el.className || '').slice(0, 80),
                globalPromptCandidate: true,
                topMost,
                z: Number(getComputedStyle(el).zIndex) || 0
              });
            }
          }
          cands.sort((a, b) => Number(b.topMost) - Number(a.topMost) || b.z - a.z || b.top - a.top || b.left - a.left);
          if (!cands.length && hasQuantityPrompt) {
            return {
              found: true,
              candidates: [{
                x: Math.round(window.innerWidth / 2 + 160),
                y: Math.round(window.innerHeight / 2 + 65),
                text: '继续导出',
                fallback: 'quantity_prompt_geometry'
              }]
            };
          }
          return {found: cands.length > 0, candidates: cands.slice(0, 8)};
        })()
        """,
        "quantity_confirm_export",
        wait=3,
    )


QUANTITY_PROMPT_MARKERS = ("导出数量提示", "是否继续导出", "访问额度不足", "剩余额度", "可导出部分")
# 强特征：单独「可导出部分」等弱词会在页面残留文案/隐藏 dialog 文本里误匹配
QUANTITY_PROMPT_STRONG = ("导出数量提示", "是否继续导出", "访问额度不足", "今日剩余额度", "剩余额度0", "剩余额度1")


def quantity_prompt_in_text(text: str):
    """True only when text looks like the real quota dialog, not generic page chrome."""
    t = text or ""
    if any(s in t for s in QUANTITY_PROMPT_STRONG):
        return True
    # 弱词需至少两个同时出现，避免误判
    weak_hits = sum(1 for s in ("剩余额度", "可导出部分", "继续导出") if s in t)
    return weak_hits >= 2


def quantity_prompt_visible(agent: CDP) -> bool:
    """Prefer visible dialog geometry; ignore display:none / 0-size wrappers."""
    dialog = read_top_quantity_dialog(agent)
    if dialog.get("found"):
        return True
    # Fallback: strong markers in visible body only (not hidden dialog stash)
    state = page_state(agent)
    body = str((state or {}).get("body") or "")
    return any(s in body for s in ("导出数量提示", "访问额度不足", "是否继续导出"))


def extract_quota_remaining(text: str) -> int | None:
    """Extract 今日剩余额度 N（含 0）。用于额度 0 fail-fast。

    常见文案：访问额度不足，是否消耗今日剩余额度0条导出可导出部分？
    """
    normalized = re.sub(r"\s+", "", text or "")
    patterns = (
        r"今日剩余额度(\d{1,5})条",
        r"剩余(?:访问)?额度[:：]?(\d{1,5})条?",
        r"剩余额度(\d{1,5})",
    )
    for pattern in patterns:
        m = re.search(pattern, normalized)
        if m:
            return int(m.group(1))
    return None


def extract_quota_export_count(text: str) -> int | None:
    """Extract the actual allowed export count from the quota prompt (positive only).

    常见文案：访问额度不足，是否消耗今日剩余额度 N 条导出可导出部分？
    注意：N=0 时由 extract_quota_remaining 处理 fail-fast，本函数不把 0 当可导出条数。
    """
    remaining = extract_quota_remaining(text)
    if remaining == 0:
        return None
    normalized = re.sub(r"\s+", "", text or "")
    patterns = (
        r"今日剩余额度(\d{1,5})条",
        r"剩余(?:访问)?额度[:：]?(\d{1,5})条?",
        r"剩余额度(\d{1,5})",
        r"可导出部分[：:]?(\d{1,5})条?",
        r"(\d{1,5})条导出可导出部分",
        r"(\d{1,5})条可导出部分",
        r"可导出(?:条数)?[:：]?(\d{1,5})条?",
        r"本次可导出[:：]?(\d{1,5})条?",
        r"导出数量[:：]?(\d{1,5})条?",
        r"额度不足[^\d]{0,30}(\d{1,5})条",
        r"仅可导出(\d{1,5})条",
        r"将导出(\d{1,5})条",
        r"消耗[^\d]{0,12}(\d{1,5})条",
    )
    values = []
    for pattern in patterns:
        values.extend(int(value) for value in re.findall(pattern, normalized))
    # 兜底：弹窗里靠近「条」的数字（排除年份）
    if not values:
        for m in re.finditer(r"(\d{1,5})条", normalized):
            n = int(m.group(1))
            if 1 <= n <= 5000:
                values.append(n)
    values = [value for value in values if 0 < value < 100000]
    return min(values) if values else None


def read_top_quantity_dialog(agent: CDP) -> dict:
    """Read topmost quota dialog text/loading state from DOM (not truncated body slice)."""
    return agent.eval(
        r"""
        (() => {
          const markers = ['导出数量提示', '是否继续导出', '访问额度不足', '剩余额度', '可导出部分', '继续导出'];
          const visible = el => {
            if (!el) return false;
            const r = el.getBoundingClientRect();
            const s = getComputedStyle(el);
            return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none' && s.opacity !== '0';
          };
          const wrappers = [...document.querySelectorAll(
            '.layui-layer, .layui-layer-dialog, .layui-layer-page, .el-message-box__wrapper, .el-dialog__wrapper, .el-message-box, .el-dialog, [role=dialog]'
          )].filter(visible);
          const scored = wrappers.map(w => {
            const text = (w.innerText || w.textContent || '').replace(/\s+/g, ' ').trim();
            const z = Number(getComputedStyle(w).zIndex) || 0;
            const hit = markers.some(m => text.includes(m));
            return {text, z, hit, top: w.getBoundingClientRect().y};
          }).filter(x => x.hit);
          scored.sort((a,b) => b.z - a.z || a.top - b.top);
          const top = scored[0];
          if (!top) {
            const body = (document.body.innerText || '').replace(/\s+/g, ' ');
            const hit = markers.some(m => body.includes(m));
            return {found: hit, loading: /加载中/.test(body), text: body.slice(-1500), source: 'body'};
          }
          return {
            found: true,
            loading: /加载中/.test(top.text),
            text: top.text.slice(0, 2000),
            source: 'dialog',
            z: top.z
          };
        })()
        """
    ) or {}


def wait_and_confirm_quantity_prompt(agent: CDP, ctx: RunContext, stage: str, max_wait: float = 25.0) -> bool:
    """Wait until quota dialog finished loading, parse quota, click topmost 继续导出 once."""
    deadline = time.time() + max_wait
    saw_prompt = False
    while time.time() < deadline:
        dialog = read_top_quantity_dialog(agent)
        text = str(dialog.get("text") or "")
        if dialog.get("found") or quantity_prompt_in_text(text):
            saw_prompt = True
            if dialog.get("loading") or "加载中" in text:
                ctx.record("quantity_prompt", "waiting_loading", {"stage": stage, "preview": text[:120]})
                time.sleep(1.2)
                continue
            # 连续两次读到非加载态再点，避免点到半开弹窗
            time.sleep(0.8)
            dialog2 = read_top_quantity_dialog(agent)
            text2 = str(dialog2.get("text") or text)
            if dialog2.get("loading") or "加载中" in text2:
                continue
            text = text2 or text
            update_expected_count_from_quantity_prompt(text, ctx)
            # Prefer CDP mouse click on topmost confirm (trusted gesture)
            if not click_quantity_confirm_export(agent, ctx):
                # retry once after short wait
                time.sleep(1.5)
                if not click_quantity_confirm_export(agent, ctx):
                    agent.screenshot(ctx, f"quantity_confirm_failed_{stage}.png")
                    raise RuntimeError(f"{stage}: 额度弹窗出现但点不到「继续导出」。")
            if stage == "generation":
                ctx.generation_quota_confirmed = True
            else:
                ctx.initial_quota_confirmed = True
            ctx.quantity_prompt_confirm_attempts += 1
            ctx.write_checkpoint()
            ctx.record(
                "quantity_prompt",
                "confirmed_after_ready",
                {
                    "stage": stage,
                    "quota_count": ctx.quota_export_count,
                    "expected_export_count": ctx.expected_export_count,
                    "dialog_source": dialog.get("source"),
                    "dialog_preview": text[:240],
                },
            )
            time.sleep(4)
            # Ensure prompt closed
            for _ in range(5):
                after = read_top_quantity_dialog(agent)
                if not after.get("found") and not quantity_prompt_in_text(str(after.get("text") or "")):
                    break
                close_known_quantity_prompt(agent, ctx)
                time.sleep(1)
            return True
        time.sleep(0.6)
    if saw_prompt:
        raise RuntimeError(f"{stage}: 额度弹窗一直加载中或无法确认。")
    return False


def submit_generate_on_target_row(agent: CDP, ctx: RunContext, row: dict) -> None:
    """点一次「生成」；有额度弹窗则继续导出；之后只刷新，绝不重复生成。"""
    row_index = row.get("index")
    actions = row.get("actions") or []
    refreshed = latest_export_record_state(agent, ctx)
    row2 = select_target_export_row((refreshed or {}).get("rows") or [], ctx) or row
    actions2 = row2.get("actions") or actions
    gen = next((a for a in actions2 if a.get("text") == "生成"), None)
    js_clicked = js_click_action_in_export_row(agent, ctx, row2.get("index", row_index), "生成")
    if gen and gen.get("x") is not None:
        agent.click(gen["x"], gen["y"], wait=0.3)
        ctx.record("export_record_generate", "cdp_mouse_after_js", gen)
    elif not js_clicked.get("ok"):
        raise RuntimeError("目标行操作列为「生成」但点击失败。")
    else:
        ctx.record("export_record_generate", "js_click_only", js_clicked)

    # 额度弹窗（多数账号额度足时不会出现）
    deadline = time.time() + 12
    while time.time() < deadline:
        dialog = read_top_quantity_dialog(agent)
        if dialog.get("found") or quantity_prompt_in_text(str(dialog.get("text") or "")):
            wait_and_confirm_quantity_prompt(agent, ctx, "generation", max_wait=20)
            break
        acts = target_row_action_set(
            select_target_export_row(
                ((latest_export_record_state(agent, ctx) or {}).get("rows") or []), ctx
            )
        )
        # 操作列已变为刷新/下载 → 生成已受理
        if "下载" in acts or ("刷新" in acts and "生成" not in acts):
            ctx.record("export_record_generate", "action_changed_after_generate", {"actions": sorted(acts)})
            break
        time.sleep(0.8)
    else:
        wait_and_confirm_quantity_prompt(agent, ctx, "generation", max_wait=3)

    body = combined_page_text(page_state(agent))
    if "已订阅成功" in body:
        ctx.record("export_record", "dismiss_subscribe_success_not_export_ok", {})
        agent.press_escape(wait=1)
        time.sleep(1)

    # 提交后：若操作列已是「刷新」则点行内刷新；否则仅 DOM 重读（禁止整页 reload 当刷新）
    refresh_export_record_page(agent, ctx, "after_generate", wait_seconds=15)
    ctx.record(
        "export_record_generate",
        "submitted_follow_op_column_thereafter",
        {"quota_count": ctx.quota_export_count, "generation_attempts": ctx.generation_attempts},
    )


def update_expected_count_from_quantity_prompt(text: str, ctx: RunContext) -> int | None:
    """Update expected count from top quota dialog; lock target row count after this."""
    remaining = extract_quota_remaining(text)
    if remaining is not None:
        ctx.quota_remaining = remaining
        ctx.record(
            "quota_remaining",
            "parsed",
            {"quota_remaining": remaining, "preview": (text or "")[:160]},
        )
        if remaining == 0:
            ctx.record(
                "quota_remaining",
                "zero_detected",
                {
                    "note": "剩余0不中断：继续导出后仍走 生成中/校验中→刷新→已生成→下载；"
                    "150为导出条数窗口上限，非禁止生成"
                },
            )

    quota_count = extract_quota_export_count(text)
    if quota_count:
        ctx.quota_export_count = quota_count
        ctx.expected_export_count = quota_count
        # After quota popup, target row may show truncated count — prefer quota as lock count.
        if ctx.target_record_count is not None and ctx.target_record_count != quota_count:
            ctx.record(
                "export_record_target_row",
                "count_aligned_to_quota",
                {"old_count": ctx.target_record_count, "quota_count": quota_count},
            )
            ctx.target_record_count = quota_count
        ctx.record(
            "expected_export_count",
            "updated_from_quota_prompt",
            {
                "filter_result_count": ctx.filter_result_count,
                "quota_export_count": quota_count,
                "quota_remaining": ctx.quota_remaining,
                "expected_export_count": ctx.expected_export_count,
            },
        )
    return quota_count


def generation_timeout_seconds(count: int | None) -> int:
    """Wait budget for 生成中/校验中/待生成 after submit, by export row count."""
    n = int(count or 0)
    if n <= 0:
        n = 200  # unknown count → mid budget
    if n <= 100:
        return 180  # ≤100 条最多 3 分钟
    if n <= 500:
        return 900  # 100-500 条 10-15 分钟，取 15
    return 1200  # 近 1000 条最多 20 分钟


def target_row_action_set(row: dict | None) -> set[str]:
    """Normalized action-column button texts on the locked target row only."""
    if not row:
        return set()
    return {str(a.get("text") or "").strip() for a in (row.get("actions") or []) if str(a.get("text") or "").strip()}


def target_row_status(row: dict | None) -> str:
    """Prefer dedicated 状态 cell; fallback to free-text guess."""
    if not row:
        return ""
    return str(row.get("status") or "").strip()


def target_row_is_terminal_fail(row: dict | None) -> bool:
    """状态失败/已过期 / 0 条且无下载 → 失败停止。"""
    if not row:
        return False
    text = str(row.get("text") or "")
    status = target_row_status(row)
    actions = target_row_action_set(row)
    # 状态栏含失败类 → 停止（网页可能新增「生成失败」「校验失败」等，用包含匹配）
    fail_tokens = ("已过期", "失败", "作废", "取消")
    if any(tok in status for tok in fail_tokens) or any(tok in text for tok in ("已过期", "生成失败", "导出失败", "校验失败")):
        # 若操作列仍有「下载」则以可下载为准（极少见）
        if "下载" not in actions:
            return True
    if row.get("count") == 0 and "下载" not in actions:
        return True
    return False


# 操作栏「可点击」文案（非穷举：未知短按钮也可尝试；这些是常见优先级锚点）
ACTION_DOWNLOAD = "下载"
ACTION_GENERATE = "生成"
ACTION_REFRESH = "刷新"
ACTION_PROGRESS_HINTS = ("刷新", "校验", "重试", "继续", "等待", "排队", "处理")  # 进度类：照做，不枚举状态
ACTION_IGNORE = {"-", "—", "无", "操作"}

def combined_page_text(state: dict | None):
    return ((state or {}).get("body", "") or "") + " " + " ".join(
        d.get("text", "") for d in (state or {}).get("dialogs", [])
    )


def close_known_quantity_prompt(agent: CDP, ctx: RunContext):
    """Only close the known export quota prompt after the export row is already ready behind it."""
    closed = click_candidate(
        agent,
        ctx,
        r"""
        (() => {
          const markers = ['导出数量提示', '是否继续导出', '访问额度不足', '剩余额度', '可导出部分'];
          const visible = el => {
            const r = el.getBoundingClientRect();
            const s = getComputedStyle(el);
            return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
          };
          const hasMarker = el => markers.some(m => (el.textContent || '').includes(m));
          const wrappers = Array.from(document.querySelectorAll(
            '.layui-layer, .layui-layer-dialog, .layui-layer-page, .el-message-box__wrapper, .el-dialog__wrapper, .el-message-box, .el-dialog'
          )).filter(el => visible(el) && hasMarker(el));
          const cands = [];
          for (const wrapper of wrappers) {
            const wrapperRect = wrapper.getBoundingClientRect();
            for (const el of wrapper.querySelectorAll(
              '.layui-layer-close, .el-message-box__close, .el-dialog__headerbtn, [aria-label="Close"], [aria-label="close"], i.el-icon-close, button[aria-label="Close"]'
            )) {
              const r = el.getBoundingClientRect();
              const s = getComputedStyle(el);
              if (r.width <= 0 || r.height <= 0 || s.visibility === 'hidden' || s.display === 'none') continue;
              cands.push({
                x: Math.round(r.x + r.width / 2),
                y: Math.round(r.y + r.height / 2),
                left: Math.round(r.x),
                top: Math.round(r.y),
                text: (el.textContent || '').replace(/\s+/g, ' ').trim() || 'close',
                tag: el.tagName,
                cls: String(el.className || '').slice(0, 80),
                wrapperTop: Math.round(wrapperRect.y),
                z: Number(getComputedStyle(wrapper).zIndex) || 0
              });
            }
          }
          cands.sort((a, b) => b.z - a.z || b.top - a.top || b.left - a.left);
          return {found: cands.length > 0, candidates: cands.slice(0, 8)};
        })()
        """,
        "quantity_prompt_close",
        wait=1,
    )
    if closed:
        return True
    ctx.record("quantity_prompt_close", "press_escape_fallback", {})
    agent.press_escape(wait=1)
    return not quantity_prompt_in_text(combined_page_text(page_state(agent)))


def handle_quantity_prompt_during_export_record(agent: CDP, ctx: RunContext, elapsed: float, confirm_attempts: int):
    """Confirm visible quota prompts with a bounded retry budget.

    千里马在“批量导出”提交后、导出记录页“生成”后都可能弹出同名提示。
    只确认一次会导致提示框残留或状态卡在“待生成”，因此这里按可见弹窗处理，
    但设置总上限，避免无限消耗额度。
    """
    state = page_state(agent)
    text = combined_page_text(state)
    update_expected_count_from_quantity_prompt(text, ctx)

    if ctx.generation_submitted:
        already_confirmed = ctx.generation_quota_confirmed
        prompt_stage = "generation"
    else:
        already_confirmed = ctx.initial_quota_confirmed
        prompt_stage = "initial"

    if ctx.quantity_prompt_confirm_attempts >= 4:
        ctx.record(
            "quantity_prompt",
            "confirm_limit_reached",
            {"elapsed": round(elapsed), "stage": prompt_stage, "attempts": ctx.quantity_prompt_confirm_attempts},
        )
    else:
        if not click_quantity_confirm_export(agent, ctx):
            raise RuntimeError("导出记录页出现导出数量提示，但未找到“导出/继续导出”按钮。")
        if prompt_stage == "generation":
            ctx.generation_quota_confirmed = True
        else:
            ctx.initial_quota_confirmed = True
        confirm_attempts += 1
        ctx.quantity_prompt_confirm_attempts += 1
        ctx.record(
            "quantity_prompt",
            "confirmed_visible_prompt",
            {
                "elapsed": round(elapsed),
                "stage": prompt_stage,
                "attempt": confirm_attempts,
                "total_attempts": ctx.quantity_prompt_confirm_attempts,
                "already_confirmed": already_confirmed,
            },
        )
        agent.screenshot(ctx, f"after_export_record_quantity_continue_{confirm_attempts}.png")
        time.sleep(5)

    state_after = page_state(agent)
    if not quantity_prompt_visible(agent) and not quantity_prompt_in_text(combined_page_text(state_after)):
        return True, confirm_attempts

    rows_state = latest_export_record_state(agent, ctx)
    target = select_target_export_row((rows_state or {}).get("rows") or [], ctx)
    action_texts = {item.get("text") for item in ((target or {}).get("actions") or [])}
    if "下载" in action_texts or (target or {}).get("status") == "已生成":
        ctx.record("quantity_prompt", "closing_ready_prompt", {"elapsed": round(elapsed)})
        close_known_quantity_prompt(agent, ctx)
    else:
        # 确认后不整页 reload；关闭可见弹窗后由主循环读操作列推进
        ctx.record("quantity_prompt", "dismiss_after_confirmed_no_page_reload", {"elapsed": round(elapsed)})
        close_known_quantity_prompt(agent, ctx)
        time.sleep(2)
    return True, confirm_attempts

def page_state(agent: CDP):
    return agent.eval(
        r"""
        (() => ({
          url: location.href,
          body: (document.body.innerText || '').replace(/\s+/g,' ').slice(-1200),
          dialogs: Array.from(document.querySelectorAll('.el-dialog__wrapper, .el-message-box__wrapper, .el-dialog, .el-message-box')).map(el => {
            const r = el.getBoundingClientRect();
            const cs = getComputedStyle(el);
            const visible = r.width > 8 && r.height > 8 && cs.display !== 'none' && cs.visibility !== 'hidden' && Number(cs.opacity || '1') > 0.05;
            return {
              text:(el.textContent||'').replace(/\s+/g,' ').trim().slice(0,240),
              x:Math.round(r.x), y:Math.round(r.y),
              w:Math.round(r.width), h:Math.round(r.height),
              cls:String(el.className||''),
              style:el.getAttribute('style')||'',
              visible
            };
          }).filter(o => o.visible)
        }))()
        """
    )


def capture_expected_export_count(agent: CDP, ctx: RunContext):
    state = agent.eval(
        r"""
        (() => {
          const body = document.body.innerText || '';
          const patterns = [
            /共\s*(\d{1,5})\s*条数据/,
            /共\s*(\d{1,5})\s*条/,
            /全部搜索结果\s*[（(]?(\d{1,5})\s*条/
          ];
          for (const re of patterns) {
            const m = body.match(re);
            if (m) return {count: Number(m[1]), source: m[0]};
          }
          return {count: null, source: ''};
        })()
        """
    ) or {}
    count = state.get("count") if isinstance(state, dict) else None
    if isinstance(count, int) and count > 0:
        ctx.filter_result_count = count
        ctx.expected_export_count = count
    ctx.record("expected_export_count", "captured", state)
    return ctx.expected_export_count


def trigger_export(agent: CDP, ctx: RunContext):
    wait_for_export_ready(agent, ctx)
    capture_expected_export_count(agent, ctx)
    agent.screenshot(ctx, "before_export.png")
    ctx.record(
        "table_select_all",
        "skipped",
        {"reason": "Select all only covers the current page; use all search results in the export dialog."},
    )
    dialog_opened = False
    for attempt in range(1, 4):
        if not click_batch_export(agent, ctx):
            ctx.record("batch_export", "retry_not_clickable", {"attempt": attempt})
            wait_for_export_ready(agent, ctx)
            continue
        agent.screenshot(ctx, f"after_batch_export_attempt_{attempt}.png")
        state = page_state(agent) or {}
        visible_text = (state.get("body") or "") + " " + " ".join(
            d.get("text", "") for d in state.get("dialogs", [])
        )
        if "\u5bfc\u51fa\u914d\u7f6e" in visible_text and "\u5bfc\u51fa\u8303\u56f4" in visible_text:
            dialog_opened = True
            break
        ctx.record("batch_export", "dialog_not_open_retry", {"attempt": attempt, "state": state})
        wait_for_export_ready(agent, ctx)
    if not dialog_opened:
        agent.screenshot(ctx, "export_dialog_not_visible.png")
        raise RuntimeError("Batch export was clicked but the export configuration dialog did not open.")
    ensure_export_dialog_options(agent, ctx)
    if not click_visible_export_button(agent, ctx, "main_dialog_export"):
        raise RuntimeError("未找到导出配置弹窗内的红色“导出”按钮。")
    ctx.export_confirm_clicked_at = datetime.now()
    agent.screenshot(ctx, "after_main_dialog_export.png")

    state = page_state(agent)
    ctx.record("after_main_dialog_export", "state", state)
    text = (state or {}).get("body", "") + " ".join(d.get("text", "") for d in (state or {}).get("dialogs", []))
    if any(s in text for s in ("导出数量提示", "是否继续导出", "访问额度不足", "剩余额度", "可导出部分")):
        update_expected_count_from_quantity_prompt(text, ctx)
        if not click_quantity_confirm_export(agent, ctx):
            raise RuntimeError("出现导出数量提示，但未找到二次确认“导出/继续导出”按钮。")
        ctx.export_confirm_clicked_at = datetime.now()
        ctx.initial_quota_confirmed = True
        agent.screenshot(ctx, "after_quantity_confirm_export.png")


def click_export_record_nav(agent: CDP, ctx: RunContext, *, step_name: str = "export_record_nav"):
    """Click '我的导出记录' / '导出记录' navigation.

    Prefers exact '我的导出记录', then short '导出记录' in menu/sidebar/<a href>.
    """
    return click_candidate(
        agent,
        ctx,
        r"""
        (() => {
          const cands = [];
          for (const el of document.querySelectorAll('a, span, div, li, button, p, em, i')) {
            const raw = (el.textContent || '').replace(/\s+/g,' ').trim();
            // 叶子节点：自身文本（不含深层子孙重复）
            let own = '';
            try {
              own = [...el.childNodes].filter(n => n.nodeType === 3)
                .map(n => (n.textContent||'').replace(/\s+/g,' ').trim()).join('');
            } catch (e) {}
            const text = (own || raw);
            if (!text.includes('导出记录')) continue;
            if (text.length > 24) continue;
            const r = el.getBoundingClientRect();
            if (r.width <= 0 || r.height <= 0 || r.width > 400 || r.height > 80) continue;
            const style = getComputedStyle(el);
            if (style.visibility === 'hidden' || style.display === 'none' || Number(style.opacity) === 0) continue;
            const hasHref = !!(el.href || (el.closest && el.closest('a') && el.closest('a').href));
            const inMenu = !!(el.closest && el.closest(
              '[class*=menu], [class*=sidebar], [class*=nav], [class*=aside], .menu-name, .el-menu, .el-submenu'
            ));
            const exactMine = text === '我的导出记录' || text.replace(/\s/g,'') === '我的导出记录';
            const exactShort = text === '导出记录';
            cands.push({
              x: Math.round(r.x + r.width/2), y: Math.round(r.y + r.height/2),
              left: Math.round(r.x), top: Math.round(r.y),
              text, tag: el.tagName, cls: String(el.className||'').slice(0,80),
              hasHref, inMenu, exactMine, exactShort
            });
          }
          cands.sort((a,b) =>
            (Number(b.exactMine)-Number(a.exactMine)) ||
            (Number(b.exactShort)-Number(a.exactShort)) ||
            (Number(b.inMenu)-Number(a.inMenu)) ||
            (Number(b.hasHref)-Number(a.hasHref)) ||
            a.top - b.top || a.left - b.left
          );
          return {found: cands.length > 0, candidates: cands.slice(0, 8)};
        })()
        """,
        step_name,
        wait=3,
    )


def click_my_focus_menu(agent: CDP, ctx: RunContext) -> bool:
    """先点「我的关注」展开子菜单，便于露出「导出记录」。"""
    return click_candidate(
        agent,
        ctx,
        r"""
        (() => {
          const cands = [];
          for (const el of document.querySelectorAll('a, span, div, li, button, p')) {
            let own = '';
            try {
              own = [...el.childNodes].filter(n => n.nodeType === 3)
                .map(n => (n.textContent||'').replace(/\s+/g,' ').trim()).join('');
            } catch (e) {}
            const text = (own || (el.textContent || '').replace(/\s+/g,' ').trim());
            if (text !== '我的关注' && text !== '我的关注 ') continue;
            if (text.length > 12) continue;
            const r = el.getBoundingClientRect();
            if (r.width <= 0 || r.height <= 0) continue;
            const style = getComputedStyle(el);
            if (style.visibility === 'hidden' || style.display === 'none') continue;
            cands.push({
              x: Math.round(r.x + r.width/2), y: Math.round(r.y + r.height/2),
              left: Math.round(r.x), top: Math.round(r.y),
              text, tag: el.tagName, cls: String(el.className||'').slice(0,80)
            });
          }
          cands.sort((a,b) => a.top - b.top || a.left - b.left);
          return {found: cands.length > 0, candidates: cands.slice(0, 5)};
        })()
        """,
        "my_focus_menu",
        wait=2,
    )


def _is_export_record_page_body(body: str) -> bool:
    if "导出记录" not in (body or ""):
        return False
    content_markers = ["下载", "生成中", "待生成", "校验中", "操作", "导出日期", "导出条数", "生成"]
    return any(marker in body for marker in content_markers)


def _poll_export_record_page(agent: CDP, ctx: RunContext, status: str, rounds: int = 12) -> bool:
    for _ in range(rounds):
        state = page_state(agent)
        body = (state or {}).get("body", "")
        if _is_export_record_page_body(body):
            ctx.record("export_record_page", status, {"url": (state or {}).get("url", "")})
            return True
        time.sleep(2)
    return False


def wait_for_export_record_page(
    agent: CDP,
    ctx: RunContext,
    subscribe_url: str,
    timeout=300,
    *,
    cdp_port: int = 9222,
    config_path: str | None = None,
) -> CDP:
    """导出确认后进入导出记录页（P1-01 加固）。返回可能重附着的 agent。

    策略（由快到稳）：
    1) 短等自然跳转（约 8s）
    2) 站内点「导出记录 / 我的导出记录」
    3) 先点「我的关注」再点子项「导出记录」
    4) 回 VIP 首页后再做 2/3
    5) 直达已验证入口 URL
    6) 若登录失效：自动登录后再走 5）
    """
    start = time.time()
    natural_wait = min(8, timeout)

    while time.time() - start < natural_wait:
        state = page_state(agent)
        body = (state or {}).get("body", "")
        if _is_export_record_page_body(body):
            ctx.record("export_record_page", "found", {"url": (state or {}).get("url", "")})
            return agent
        time.sleep(2)

    ctx.record("export_record_page", "natural_transition_timeout", {"elapsed": round(time.time() - start)})
    agent.screenshot(ctx, "export_record_page_not_reached_trying_nav.png")

    # 2) 直接点子菜单
    if click_export_record_nav(agent, ctx, step_name="export_record_nav_direct"):
        if _poll_export_record_page(agent, ctx, "found_after_nav_click"):
            return agent

    # 3) 展开「我的关注」再点导出记录
    if click_my_focus_menu(agent, ctx):
        ctx.record("export_record_page", "expanded_my_focus", {})
        time.sleep(1)
        if click_export_record_nav(agent, ctx, step_name="export_record_nav_after_focus"):
            if _poll_export_record_page(agent, ctx, "found_after_focus_nav"):
                return agent

    # 4) 回 VIP 根再导航（订阅页有时菜单折叠/遮罩）
    ctx.record("export_record_page", "retry_from_vip_root", {})
    agent.navigate("https://vip.qianlima.com/", wait=3)
    dismiss_popups(agent, ctx)
    if click_my_focus_menu(agent, ctx):
        time.sleep(1)
    if click_export_record_nav(agent, ctx, step_name="export_record_nav_from_vip_root"):
        if _poll_export_record_page(agent, ctx, "found_after_vip_root_nav"):
            return agent

    # 5) URL 兜底（保留；目标行仍由时间+条数锁定，非下历史）
    def _open_entry_url(status: str) -> bool:
        ctx.record(
            "export_record_page",
            "entry_url_fallback_start",
            {
                "url": EXPORT_RECORD_ENTRY_URL,
                "reason": status,
                "elapsed": round(time.time() - start),
            },
        )
        agent.navigate(EXPORT_RECORD_ENTRY_URL, wait=5)
        return _poll_export_record_page(agent, ctx, "found_after_entry_url", rounds=15)

    if _open_entry_url("natural/nav/my-focus/vip-root all failed; open verified export-record entry"):
        return agent

    # 6) 登录失效时自动登录再进入口
    login_st = read_login_state(agent)
    if not login_st.get("logged") or login_st.get("onLoginPage") or login_st.get("loginGate"):
        ctx.record("export_record_page", "login_recover_before_entry_url", login_st)
        try:
            agent.close()
        except Exception:
            pass
        recover_login_via_auto_login(ctx, cdp_port, config_path)
        agent = attach_qianlima(cdp_port, subscribe_url, ctx)
        if _open_entry_url("after_auto_login open export-record entry"):
            return agent

    agent.screenshot(ctx, "export_record_page_final_failure.png")
    raise RuntimeError(
        "点击导出确认后未进入导出记录页；自然跳转、站内导航、我的关注展开、VIP 首页导航、"
        "已验证导出记录入口、自动登录恢复后均失败。"
    )


def click_download_on_export_record(agent: CDP, ctx: RunContext):
    return click_candidate(
        agent,
        ctx,
        r"""
        (() => {
          const cands = [];
          for (const el of document.querySelectorAll('span, a, button')) {
            const text = (el.textContent || '').replace(/\s+/g,' ').trim();
            const r = el.getBoundingClientRect();
            if (text !== '下载' || r.width <= 0 || r.height <= 0) continue;
            cands.push({x:Math.round(r.x+r.width/2), y:Math.round(r.y+r.height/2), left:Math.round(r.x), top:Math.round(r.y), text, tag:el.tagName, cls:String(el.className||'').slice(0,80)});
          }
          cands.sort((a,b) => a.top-b.top || a.left-b.left);
          return {found:cands.length>0, candidates:cands.slice(0,8)};
        })()
        """,
        "export_record_download",
        wait=2,
    )


def parse_export_record_time(text: str):
    matches = re.findall(r"20\d{2}[-/]\d{1,2}[-/]\d{1,2}\s+\d{1,2}:\d{2}(?::\d{2})?", text or "")
    if not matches:
        return None
    raw = matches[0].replace("/", "-")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    return None


def latest_export_record_state(agent: CDP, ctx: RunContext):
    """Read export-record table rows: prefer 状态/操作 cells; actions from op column buttons.

    Live columns (2026-07-15): 导出日期 | 导出条数 | 导出格式 | 状态 | 操作 | 备注
    Observed pairs: 待生成→生成；生成中→刷新；已生成→下载；已过期→-
    """
    state = agent.eval(
        r"""
        (() => {
          const rows = [];
          for (const row of document.querySelectorAll('.el-table__body-wrapper tbody tr, tbody tr')) {
            const rr = row.getBoundingClientRect();
            const text = (row.textContent || '').replace(/\s+/g, ' ').trim();
            if (!text || rr.width <= 0 || rr.height <= 0) continue;
            if (!/20\d{2}[-/]\d{1,2}[-/]\d{1,2}\s+\d{1,2}:\d{2}/.test(text)) continue;
            const rowIndex = rows.length;
            const cells = [...row.querySelectorAll('td')].map(td => (td.innerText || '').replace(/\s+/g, ' ').trim());
            // Prefer dedicated columns when table has ≥5 cells
            let statusCell = cells.length >= 5 ? (cells[3] || '') : '';
            let opCell = cells.length >= 5 ? (cells[4] || '') : '';
            const actions = [];
            const seenAct = new Set();
            const pushAct = (t, el) => {
              const key = (t || '').trim();
              if (!key || key.length > 16 || seenAct.has(key)) return;
              if (key === '-' || key === '—') return;
              const r = el ? el.getBoundingClientRect() : {x:0,y:0,width:0,height:0};
              const cls = el ? String(el.className || '') : '';
              seenAct.add(key);
              actions.push({
                text: key,
                x: Math.round(r.x + r.width/2),
                y: Math.round(r.y + r.height/2),
                left: Math.round(r.x),
                top: Math.round(r.y),
                tag: el ? el.tagName : '',
                cls: cls.slice(0, 80)
              });
            };
            // Operation column buttons first (cu-btn etc.)
            const opRoot = cells.length >= 5 ? row.querySelectorAll('td')[4] : row;
            const scanRoot = opRoot || row;
            for (const el of scanRoot.querySelectorAll('button, span, a')) {
              const t = (el.textContent || '').replace(/\s+/g, ' ').trim();
              const r = el.getBoundingClientRect();
              const cls = String(el.className || '');
              const looks = /下载|刷新|生成|校验|重试|继续|取消/.test(t) || /cu-btn|sure-btn|el-button/.test(cls);
              if (looks && r.width > 0 && r.height > 0 && t.length <= 16) pushAct(t, el);
            }
            // If op cell is plain text like 生成/刷新/下载 without button match
            if (opCell && opCell.length <= 16 && !seenAct.has(opCell) && opCell !== '-') {
              pushAct(opCell, null);
            }
            // Status: cell first, else free-text guess (do not invent states)
            let status = (statusCell || '').trim();
            if (!status) {
              if (text.includes('已过期')) status = '已过期';
              else if (text.includes('待生成')) status = '待生成';
              else if (text.includes('校验中')) status = '校验中';
              else if (text.includes('生成中')) status = '生成中';
              else if (text.includes('已生成') || text.includes('已完成')) status = '已生成';
              else if (/失败/.test(text)) status = '失败';
            }
            const countMatch = text.match(/\s(\d+)\s+excel/i);
            const countFromCell = cells.length >= 2 && /^\d+$/.test(cells[1] || '') ? Number(cells[1]) : null;
            rows.push({
              index: rowIndex,
              text,
              cells,
              top: Math.round(rr.y),
              count: countFromCell != null ? countFromCell : (countMatch ? Number(countMatch[1]) : null),
              status,
              op_cell: opCell,
              actions
            });
          }
          rows.sort((a,b) => a.top - b.top);
          return {url: location.href, rows: rows.slice(0, 12)};
        })()
        """
    )
    ctx.record("export_record_rows", "state", state)
    return state


def select_target_export_row(rows: list[dict], ctx: RunContext) -> dict | None:
    """Select and lock the row created by this run; readiness never changes identity."""
    if not rows:
        return None

    candidates = []
    for pos, row in enumerate(rows):
        record_time = parse_export_record_time(row.get("text", ""))
        if not record_time:
            continue

        signed_time_delta = None
        if ctx.export_confirm_clicked_at:
            signed_time_delta = (record_time - ctx.export_confirm_clicked_at).total_seconds()
            # Record timestamps have minute precision. Allow one minute before the click,
            # but never accept old historical rows or unrelated later exports.
            if signed_time_delta < -90 or signed_time_delta > 900:
                continue

        count = row.get("count")
        count_delta = None
        if ctx.expected_export_count and isinstance(count, int):
            count_delta = abs(count - ctx.expected_export_count)

        if ctx.target_record_time is not None:
            # 服务端记录时间常按分钟取整，生成提交后可能漂移 1–5 分钟；
            # 旧逻辑 same_minute≤60s 会在 15:52 锁 → 行显示 15:54 时永久 not_found。
            time_window = 300  # 5 分钟
            same_window = abs((record_time - ctx.target_record_time).total_seconds()) <= time_window
            count_match = (
                count is not None
                and ctx.target_record_count is not None
                and count == ctx.target_record_count
            )
            # 优先：时间窗内 + 条数一致；其次：时间窗内（允许条数漂移）
            if same_window and (count_match or ctx.target_record_count is None):
                if count is not None and ctx.target_record_count is not None and count != ctx.target_record_count:
                    ctx.record(
                        "export_record_target_row",
                        "count_drift_same_window",
                        {
                            "locked_count": ctx.target_record_count,
                            "current_count": count,
                            "row_text": row.get("text", "")[:260],
                            "time_window_seconds": time_window,
                        },
                    )
                    ctx.target_record_count = count
                # 漂移后回写时间，避免后续轮询再次偏出窗口
                if abs((record_time - ctx.target_record_time).total_seconds()) > 60:
                    ctx.target_record_time = record_time
                return row
            if same_window:
                # 时间窗内但条数不一致：仍可作为候选（生成后条数可能被额度截断）
                time_score = abs((record_time - ctx.target_record_time).total_seconds())
                count_score = count_delta if count_delta is not None else 1000
                candidates.append((time_score, count_score, pos, row, record_time, count_delta))
                continue
            # 时间窗外但条数精确匹配且靠近点击时刻：允许重锚
            if count_match and ctx.export_confirm_clicked_at:
                signed = (record_time - ctx.export_confirm_clicked_at).total_seconds()
                if -90 <= signed <= 900:
                    time_score = abs(signed)
                    candidates.append((time_score, 0, pos, row, record_time, 0))
            continue

        time_score = abs(signed_time_delta) if signed_time_delta is not None else 600
        count_score = count_delta if count_delta is not None else 1000
        candidates.append((time_score, count_score, pos, row, record_time, count_delta))

    if not candidates:
        # 生成已提交时：兜底锁到点击时刻附近的「生成中/校验中/已生成/刷新」行
        if ctx.generation_submitted and ctx.export_confirm_clicked_at:
            fallback = []
            for pos, row in enumerate(rows):
                record_time = parse_export_record_time(row.get("text", ""))
                if not record_time:
                    continue
                signed = (record_time - ctx.export_confirm_clicked_at).total_seconds()
                if signed < -90 or signed > 900:
                    continue
                status = str(row.get("status") or "")
                op = str(row.get("op_cell") or "")
                text = str(row.get("text") or "")
                busy = any(
                    k in (status + op + text)
                    for k in ("生成中", "校验中", "已生成", "刷新", "下载")
                )
                if not busy:
                    continue
                count = row.get("count")
                count_delta = (
                    abs(count - ctx.expected_export_count)
                    if ctx.expected_export_count and isinstance(count, int)
                    else 1000
                )
                # 偏好条数接近 + 时间接近 + 列表靠前
                fallback.append((abs(signed), count_delta, pos, row, record_time))
            if fallback:
                fallback.sort(key=lambda item: (item[0], item[1], item[2]))
                best_fb = fallback[0]
                row = best_fb[3]
                old_time = ctx.target_record_time
                old_count = ctx.target_record_count
                ctx.target_record_time = best_fb[4]
                ctx.target_record_count = row.get("count")
                ctx.record(
                    "export_record_target_row",
                    "relock_after_time_drift",
                    {
                        "old_time": old_time.strftime("%Y-%m-%d %H:%M:%S") if old_time else None,
                        "old_count": old_count,
                        "new_time": ctx.target_record_time.strftime("%Y-%m-%d %H:%M:%S"),
                        "new_count": ctx.target_record_count,
                        "row_text": row.get("text", "")[:260],
                    },
                )
                return row
        ctx.record(
            "export_record_target_row",
            "not_found",
            {
                "expected_export_count": ctx.expected_export_count,
                "quota_export_count": ctx.quota_export_count,
                "clicked_at": (
                    ctx.export_confirm_clicked_at.strftime("%Y-%m-%d %H:%M:%S")
                    if ctx.export_confirm_clicked_at else None
                ),
                "target_record_time": (
                    ctx.target_record_time.strftime("%Y-%m-%d %H:%M:%S")
                    if ctx.target_record_time else None
                ),
                "target_record_count": ctx.target_record_count,
            },
        )
        return None

    candidates.sort(key=lambda item: (item[0], item[1], item[2]))
    best = candidates[0]
    row = best[3]
    ctx.target_record_time = best[4]
    ctx.target_record_count = row.get("count")
    ctx.record(
        "export_record_target_row",
        "locked",
        {
            "row_index": row.get("index"),
            "row_text": row.get("text", "")[:260],
            "record_time": ctx.target_record_time.strftime("%Y-%m-%d %H:%M:%S"),
            "record_count": ctx.target_record_count,
            "time_delta_seconds": round(best[0]),
            "count_delta": best[5],
            "filter_result_count": ctx.filter_result_count,
            "quota_export_count": ctx.quota_export_count,
            "expected_export_count": ctx.expected_export_count,
        },
    )
    return row

def js_click_action_in_export_row(agent: CDP, ctx: RunContext, row_index: int | None, action_text: str) -> dict:
    """Click an export-record action by DOM row/action text, not by stale coordinates."""
    if row_index is None:
        return {"ok": False, "reason": "missing_row_index"}
    state = agent.eval(
        fr"""
        (() => {{
          const normalize = s => (s || '').replace(/\s+/g, '').trim();
          const wanted = normalize({json.dumps(action_text, ensure_ascii=False)});
          const rows = [...document.querySelectorAll('.el-table__body-wrapper tbody tr, tbody tr')]
            .filter(r => /20\d{{2}}[-/]\d{{1,2}}[-/]\d{{1,2}}\s+\d{{1,2}}:\d{{2}}/.test(r.textContent || ''));
          const row = rows[{json.dumps(row_index)}];
          if (!row) return {{ok:false, reason:'row_not_found', rowCount: rows.length}};
          row.scrollIntoView({{block:'center', inline:'center'}});
          const candidates = [...row.querySelectorAll('button, span, a, div')]
            .filter(el => normalize(el.textContent) === wanted);
          for (const el of candidates) {{
            const r = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            if (r.width <= 0 || r.height <= 0 || style.visibility === 'hidden' || style.display === 'none') continue;
            el.scrollIntoView({{block:'center', inline:'center'}});
            el.dispatchEvent(new MouseEvent('mouseover', {{bubbles:true}}));
            el.dispatchEvent(new MouseEvent('mousedown', {{bubbles:true}}));
            el.dispatchEvent(new MouseEvent('mouseup', {{bubbles:true}}));
            el.click();
            return {{
              ok:true,
              action:wanted,
              x:Math.round(r.x+r.width/2),
              y:Math.round(r.y+r.height/2),
              tag:el.tagName,
              cls:String(el.className||'').slice(0,80),
              rowText:normalize(row.textContent).slice(0,240)
            }};
          }}
          const fallback = [...row.querySelectorAll('button, span, a, div')]
            .filter(el => /cu-btn|sure-btn|el-button/.test(String(el.className || '')))
            .filter(el => {{
              const r = el.getBoundingClientRect();
              const style = getComputedStyle(el);
              return r.width > 0 && r.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
            }})[0];
          if (fallback) {{
            const r = fallback.getBoundingClientRect();
            fallback.scrollIntoView({{block:'center', inline:'center'}});
            fallback.dispatchEvent(new MouseEvent('mouseover', {{bubbles:true}}));
            fallback.dispatchEvent(new MouseEvent('mousedown', {{bubbles:true}}));
            fallback.dispatchEvent(new MouseEvent('mouseup', {{bubbles:true}}));
            fallback.click();
            return {{
              ok:true,
              action:wanted,
              fallback:true,
              fallbackText:normalize(fallback.textContent),
              x:Math.round(r.x+r.width/2),
              y:Math.round(r.y+r.height/2),
              tag:fallback.tagName,
              cls:String(fallback.className||'').slice(0,80),
              rowText:normalize(row.textContent).slice(0,240)
            }};
          }}
          return {{
            ok:false,
            reason:'action_not_found',
            wanted,
            rowText:normalize(row.textContent).slice(0,240),
            visibleTexts:[...row.querySelectorAll('button,span,a,div')].map(el=>normalize(el.textContent)).filter(Boolean).slice(0,20)
          }};
        }})()
        """
    )
    ctx.record("export_record_action_js", "clicked" if state.get("ok") else "not_clicked", state)
    return state


def is_export_record_url(url: str) -> bool:
    u = (url or "").lower()
    return "export" in u or "myfocus_getexport" in u or "导出" in (url or "")


def ensure_on_export_record_page(agent: CDP, ctx: RunContext, reason: str = "") -> None:
    """If navigated away (e.g. after 生成/订阅成功), re-enter export records via VIP menu."""
    state = page_state(agent)
    url = str((state or {}).get("url") or "")
    body = combined_page_text(state)
    if is_export_record_url(url) and ("导出记录" in body or "下载" in body or "生成" in body or "刷新" in body):
        return
    ctx.record("export_record_page", "left_page_reenter", {"url": url, "reason": reason})
    # 关掉「已订阅成功」等遮挡
    if "已订阅成功" in body or "我知道了" in body:
        agent.press_escape(wait=0.5)
        click_candidate(
            agent,
            ctx,
            r"""
            (() => {
              const cands=[];
              for (const el of document.querySelectorAll('button,span,div,a')) {
                const t=(el.textContent||'').replace(/\s+/g,'').trim();
                const r=el.getBoundingClientRect();
                if ((t==='我知道了'||t==='确定'||t==='关闭') && r.width>0 && r.height>0)
                  cands.push({x:Math.round(r.x+r.width/2),y:Math.round(r.y+r.height/2),text:t});
              }
              return {found:cands.length>0, candidates:cands.slice(0,3)};
            })()
            """,
            "dismiss_subscribe_ok",
            wait=1,
        )
    agent.navigate("https://vip.qianlima.com/", wait=3)
    if not click_export_record_nav(agent, ctx):
        # last resort known entry (still verify rows after)
        agent.navigate("https://vip.qianlima.com/myFocus_getExport.html", wait=4)
        ctx.record("export_record_page", "fallback_entry_url", {})
    time.sleep(2)


def refresh_export_record_page(
    agent: CDP,
    ctx: RunContext,
    reason: str,
    wait_seconds: int = 15,
    row: dict | None = None,
) -> None:
    """推进目标行状态：优先点操作列「刷新」按钮（不是整页 F5）。

    - 有「刷新」→ js 点目标行刷新
    - 无「刷新」但人还在导出记录页 → 仅重新读 DOM（sleep），禁止 Page.reload 当「刷新」
    - 已离开导出页 → re-enter 导出记录入口
    """
    ensure_on_export_record_page(agent, ctx, reason=f"before_refresh:{reason}")
    state = latest_export_record_state(agent, ctx)
    if row is None:
        row = select_target_export_row((state or {}).get("rows") or [], ctx)
    actions = (row or {}).get("actions") or []
    refresh = next((a for a in actions if a.get("text") == ACTION_REFRESH), None)
    if row and refresh:
        clicked = js_click_action_in_export_row(agent, ctx, row.get("index"), ACTION_REFRESH)
        if not clicked.get("ok") and refresh.get("x") is not None:
            agent.click(refresh["x"], refresh["y"])
            ctx.record("export_record_refresh", "clicked_coordinate_fallback", refresh)
        else:
            ctx.record(
                "export_record_refresh",
                "clicked_row_refresh_btn",
                {"reason": reason, "row_index": row.get("index"), "status": row.get("status")},
            )
    else:
        st2 = page_state(agent)
        if not is_export_record_url(str((st2 or {}).get("url") or "")):
            ensure_on_export_record_page(agent, ctx, reason=f"left_page_reenter:{reason}")
            ctx.record("export_record_refresh", "reentered_export_page", {"reason": reason})
        else:
            # 无行内「刷新」：只重读表格，不做 Page.reload（用户明确：刷新≠刷新网页）
            ctx.record(
                "export_record_refresh",
                "dom_reread_only_no_page_reload",
                {
                    "reason": reason,
                    "actions": sorted(target_row_action_set(row)),
                    "status": target_row_status(row),
                },
            )
    time.sleep(max(8, int(wait_seconds)))


def click_latest_export_action(agent: CDP, ctx: RunContext):
    """目标法则（不枚举状态名）。

    唯一目标：本次目标行操作列出现「下载」并点成功。
    法则：
      1) 只动锁定的目标行，禁止历史行下载
      2) 状态栏用于理解/判死（失败、已过期）；动作以操作列为准
      3) 操作列写什么点什么：下载→下载；生成→默认一次；刷新/其它进度钮→照点
      4) 刷新=行内按钮，≠ Page.reload；进度不明默认重读/行内刷新（约10–20s）
      5) 额度弹窗继续导出；剩余0不 fail-fast（150=条数窗口）
      6) 页面新增「等待」等第5/6步：只要操作是刷新/进度钮即照做，无需改枚举表
    """
    ensure_on_export_record_page(agent, ctx, reason="before_action")
    state = latest_export_record_state(agent, ctx)
    rows = (state or {}).get("rows") or []
    if not rows:
        ensure_on_export_record_page(agent, ctx, reason="no_rows")
        refresh_export_record_page(agent, ctx, "no_rows", wait_seconds=12)
        return False

    row = select_target_export_row(rows, ctx)
    if not row:
        refresh_export_record_page(agent, ctx, "no_target_row", wait_seconds=12)
        return False
    record_time = parse_export_record_time(row.get("text", ""))
    if ctx.export_confirm_clicked_at and record_time:
        delta = abs((record_time - ctx.export_confirm_clicked_at).total_seconds())
        ctx.record(
            "export_record_target_time",
            "checked",
            {
                "record_time": record_time.strftime("%Y-%m-%d %H:%M:%S"),
                "clicked_at": ctx.export_confirm_clicked_at.strftime("%Y-%m-%d %H:%M:%S"),
                "delta_seconds": round(delta),
                "row_index": row.get("index"),
            },
        )
        if delta > 900:
            raise RuntimeError("未找到与本次导出时间匹配的导出记录，禁止下载旧记录。")

    actions = row.get("actions") or []
    action_texts = target_row_action_set(row)
    status = target_row_status(row) or ""
    count = row.get("count")
    wait_budget = generation_timeout_seconds(
        ctx.target_record_count or ctx.expected_export_count or count
    )
    progress_refresh_wait = 15  # 10–20s 一轮

    def waited_sec() -> float:
        if not ctx.generation_wait_started_at:
            return 0.0
        return (datetime.now() - ctx.generation_wait_started_at).total_seconds()

    def waited_too_long() -> bool:
        return waited_sec() >= wait_budget

    # 每步自查快照（状态栏 + 操作列 + 目标行文本）
    ctx.record(
        "goal_law_tick",
        "observe",
        {
            "goal": "target_row_download",
            "status": status,
            "actions": sorted(action_texts),
            "count": count,
            "row_preview": str(row.get("text") or "")[:200],
            "generation_submitted": ctx.generation_submitted,
            "generation_attempts": ctx.generation_attempts,
            "download_clicked": ctx.download_clicked,
            "quota_remaining": ctx.quota_remaining,
            "waited_sec": round(waited_sec()) if ctx.generation_wait_started_at else 0,
        },
    )

    # 真失败：失败/已过期（不是剩余0）
    if target_row_is_terminal_fail(row):
        agent.screenshot(ctx, "export_record_expired.png")
        raise RuntimeError(
            f"目标导出记录失败或已过期且无下载，停止。 status={status!r} count={count} "
            f"row={str(row.get('text',''))[:160]}"
        )

    if ctx.generation_submitted and waited_too_long() and ACTION_DOWNLOAD not in action_texts:
        agent.screenshot(ctx, "export_record_backend_timeout.png")
        raise RuntimeError(
            f"后台生成超时：操作列仍未出现「下载」（等待上限 {wait_budget}s，"
            f"条数≈{ctx.target_record_count or count}，status={status!r}，actions={sorted(action_texts)}）。"
        )

    # —— 法则3：操作列驱动（不 if 状态名枚举）——

    # A. 下载 → 点下载（目标达成）
    if ACTION_DOWNLOAD in action_texts:
        if ctx.download_clicked:
            ctx.record("export_record_download", "already_clicked_waiting_for_file", {})
            time.sleep(5)
            return False
        ctx.record(
            "export_record",
            "button_download",
            {"actions": sorted(action_texts), "count": count, "status": status, "law": "op=下载"},
        )
        clicked = js_click_action_in_export_row(agent, ctx, row.get("index"), ACTION_DOWNLOAD)
        if not clicked.get("ok"):
            download = next((a for a in actions if a.get("text") == ACTION_DOWNLOAD), None)
            if not download:
                raise RuntimeError("目标行操作列显示下载但点击失败。")
            agent.click(download["x"], download["y"])
            ctx.record("export_record_download", "clicked_coordinate_fallback", download)
        else:
            ctx.record("export_record_download", "clicked_latest", clicked)
        ctx.download_clicked = True
        try:
            agent.screenshot(ctx, f"after_download_click_{datetime.now().strftime('%H%M%S')}.png")
        except Exception:
            pass
        time.sleep(3)
        return True

    # B. 生成 → 默认一次；长时间未受理可再点一次
    if ACTION_GENERATE in action_texts:
        # 若同时有「刷新」，优先刷新（操作列同时给刷新时：进度优先于再生成）
        if ACTION_REFRESH in action_texts and ctx.generation_submitted:
            ctx.record(
                "export_record",
                "prefer_refresh_over_regenerate",
                {"status": status, "actions": sorted(action_texts), "law": "进度钮优先"},
            )
            refresh_export_record_page(
                agent, ctx, "refresh_over_generate", wait_seconds=progress_refresh_wait, row=row
            )
            return False
        allow_generate = not ctx.generation_submitted
        retry_generate = (
            ctx.generation_submitted
            and ctx.generation_attempts < 2
            and waited_sec() >= 45
            and ACTION_REFRESH not in action_texts
        )
        if allow_generate or retry_generate:
            now = datetime.now()
            ctx.generation_attempts = int(ctx.generation_attempts or 0) + 1
            ctx.generation_submitted = True
            ctx.last_generation_clicked_at = now
            if not ctx.generation_wait_started_at:
                ctx.generation_wait_started_at = now
            ctx.write_checkpoint()
            ctx.record(
                "export_record",
                "button_generate",
                {
                    "actions": sorted(action_texts),
                    "status": status,
                    "attempt": ctx.generation_attempts,
                    "retry": bool(retry_generate),
                    "law": "op=生成",
                },
            )
            try:
                agent.screenshot(ctx, f"before_generate_{datetime.now().strftime('%H%M%S')}.png")
            except Exception:
                pass
            submit_generate_on_target_row(agent, ctx, row)
            try:
                agent.screenshot(ctx, f"after_generate_{datetime.now().strftime('%H%M%S')}.png")
            except Exception:
                pass
            return False
        # 已提交、未到重试：不连点生成 → 刷新/重读
        ctx.record(
            "export_record",
            "after_generate_wait",
            {
                "status": status,
                "waited_sec": round(waited_sec()),
                "quota_remaining": ctx.quota_remaining,
                "law": "未到下载前默认刷新/重读",
            },
        )
        refresh_export_record_page(
            agent, ctx, "after_generate_wait", wait_seconds=progress_refresh_wait, row=row
        )
        return False

    # C. 刷新 / 校验 / 重试 / 继续… 等进度钮 → 照做（覆盖「等待」等新状态若按钮仍是刷新）
    for hint in ACTION_PROGRESS_HINTS:
        match = next((a for a in actions if hint in str(a.get("text") or "")), None)
        if not match:
            continue
        label = str(match.get("text") or hint)
        if label in ACTION_IGNORE:
            continue
        ctx.record(
            "export_record",
            "button_follow_op_column",
            {
                "click": label,
                "actions": sorted(action_texts),
                "status": status,
                "law": "op列进度钮照做",
            },
        )
        if label == ACTION_REFRESH or hint == ACTION_REFRESH:
            refresh_export_record_page(
                agent, ctx, "op_column_refresh", wait_seconds=progress_refresh_wait, row=row
            )
        else:
            clicked = js_click_action_in_export_row(agent, ctx, row.get("index"), label)
            if not clicked.get("ok") and match.get("x") is not None:
                agent.click(match["x"], match["y"])
            time.sleep(progress_refresh_wait)
        return False

    # D. 未知短操作文案：照做（页面改版新按钮）
    for a in actions:
        label = str(a.get("text") or "").strip()
        if not label or label in ACTION_IGNORE:
            continue
        if label in (ACTION_DOWNLOAD, ACTION_GENERATE):
            continue
        if any(x in label for x in ("失败", "删除", "详情", "备注")):
            continue
        if len(label) > 12:
            continue
        ctx.record(
            "export_record",
            "button_unknown_follow",
            {"click": label, "status": status, "actions": sorted(action_texts), "law": "未知按钮照做"},
        )
        clicked = js_click_action_in_export_row(agent, ctx, row.get("index"), label)
        if not clicked.get("ok") and a.get("x") is not None:
            agent.click(a["x"], a["y"])
        time.sleep(progress_refresh_wait)
        return False

    # E. 无可点：目标行 DOM 重读（默认等待），不整页 reload
    ctx.record(
        "export_record",
        "wait_reread_target_row",
        {
            "actions": sorted(action_texts),
            "status": status,
            "generation_submitted": ctx.generation_submitted,
            "waited_sec": round(waited_sec()),
            "law": "默认重读直至出现下载",
        },
    )
    refresh_export_record_page(agent, ctx, "wait_reread", wait_seconds=progress_refresh_wait, row=row)
    return False

def snapshot_downloads(ctx: RunContext):
    roots = [ctx.cdp_download_dir, ctx.download_dir, Path.home() / "Downloads"]
    files = []
    for root in roots:
        if not root.exists():
            continue
        for p in root.iterdir():
            if not p.is_file():
                continue
            name = p.name.lower()
            # 跳过临时文件和下载中的文件
            if name.endswith(".crdownload") or name.startswith("~$") or name.startswith("."):
                continue
            # 只接受Excel文件（xlsx/xls）或CDP目录中的任何文件（可能是zip格式的xlsx）
            is_excel = name.endswith(".xlsx") or name.endswith(".xls")
            is_in_cdp_dir = str(p).startswith(str(ctx.cdp_download_dir))
            if not is_excel and not is_in_cdp_dir:
                continue
            files.append({"path": str(p), "size": p.stat().st_size, "mtime": p.stat().st_mtime})
    files.sort(key=lambda x: x["mtime"], reverse=True)
    return files


def looks_like_qianlima_export_name(path: Path) -> bool:
    """文件名启发式：拦截 Downloads 里无关 xlsx（如网页结构化数据导出）。"""
    name = path.name
    name_l = name.lower()
    if name_l.endswith(".crdownload") or name.startswith("~$"):
        return False
    # 历史主线文件名：千里马信息导出_MM-DD.xlsx
    positive = (
        "千里马" in name
        or "信息导出" in name
        or "qianlima" in name_l
        or name_l.startswith("export")
        or "招标" in name
        or "标讯" in name
    )
    # 明确拒绝常见噪声（Chrome 扩展/站点爬虫产物）
    negative = (
        "structured_data" in name_l
        or "schema.org" in name_l
        or "pages_structured" in name_l
        or name_l.startswith("www.")
    )
    if negative:
        return False
    return positive


def pick_new_export_file(
    ctx: RunContext,
    before_paths: set[str],
    *,
    require_download_clicked: bool = True,
    min_size: int = 1000,
) -> Path | None:
    """只在「已点下载」后认新文件；优先 CDP/本 run 目录，再筛文件名。"""
    if require_download_clicked and not ctx.download_clicked:
        return None
    files = snapshot_downloads(ctx)
    candidates = []
    for item in files:
        if item["path"] in before_paths or item["size"] < min_size:
            continue
        p = Path(item["path"])
        in_cdp = str(p).startswith(str(ctx.cdp_download_dir))
        in_run_dir = str(p).startswith(str(ctx.download_dir))
        name_ok = looks_like_qianlima_export_name(p)
        # CDP / 本 run 目录：下载行为已定向，放宽文件名；用户 Downloads：必须像千里马导出
        if in_cdp or in_run_dir or name_ok:
            candidates.append({**item, "rank": 0 if in_cdp else (1 if in_run_dir else 2), "name_ok": name_ok})
        else:
            ctx.record(
                "download_file",
                "ignored_unlikely_name",
                {"path": str(p), "size": item["size"], "reason": "not_qianlima_export_name"},
            )
    if not candidates:
        return None
    candidates.sort(key=lambda x: (x["rank"], -x["mtime"]))
    chosen = candidates[0]
    ctx.record(
        "download_file",
        "found",
        {
            "path": chosen["path"],
            "size": chosen["size"],
            "mtime": chosen["mtime"],
            "download_clicked": ctx.download_clicked,
            "name_ok": chosen.get("name_ok"),
        },
    )
    return Path(chosen["path"])


def wait_for_download_file(agent: CDP, ctx: RunContext, before_paths: set[str], timeout: int | None = None):
    """Run the target export row 5-state machine until the new Excel file exists."""
    count_hint = ctx.target_record_count or ctx.expected_export_count or ctx.filter_result_count
    if timeout is None:
        # 进入记录页 + 生成等待 + 下载缓冲
        timeout = max(generation_timeout_seconds(count_hint) + 180, 600)
    start = time.time()
    quantity_confirm_attempts = 0
    last_heartbeat = 0.0
    ctx.record(
        "export_record_wait",
        "started",
        {
            "timeout_seconds": timeout,
            "generation_budget_seconds": generation_timeout_seconds(count_hint),
            "count_hint": count_hint,
        },
    )

    while time.time() - start < timeout:
        elapsed = time.time() - start
        if elapsed - last_heartbeat >= 15:
            gen_waited = None
            if ctx.generation_wait_started_at:
                gen_waited = round((datetime.now() - ctx.generation_wait_started_at).total_seconds())
            ctx.record(
                "phase1_heartbeat",
                "waiting",
                {
                    "elapsed_seconds": round(elapsed),
                    "timeout_seconds": timeout,
                    "target_record_time": (
                        ctx.target_record_time.strftime("%Y-%m-%d %H:%M:%S")
                        if ctx.target_record_time else None
                    ),
                    "target_record_count": ctx.target_record_count,
                    "generation_submitted": ctx.generation_submitted,
                    "generation_attempts": ctx.generation_attempts,
                    "generation_waited_seconds": gen_waited,
                    "quantity_prompt_confirm_attempts": ctx.quantity_prompt_confirm_attempts,
                    "download_clicked": ctx.download_clicked,
                },
            )
            last_heartbeat = elapsed
            ctx.write_checkpoint()

        state = page_state(agent)
        all_text = combined_page_text(state)
        url = str((state or {}).get("url") or "")

        # ③ 导出记录页 404：回 VIP 从站内菜单重进
        if "404" in all_text or "/404" in url.lower() or "页面不存在" in all_text:
            ctx.record("export_record_page", "got_404_reenter_via_menu", {"url": url})
            agent.screenshot(ctx, "export_record_404.png")
            agent.navigate("https://vip.qianlima.com/", wait=4)
            if not click_export_record_nav(agent, ctx):
                raise RuntimeError("导出记录页 404 后无法从 VIP 站内菜单重新进入导出记录。")
            time.sleep(3)
            continue

        if "已在其他设备登录" in all_text or "账号已在其他设备" in all_text:
            agent.screenshot(ctx, "account_logged_in_elsewhere.png")
            raise RuntimeError(
                "千里马账号已在其他设备登录（被异地登录踢出）。"
                "请确认未在别处同时使用同一账号后重试，不能继续等待下载。"
            )

        # ① 最上层额度弹窗（必须“可见”才处理；隐藏 dialog 文本不得劫持主循环）
        if quantity_prompt_visible(agent):
            if ctx.generation_submitted and ctx.generation_quota_confirmed:
                ctx.record("quantity_prompt", "already_confirmed_visible_dismiss", {})
                close_known_quantity_prompt(agent, ctx)
                time.sleep(1)
                # 不 continue 死循环：若关闭失败也交给操作列状态机
            elif ctx.quantity_prompt_confirm_attempts >= 2 and ctx.generation_submitted:
                raise RuntimeError(
                    "生成后额度弹窗已处理仍反复出现，停止以免重复提交多个导出任务。"
                )
            else:
                _, quantity_confirm_attempts = handle_quantity_prompt_during_export_record(
                    agent, ctx, elapsed, quantity_confirm_attempts
                )
                continue
        elif quantity_prompt_in_text(all_text) and not (
            ctx.generation_submitted and ctx.generation_quota_confirmed
        ):
            # 文本像弹窗但几何不可见：记一次日志后走操作列，避免 520 次空转
            ctx.record(
                "quantity_prompt",
                "text_match_but_not_visible_skip",
                {"preview": all_text[:160]},
            )

        # ② 「已订阅成功」不是导出成功
        if "已订阅成功" in all_text and "导出记录" not in all_text[:80]:
            ctx.record("export_record", "dismiss_subscribe_success_popup", {})
            agent.press_escape(wait=1)
            time.sleep(1)

        # 必须已点目标行「下载」后才认新文件，避免 Downloads 里无关 xlsx 假成功
        picked = pick_new_export_file(ctx, before_paths, require_download_clicked=True)
        if picked is not None:
            return picked

        click_latest_export_action(agent, ctx)
        time.sleep(2)

    agent.screenshot(ctx, "download_timeout.png")
    raise RuntimeError(
        f"等待下载超时（{timeout}s），未发现本次目标行的新 xlsx。"
        f" generation_submitted={ctx.generation_submitted}"
        f" download_clicked={ctx.download_clicked}"
        f" target_count={ctx.target_record_count}。"
    )


def validate_xlsx(path: Path, ctx: RunContext):
    if not path.exists() or path.stat().st_size <= 0:
        raise RuntimeError(f"下载文件不存在或为空: {path}")
    with zipfile.ZipFile(path) as zf:
        names = set(zf.namelist())
        ok = "xl/workbook.xml" in names and any(n.startswith("xl/worksheets/") for n in names)
    detail = {"path": str(path), "size": path.stat().st_size, "xlsx_structure_ok": ok}
    header_ok = False
    headers: list[str] = []
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        detail.update({"sheet": ws.title, "rows": ws.max_row, "cols": ws.max_column})
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first:
            headers = [str(c).strip() if c is not None else "" for c in first]
            detail["headers_preview"] = headers[:12]
            joined = " ".join(headers)
            # 千里马标讯导出典型表头：标题 + 基本信息 / 地区 / 发布时间 等
            header_ok = (
                ("标题" in headers or "标题" in joined)
                and (
                    "基本信息" in joined
                    or "发布时间" in joined
                    or "招标编号" in joined
                    or "地区" in joined
                    or "详情" in joined
                )
            )
        wb.close()
    except Exception as exc:
        detail["openpyxl_warning"] = str(exc)
    detail["header_ok"] = header_ok
    content_ok = ok and header_ok
    ctx.record("xlsx_validate", "ok" if content_ok else "failed", detail)
    if not ok:
        raise RuntimeError("xlsx 结构校验失败。")
    if not header_ok:
        raise RuntimeError(
            f"xlsx 表头不像千里马标讯导出（拒绝误认 Downloads 噪声文件）: {path.name}; "
            f"headers={headers[:8]}"
        )
    return detail


def apply_resume_context(ctx: RunContext, payload: dict, source: Path) -> None:
    def parse_dt(value):
        return datetime.fromisoformat(value) if value else None

    clicked_at = parse_dt(payload.get("export_confirm_clicked_at"))
    if not clicked_at:
        raise RuntimeError("恢复检查点缺少 export_confirm_clicked_at，禁止猜测或重新提交导出。")
    ctx.export_confirm_clicked_at = clicked_at
    ctx.expected_export_count = payload.get("expected_export_count")
    ctx.filter_result_count = payload.get("filter_result_count")
    ctx.quota_export_count = payload.get("quota_export_count")
    ctx.target_record_time = parse_dt(payload.get("target_record_time"))
    ctx.target_record_count = payload.get("target_record_count")
    ctx.generation_submitted = bool(payload.get("generation_submitted"))
    ctx.generation_attempts = int(payload.get("generation_attempts") or 0)
    ctx.last_generation_clicked_at = parse_dt(payload.get("last_generation_clicked_at"))
    ctx.generation_wait_started_at = parse_dt(payload.get("generation_wait_started_at"))
    if ctx.generation_submitted and not ctx.generation_wait_started_at:
        ctx.generation_wait_started_at = ctx.last_generation_clicked_at or datetime.now()
    ctx.quantity_prompt_confirm_attempts = int(payload.get("quantity_prompt_confirm_attempts") or 0)
    ctx.initial_quota_confirmed = bool(payload.get("initial_quota_confirmed"))
    ctx.generation_quota_confirmed = bool(payload.get("generation_quota_confirmed"))
    ctx.download_clicked = False
    ctx.record(
        "resume_context",
        "loaded",
        {
            "source": str(source),
            "original_run_id": payload.get("run_id"),
            "target_record_time": payload.get("target_record_time"),
            "target_record_count": payload.get("target_record_count"),
            "generation_submitted": ctx.generation_submitted,
            "generation_attempts": ctx.generation_attempts,
            "quantity_prompt_confirm_attempts": ctx.quantity_prompt_confirm_attempts,
        },
    )

def yesterday_today():
    today = datetime.now().date()
    return (today - timedelta(days=1)).isoformat(), today.isoformat()


def try_download_existing_generated_export(
    agent: CDP,
    ctx: RunContext,
    before_paths: set[str],
    *,
    prefer_count: int | None = 150,
) -> Path:
    """下载导出记录中「已生成+下载」行（不新建导出任务）。

    测试额度常固定 150 条循环：优先选 count==prefer_count 的已生成行，
    避免误点其它条数；无匹配则退回最新已生成行。
    """
    ensure_on_export_record_page(agent, ctx, reason="fallback_existing")
    state = latest_export_record_state(agent, ctx)
    rows = (state or {}).get("rows") or []
    candidates = []
    for row in rows:
        actions = target_row_action_set(row)
        status = str(row.get("status") or "")
        if "下载" in actions and ("已生成" in status or "已生成" in str(row.get("text") or "")):
            candidates.append(row)
    if not candidates:
        raise RuntimeError(
            "fallback 失败：导出记录中无「已生成+下载」行可用。"
            "额度 0 且无历史已生成文件，无法继续。"
        )
    preferred = []
    if prefer_count is not None:
        preferred = [r for r in candidates if int(r.get("count") or 0) == int(prefer_count)]
    # Prefer matching count (150 额度模式)，否则最新已生成
    row = preferred[0] if preferred else candidates[0]
    ctx.record(
        "fallback_existing_export",
        "selected_row",
        {
            "row": str(row.get("text") or "")[:200],
            "count": row.get("count"),
            "index": row.get("index"),
            "prefer_count": prefer_count,
            "preferred_hit": bool(preferred),
            "candidate_counts": [r.get("count") for r in candidates[:8]],
        },
    )
    clicked = js_click_action_in_export_row(agent, ctx, row.get("index"), "下载")
    if not clicked.get("ok"):
        download = next((a for a in (row.get("actions") or []) if a.get("text") == "下载"), None)
        if not download:
            raise RuntimeError("fallback：找到已生成行但无法点击下载。")
        agent.click(download["x"], download["y"])
        ctx.record("fallback_existing_export", "clicked_coordinate", download)
    else:
        ctx.record("fallback_existing_export", "clicked_js", clicked)
    ctx.download_clicked = True
    ctx.export_source = "fallback_existing_export"
    # wait for file (shorter timeout)；与主路径同一套认文件规则
    deadline = time.time() + 120
    while time.time() < deadline:
        picked = pick_new_export_file(ctx, before_paths, require_download_clicked=True, min_size=1000)
        if picked is not None:
            ctx.record("fallback_existing_export", "file_ready", {"path": str(picked), "size": picked.stat().st_size})
            return picked
        time.sleep(2)
    raise RuntimeError("fallback：已点击「下载」但 120s 内未出现 xlsx 文件。")


def parse_args():
    parser = argparse.ArgumentParser(description="千里马 VIP 主线导出 CDP 修复副本")
    parser.add_argument("--cdp-port", type=int, default=9222, help="Chrome remote debugging port")
    parser.add_argument("--subscribe-url", default=DEFAULT_SUBSCRIBE_URL, help="千里马标讯订阅页 URL")
    parser.add_argument("--start-date", help="开始日期 YYYY-MM-DD；不填则自动使用昨天")
    parser.add_argument("--end-date", help="结束日期 YYYY-MM-DD；不填则自动使用今天")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT), help="输出目录")
    parser.add_argument(
        "--resume-context",
        help="Phase 1 checkpoint JSON；恢复既有目标行并继续生成/刷新/下载，禁止重新提交导出",
    )
    parser.add_argument(
        "--allow-existing-export-download",
        action="store_true",
        help="S5：额度 0 等导致新导出无法下载时，允许下载导出记录中已生成行（默认关）",
    )
    parser.add_argument(
        "--download-existing-only",
        action="store_true",
        help="只下载导出记录里已生成行（默认优先 150 条），不批量导出、不消耗新额度；用于 150 额度循环测试",
    )
    parser.add_argument(
        "--prefer-export-count",
        type=int,
        default=150,
        help="download-existing-only / fallback 时优先选择的导出条数（默认 150）",
    )
    parser.add_argument(
        "--config",
        default=os.environ.get("QLM_BID_CONFIG") or "",
        help="本机敏感配置（账号密码）；登录超时后自动登录必需",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    resume_payload = {}
    resume_path = Path(args.resume_context).resolve() if args.resume_context else None
    if resume_path:
        if not resume_path.exists():
            raise FileNotFoundError(f"Phase 1 resume checkpoint not found: {resume_path}")
        resume_payload = json.loads(resume_path.read_text(encoding="utf-8-sig"))
    default_start, default_end = yesterday_today()
    start_date = args.start_date or resume_payload.get("start_date") or default_start
    end_date = args.end_date or resume_payload.get("end_date") or default_end
    ctx = RunContext(Path(args.output_root), start_date, end_date)
    ctx.allow_existing_export_download = bool(args.allow_existing_export_download)
    download_existing_only = bool(args.download_existing_only)
    prefer_export_count = args.prefer_export_count
    agent = None
    before_paths: set[str] = set()
    try:
        if resume_path:
            apply_resume_context(ctx, resume_payload, resume_path)
        before_paths = {f["path"] for f in snapshot_downloads(ctx)}
        ctx.record(
            "start",
            "mainline",
            {
                "target": f"{start_date} to {end_date}",
                "download_dir": str(ctx.download_dir),
                "cdp_download_dir": str(ctx.cdp_download_dir),
                "subscribe_url": args.subscribe_url,
                "cdp_port": args.cdp_port,
                "allow_existing_export_download": ctx.allow_existing_export_download,
                "download_existing_only": download_existing_only,
                "prefer_export_count": prefer_export_count,
            },
        )
        agent = attach_qianlima(args.cdp_port, args.subscribe_url, ctx)
        config_path = (args.config or "").strip() or None
        if download_existing_only:
            # 150 额度循环测试：只验 状态=已生成 → 操作=下载，禁止再提交批量导出
            # 仍须先保证登录态（超时弹窗→自动登录）
            agent = ensure_on_subscribe_page(
                agent,
                ctx,
                args.subscribe_url,
                cdp_port=args.cdp_port,
                config_path=config_path,
            )
            ctx.record(
                "download_existing_only",
                "skip_batch_export",
                {
                    "prefer_count": prefer_export_count,
                    "note": "不消耗新额度；筛选条数与导出150差额另记，待满额度再测",
                },
            )
            downloaded = try_download_existing_generated_export(
                agent, ctx, before_paths, prefer_count=prefer_export_count
            )
        else:
            agent = ensure_on_subscribe_page(
                agent,
                ctx,
                args.subscribe_url,
                cdp_port=args.cdp_port,
                config_path=config_path,
            )
            if resume_path:
                ctx.record(
                    "resume_context",
                    "skip_new_export",
                    {"checkpoint": str(resume_path), "rule": "continue existing target row only"},
                )
            else:
                set_date_filter_by_vuex(agent, ctx, start_date, end_date)
                trigger_export(agent, ctx)
            agent = wait_for_export_record_page(
                agent,
                ctx,
                args.subscribe_url,
                cdp_port=args.cdp_port,
                config_path=config_path,
            )
            try:
                downloaded = wait_for_download_file(agent, ctx, before_paths)
            except RuntimeError as exc:
                msg = str(exc)
                if ctx.allow_existing_export_download and (
                    "QUOTA_ZERO_NEED_FALLBACK" in msg
                    or "剩余 0 条" in msg
                    or "访问额度不足" in msg
                ):
                    ctx.record("fallback_existing_export", "triggered", {"reason": msg[:300]})
                    downloaded = try_download_existing_generated_export(
                        agent, ctx, before_paths, prefer_count=prefer_export_count
                    )
                else:
                    raise
        # 如果文件在 cdp_download_dir，复制到 download_dir
        if str(downloaded).startswith(str(ctx.cdp_download_dir)):
            import shutil
            dest = ctx.download_dir / downloaded.name
            shutil.copy2(downloaded, dest)
            ctx.record("copy_from_cdp", "ok", {"from": str(downloaded), "to": str(dest)})
            downloaded = dest
        validate_xlsx(downloaded, ctx)
        latest_pointer = Path(args.output_root) / "latest_qianlima_export_path.txt"
        latest_pointer.write_text(str(downloaded), encoding="utf-8")
        source_meta = Path(args.output_root) / "latest_qianlima_export_source.json"
        source_meta.write_text(
            json.dumps(
                {
                    "source": ctx.export_source,
                    "file": str(downloaded),
                    "quota_remaining": ctx.quota_remaining,
                    "run_id": ctx.run_id,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        ctx.record(
            "latest_export_pointer",
            "written",
            {"path": str(latest_pointer), "file": str(downloaded), "source": ctx.export_source},
        )
        ctx.record("result", "success", {"file": str(downloaded), "source": ctx.export_source})
        return 0
    except Exception as exc:
        ctx.record("result", "failed", {"error": str(exc)})
        if agent:
            try:
                agent.screenshot(ctx, "failure.png")
            except Exception:
                pass
        return 1
    finally:
        if agent:
            agent.close()
        ctx.save_logs()


if __name__ == "__main__":
    raise SystemExit(main())

