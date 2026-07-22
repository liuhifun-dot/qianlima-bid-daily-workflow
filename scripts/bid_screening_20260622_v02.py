#!/usr/bin/env python3
"""
标讯智能筛选 — 完整流程 v8.0
Step 1: 读取千里马 Excel（含超链接 URL）
Step 2: 6 维度评分筛选（规则从 YAML 配置文件读取）
Step 3: 输出 JSON 供后续 VIP 原文读取

用法:
  python bid_screening_20260622_v02.py [千里马Excel路径] [日期YYYYMMDD]
  python bid_screening_20260622_v02.py --input <千里马Excel路径> --date 20260527
  python bid_screening_20260622_v02.py

  不传 Excel 路径时，优先读取 latest_qianlima_export_path.txt。

  不传日期则优先从文件名提取，否则使用今天。
  输出目录和 YAML 规则路径相对于本脚本所在目录的上级自动推断。
"""
import argparse
import pandas as pd
import json
import re
import sys
import unicodedata
import yaml
from datetime import datetime, date
from openpyxl import load_workbook
from pathlib import Path


# ============ 路径配置 ============
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
RULES_FILE = PROJECT_DIR / "filter_rules" / "rules_current.yaml"
OUTPUT_DIR = PROJECT_DIR / "output" / "v2_4"
PIPELINE_OUTPUT_DIR = PROJECT_DIR / "output"

REQUIRED_COLUMNS = [
    "标题", "地区", "二级类型", "招标估价", "投标截止时间",
    "发布时间", "招标单位", "招标编号", "采购类型", "命中的关键词", "内容预览",
]

CRITICAL_COLUMNS = [
    "标题", "地区", "招标估价", "投标截止时间", "发布时间", "招标单位", "招标编号",
]


_DEDUP_TRAILING_LABELS = (
    "招标公告", "采购公告", "竞争性磋商公告", "竞争性谈判公告", "询价公告",
    "比选公告", "遴选公告", "交易公告", "选取公告", "竞价公告",
)


def normalize_title_for_dedup(title):
    """Normalize presentation-only differences without merging different procurement stages."""
    value = unicodedata.normalize("NFKC", str(title or "")).strip()
    value = re.sub(r"[\s　]+", "", value)
    value = re.sub(r"[。；;，,：:]+$", "", value)
    changed = True
    while changed:
        changed = False
        for label in _DEDUP_TRAILING_LABELS:
            if value.endswith(label):
                value = value[:-len(label)].rstrip("-—_()（）[]【】")
                changed = True
                break
    return value


def load_rules():
    """从 YAML 加载筛选规则"""
    with open(RULES_FILE, "r", encoding="utf-8") as f:
        rules = yaml.safe_load(f)
    return rules


def load_excel_with_urls(filepath):
    """读取 Excel，提取标题超链接 URL"""
    wb = load_workbook(filepath)
    ws = wb.active

    df = pd.read_excel(filepath, header=0)
    real_headers = df.iloc[0].tolist()
    df.columns = real_headers
    df = df.iloc[1:].reset_index(drop=True)
    missing_critical = [c for c in CRITICAL_COLUMNS if c not in df.columns]
    if missing_critical:
        raise ValueError(f"千里马 Excel 缺少必要字段: {missing_critical}")

    if "二级类型" not in df.columns:
        if "一级类型" in df.columns:
            df["二级类型"] = df["一级类型"].apply(
                lambda v: "招标公告" if "公告" in str(v) else str(v)
            )
        else:
            df["二级类型"] = "招标公告"
    for optional_col in ["采购类型", "命中的关键词", "内容预览"]:
        if optional_col not in df.columns:
            df[optional_col] = ""

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"千里马 Excel 缺少必要字段: {missing}")

    urls = []
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        url = None
        if cell.hyperlink:
            url = cell.hyperlink.target
        if not url and "详情链接" in df.columns and row - 3 < len(df):
            detail_url = df.iloc[row - 3].get("详情链接")
            if pd.notna(detail_url) and str(detail_url).strip():
                url = str(detail_url).strip()
        urls.append(url)

    df["原文链接"] = urls
    return df


def normalize_run_date(value):
    if not value:
        return None
    value = str(value).strip()
    if re.fullmatch(r"\d{8}", value):
        return value
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        return value.replace("-", "")
    raise ValueError(f"日期格式不支持: {value}，请使用 YYYYMMDD 或 YYYY-MM-DD")


def infer_run_date_from_file(path):
    name = Path(path).name
    m = re.search(r"(20\d{2})[-_]?(\d{2})[-_]?(\d{2})", name)
    if m:
        return f"{m.group(1)}{m.group(2)}{m.group(3)}"
    m = re.search(r"(?<!\d)(\d{2})-(\d{2})(?!\d)", name)
    if m:
        return f"{date.today().year}{m.group(1)}{m.group(2)}"
    return date.today().strftime("%Y%m%d")


def format_run_date(run_date):
    run_date = normalize_run_date(run_date)
    return f"{run_date[:4]}-{run_date[4:6]}-{run_date[6:8]}"


def pointer_candidates():
    candidates = []
    env_path = None
    try:
        import os
        env_path = os.environ.get("QIANLIMA_EXPORT_POINTER")
    except Exception:
        env_path = None
    if env_path:
        candidates.append(Path(env_path))
    candidates.extend([
        Path.cwd() / "latest_qianlima_export_path.txt",
        PROJECT_DIR / "latest_qianlima_export_path.txt",
        PIPELINE_OUTPUT_DIR / "latest_qianlima_export_path.txt",
    ])
    seen = set()
    unique = []
    for p in candidates:
        key = str(p).lower()
        if key not in seen:
            unique.append(p)
            seen.add(key)
    return unique


def resolve_input_file(explicit_input=None, explicit_pointer=None):
    if explicit_input:
        path = Path(explicit_input)
        if not path.exists():
            raise FileNotFoundError(f"指定 Excel 不存在: {path}")
        return path.resolve(), None

    candidates = [Path(explicit_pointer)] if explicit_pointer else pointer_candidates()
    checked = []
    for pointer in candidates:
        checked.append(str(pointer))
        if not pointer.exists():
            continue
        raw = pointer.read_text(encoding="utf-8-sig").strip().strip('"').lstrip("\ufeff")
        if not raw:
            continue
        path = Path(raw)
        if not path.exists():
            raise FileNotFoundError(f"指针文件存在但指向的 Excel 不存在: {pointer} -> {path}")
        return path.resolve(), pointer.resolve()

    raise FileNotFoundError(
        "未提供千里马 Excel，也未找到可用 latest_qianlima_export_path.txt。"
        f" 已检查: {checked}"
    )


def next_output_path(run_date):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    pattern = re.compile(rf"^标讯筛选结果_{re.escape(run_date)}_v(\d+)\.json$")
    versions = []
    for p in OUTPUT_DIR.glob(f"标讯筛选结果_{run_date}_v*.json"):
        m = pattern.match(p.name)
        if m:
            versions.append(int(m.group(1)))
    next_version = max(versions, default=0) + 1
    return OUTPUT_DIR / f"标讯筛选结果_{run_date}_v{next_version:02d}.json"


def write_latest_pointers(output_path):
    PIPELINE_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    pointer_paths = [
        OUTPUT_DIR / "latest_screening_result_path.txt",
        PIPELINE_OUTPUT_DIR / "latest_screening_result_path.txt",
    ]
    for pointer in pointer_paths:
        pointer.write_text(str(output_path), encoding="utf-8")
    return pointer_paths


def score_amount(amount_wan):
    """金额评分（30分）"""
    if amount_wan is None or pd.isna(amount_wan):
        return 0
    if 50 <= amount_wan <= 350:
        return 30
    elif 20 <= amount_wan < 50:
        return 22
    elif 100 <= amount_wan <= 500:
        return 28
    elif 5 <= amount_wan < 20:
        return 12
    elif amount_wan > 500:
        return 20
    else:
        return 5


def score_industry(search_text, core_keywords, related_keywords):
    """行业匹配评分（25分）— 标题+简介宽召回，最终判断交给正文复判。"""
    score = 0
    matched_core = []
    matched_related = []

    for kw in core_keywords:
        if kw in search_text:
            score += 5
            matched_core.append(kw)

    if not matched_core:
        for kw in related_keywords:
            if kw in search_text:
                score += 3
                matched_related.append(kw)

    return min(score, 25), matched_core, matched_related


def score_region(region, region_scores):
    """地区评分（15分）— 从配置文件读取"""
    if not region or pd.isna(region):
        return 0
    for item in region_scores:
        if item["prefix"] and region.startswith(item["prefix"]):
            return item["score"]
    return 3  # 默认


def score_time(deadline_str):
    """时间窗口评分（10分）"""
    if not deadline_str or pd.isna(deadline_str):
        return 3

    try:
        if "T" in str(deadline_str):
            dt = pd.to_datetime(deadline_str)
        else:
            dt = pd.to_datetime(str(deadline_str))
        days_left = (dt - pd.Timestamp.now()).days
        if days_left >= 15:
            return 10
        elif days_left >= 7:
            return 7
        elif days_left >= 3:
            return 4
        else:
            return 1
    except:
        return 3


def score_qualification(title):
    """资质匹配推断（15分）— 基于标题内容推断"""
    score = 5
    if any(kw in title for kw in ["市政", "道路", "公路", "照明工程", "路灯", "亮化"]):
        score = 15
    elif any(kw in title for kw in ["装修", "改造", "修缮", "安装"]):
        score = 10
    elif any(kw in title for kw in ["设计", "监理", "咨询", "检测"]):
        score = 8
    return score


def run_scoring(df, rules, run_date=None, input_file=None, pointer_file=None):
    """执行完整评分流程"""
    hf = rules["hard_filter"]
    exclude_kw = rules["exclude_keywords"]
    deferred_kw_list = rules["deferred_exclude_keywords"]
    core_kw = rules["core_keywords"]
    related_kw = rules["related_keywords"]
    region_scores = rules["region_scores"]

    results = []
    excluded = []
    no_url = []

    # 硬过滤：省份 + 招标公告
    bid_df = df[
        (df["地区"].str.startswith(hf["province"], na=False)) &
        (df["二级类型"] == hf["bid_type"])
    ].copy()

    bid_df["估价万"] = pd.to_numeric(bid_df["招标估价"], errors="coerce") / 10000

    # 金额范围过滤
    target_df = bid_df[
        (bid_df["估价万"] >= hf["amount_min_wan"]) &
        (bid_df["估价万"] <= hf["amount_max_wan"])
    ].copy()

    # 标题去重：先处理完全相同标题，再处理仅公告后缀不同的同一采购项目。
    before_dedup = len(target_df)
    target_df["_dedup_title"] = target_df["标题"].map(normalize_title_for_dedup)
    target_df = target_df.drop_duplicates(subset=["_dedup_title"], keep="first")
    target_df = target_df.drop(columns=["_dedup_title"])
    after_dedup = len(target_df)
    dedup_count = before_dedup - after_dedup

    print(f"原始数据: {len(df)} 条")
    print(f"广东招标公告: {len(bid_df)} 条")
    print(f"金额 {hf['amount_min_wan']}-{hf['amount_max_wan']}万: {before_dedup} 条")
    if dedup_count > 0:
        print(f"标题去重: 去掉 {dedup_count} 条，剩余 {after_dedup} 条")
    print(f"规则文件: {RULES_FILE}")
    print(f"排除词: {len(exclude_kw)} 个, 延迟排除词: {len(deferred_kw_list)} 个")

    for idx, row in target_df.iterrows():
        title = str(row["标题"])
        url = row.get("原文链接")
        keywords_hit = str(row.get("命中的关键词", "")) if pd.notna(row.get("命中的关键词")) else ""
        preview = str(row.get("内容预览", "")) if pd.notna(row.get("内容预览")) else ""
        # 行业评分只看标题，避免千里马导出的订阅全局关键词污染判断。
        # “命中的关键词/内容预览”只作为复核参考，不得把泛道路、监控、评估类项目推成推荐候选。
        industry_search_text = title

        # 核心照明关键词保护：标题命中这些词时，不允许被泛排除词直接排除
        CORE_LIGHTING_PROTECTION = [
            "应急照明", "照明", "灯光", "照明工程", "照明更换", "灯具", "灯具招标",
            "灯具更换", "路灯", "路灯设施", "道路照明", "户外照明",
            "亮化", "亮化工程", "景观亮化", "太阳能路灯", "LED路灯",
            "光伏施工", "光伏EPC", "光伏安装",
        ]
        DISPLAY_ONLY_TITLE = [
            "LED显示屏", "LED屏", "LED大屏", "显示屏", "显示系统",
            "屏采购", "护目灯",
        ]
        ABSOLUTE_SERVICE_EXCLUDE = [
            "试验及检测校验", "检测校验辅助服务", "路灯巡检服务",
            "光伏组件清洗", "舞台灯光", "舞台照明",
        ]
        title_has_core_lighting = any(kw in title for kw in CORE_LIGHTING_PROTECTION)
        absolute_service_hit = next((kw for kw in ABSOLUTE_SERVICE_EXCLUDE if kw in title), None)
        title_is_display_only = (
            any(kw in title for kw in DISPLAY_ONLY_TITLE)
            and not any(kw in title for kw in ["路灯", "照明", "灯具", "亮化", "太阳能路灯"])
        )

        ex_kw = None
        for kw in exclude_kw:
            if kw in title:
                ex_kw = kw
                break
        # Proven non-target service categories override broad lighting keyword protection.
        if absolute_service_hit:
            excluded.append({
                "title": title[:80], "region": row["地区"],
                "amount_wan": row["估价万"],
                "exclude_reason": f"明确非主营服务: {absolute_service_hit}", "url": url,
            })
            continue

        # Core keyword protection: display-only LED screens are excluded first.
        if title_is_display_only:
            excluded.append({
                "title": title[:80], "region": row["地区"],
                "amount_wan": row["估价万"],
                "exclude_reason": "LED显示屏/护目灯类非主营照明工程", "url": url,
            })
            continue

        if ex_kw and title_has_core_lighting:
            # 记录但不排除，进入正文复判
            pass
        elif ex_kw:
            excluded.append({
                "title": title[:80], "region": row["地区"],
                "amount_wan": row["估价万"],
                "exclude_reason": f"排除关键词: {ex_kw}", "url": url,
            })
            continue

        # 检查延迟排除词
        deferred_kw = None
        for kw in deferred_kw_list:
            if kw in title:
                deferred_kw = kw
                break

        # 评分
        s_amount = score_amount(row["估价万"])
        s_industry, ckw, rkw = score_industry(industry_search_text, core_kw, related_kw)
        s_region = score_region(row["地区"], region_scores)
        s_time = score_time(row.get("投标截止时间"))
        s_qual = score_qualification(title)
        s_black = 5

        # 行业匹配为 0 分的直接排除
        if s_industry == 0:
            excluded.append({
                "title": title[:80], "region": row["地区"],
                "amount_wan": row["估价万"],
                "exclude_reason": "行业不匹配（标题无相关关键词）", "url": url,
            })
            continue

        raw_total = s_amount + s_industry + s_region + s_time + s_qual + s_black
        total = raw_total
        score_notes = []

        # Explicit lighting titles must enter evidence review unless a hard exclusion matched.
        if title_has_core_lighting and total < 50:
            total = 50
            score_notes.append("标题命中明确照明/灯光词，最低进入正文复判")

        # 经验门槛：没有命中核心业务词时，相关词只负责召回，不能直接进入推荐候选。
        # 例如“道路安全提升/附属设施整修/市政道路工程”可能包含少量照明，但主标的通常不是我方主营。
        if not ckw and total >= 70:
            total = 69
            score_notes.append("未命中核心业务词，最高进入待复核")

        url_status = "无链接" if not url else "有链接"

        result = {
            "title": title, "url": url, "url_status": url_status,
            "region": row["地区"],
            "amount_wan": round(row["估价万"], 1),
            "publish_time": str(row.get("发布时间", "")),
            "deadline": str(row.get("投标截止时间", "")),
            "bid_unit": str(row.get("招标单位", "")) if pd.notna(row.get("招标单位")) else "",
            "bid_no": str(row.get("招标编号", "")) if pd.notna(row.get("招标编号")) else "",
            "purchase_type": str(row.get("采购类型", "")),
            "keywords_hit": keywords_hit,
            "preview": preview[:300],
            "recall_basis": "标题+命中关键词+内容预览",
            "deferred_exclude_kw": deferred_kw,
            "scores": {
                "金额": s_amount, "行业": s_industry, "地区": s_region,
                "资质": s_qual, "时间": s_time, "黑名单": s_black,
                "原始总分": raw_total, "总分": total,
            },
            "score_notes": score_notes,
            "matched_core_kw": ckw, "matched_related_kw": rkw,
        }
        results.append(result)

    results.sort(key=lambda x: x["scores"]["总分"], reverse=True)

    recommended = [r for r in results if r["scores"]["总分"] >= 70]
    considered = [r for r in results if 50 <= r["scores"]["总分"] < 70]
    low_score = [r for r in results if r["scores"]["总分"] < 50]

    print(f"\n排除: {len(excluded)} 条")
    print(f"行业不匹配: {sum(1 for e in excluded if '行业不匹配' in e['exclude_reason'])} 条")
    print(f"排除关键词: {sum(1 for e in excluded if '排除关键词' in e['exclude_reason'])} 条")
    print(f"\n高评分(>=70): {len(recommended)} 条")
    print(f"中等(50-69): {len(considered)} 条")
    print(f"低分(<50): {len(low_score)} 条")

    return {
        "meta": {
            "schema_version": "screening_20260527_v02",
            "date": format_run_date(run_date) if run_date else date.today().isoformat(),
            "run_date": run_date,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "source_file": str(input_file) if input_file else "",
            "source_file_name": Path(input_file).name if input_file else "",
            "source_pointer": str(pointer_file) if pointer_file else "",
            "rules_file": str(RULES_FILE),
            "total_raw": len(df), "total_gd_bid": len(bid_df),
            "total_filtered": len(target_df), "total_excluded": len(excluded),
            "total_recommended": len(recommended), "total_considered": len(considered),
            "total_low": len(low_score),
        },
        "recommended": recommended, "considered": considered,
        "low_score": low_score, "excluded": excluded, "no_url_titles": no_url,
    }


def main():
    parser = argparse.ArgumentParser(description="标讯智能筛选 v20260527_v02")
    parser.add_argument("legacy_input", nargs="?", help="兼容旧用法：千里马 Excel 路径")
    parser.add_argument("legacy_date", nargs="?", help="兼容旧用法：日期 YYYYMMDD")
    parser.add_argument("--input", dest="input_file", help="千里马 Excel 路径")
    parser.add_argument("--date", dest="run_date", help="日期 YYYYMMDD 或 YYYY-MM-DD")
    parser.add_argument("--latest-export-pointer", dest="latest_export_pointer", help="latest_qianlima_export_path.txt 路径")
    args = parser.parse_args()

    input_file, pointer_file = resolve_input_file(args.input_file or args.legacy_input, args.latest_export_pointer)
    run_date = normalize_run_date(args.run_date or args.legacy_date) if (args.run_date or args.legacy_date) else infer_run_date_from_file(input_file)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 加载规则
    rules = load_rules()

    print("=" * 60)
    print("标讯智能筛选 v20260527_v02 — 千里马 Excel 评分")
    print(f"输入: {input_file}")
    if pointer_file:
        print(f"指针: {pointer_file}")
    print(f"日期: {run_date}")
    print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Step 1: 读取数据
    print("\n[Step 1] 读取 Excel...")
    df = load_excel_with_urls(input_file)
    print(f"  读取 {len(df)} 条记录")

    # Step 2: 评分
    print("\n[Step 2] 评分筛选...")
    scored = run_scoring(df, rules, run_date=run_date, input_file=input_file, pointer_file=pointer_file)

    # Step 3: 保存
    output_path = next_output_path(run_date)
    scored["meta"]["output_path"] = str(output_path)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(scored, f, ensure_ascii=False, indent=2)
    pointer_paths = write_latest_pointers(output_path)
    print(f"\n[Step 3] 结果已保存: {output_path}")
    print("[Step 3] 最新筛选结果指针:")
    for pointer in pointer_paths:
        print(f"  - {pointer}")

    # 打印高分列表
    print("\n" + "=" * 60)
    print("高评分项目（>=70分）")
    print("=" * 60)
    for i, r in enumerate(scored["recommended"]):
        s = r["scores"]
        print(f"\n  #{i+1} [{s['总分']}分] {r['title'][:60]}")
        print(f"       地区: {r['region']} | 金额: {r['amount_wan']}万")
        if r["url"]:
            print(f"       URL: {r['url']}")
        if r.get("deferred_exclude_kw"):
            print(f"       ⏳ 延迟排除: {r['deferred_exclude_kw']}（需VIP正文确认）")

    return scored


if __name__ == "__main__":
    main()
