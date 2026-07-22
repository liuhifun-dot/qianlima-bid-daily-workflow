# -*- coding: utf-8 -*-
"""
Formal Phase 3/4 report, archive, Dingpan upload, and DingTalk push script.

This is a versioned repair copy. It does not depend on the old local-draft
report script and does not read latest_* pointers by default.

Automation contract:
- input paths come from --manifest or explicit arguments;
- Phase 2 JSON is validated before report generation;
- Excel is copied from the fixed template workbook;
- local archive, shared archive, Dingpan upload, and DingTalk message are gates;
- if any required Phase 4 gate fails, the script stops before sending DingTalk.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import urllib.request
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from report_builder_20260615_v1 import build_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


SCRIPT_DIR = Path(__file__).resolve().parent
SOP_ROOT = SCRIPT_DIR.parent
_TEMPLATE_CANDIDATES = sorted((SOP_ROOT / "02_模板").glob("*.xlsx")) or sorted((SOP_ROOT / "assets").glob("*.xlsx"))
if not _TEMPLATE_CANDIDATES:
    raise FileNotFoundError(f"No Excel template found under {SOP_ROOT / '02_模板'} or {SOP_ROOT / 'assets'}")
DEFAULT_TEMPLATE = _TEMPLATE_CANDIDATES[0]
DEFAULT_CONFIG = os.environ.get("QLM_BID_CONFIG", "")
DEFAULT_RUNS_DIR = SOP_ROOT / "runs"
# 归档路径：仅环境变量/CLI/外部配置；禁止写死本机路径（换机必炸且泄露环境指纹）。
DEFAULT_LOCAL_ARCHIVE_DIR = os.environ.get("QLM_LOCAL_ARCHIVE", "")
DEFAULT_SHARE_ARCHIVE_DIR = os.environ.get("QLM_SHARE_ARCHIVE", "")
PHASE2_VALIDATOR = SCRIPT_DIR / "validate_phase2_rejudge_20260622_v02.py"
EXCEL_VALIDATOR = SCRIPT_DIR / "validate_excel_template_contract_20260603_v02.py"

BAD_MESSAGE_TOKENS = [
    "未上传",
    "共享盘未同步",
    "local-draft",
    "archive_share_skipped",
    "OAuth token 过期",
    "OSS 签名失败",
    "确认投标",
    "确认跟进",
    "\u003f\u003f\u003f\u003f",
    "attachment_required=",
    "attachment_read_ok=",
    "procurement-demand",
    "valid tender/spec/BOQ",
]

KEEP_TEXT = {"推荐", "推荐候选", "keep"}
REVIEW_TEXT = {"待人工复核", "待复核", "待定", "needs_review"}
REJECT_TEXT = {"排除", "放弃", "reject"}

HEADER_ALL = {
    "seq": "A",
    "notice_type": "B",
    "title": "C",
    "province": "D",
    "city": "E",
    "buyer": "F",
    "project_type": "H",
    "amount": "I",
    "publish_date": "J",
    "deadline": "K",
    "duration": "L",
    "agency": "M",
    "payment": "N",
    "url": "O",
    "exclude_reason": "Z",
    "analysis": "AA",
    "score_detail": "AB",
    "review_status": "AC",
    "second_decision": "AD",
    "second_reason": "AE",
}

HEADER_TODAY = {
    "seq": "A",
    "notice_type": "B",
    "title": "C",
    "province": "D",
    "city": "E",
    "buyer": "F",
    "project_type": "H",
    "amount": "I",
    "publish_date": "J",
    "deadline": "K",
    "duration": "L",
    "agency": "M",
    "payment": "N",
    "url": "O",
    "condition": "Z",
    "analysis": "AA",
    "review_status": "AB",
    "second_decision": "AC",
    "second_reason": "AD",
}


class GateError(RuntimeError):
    pass


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def compact_date(value: str | None) -> str:
    if value:
        digits = re.sub(r"\D", "", value)
        if len(digits) >= 8:
            return digits[:8]
    return datetime.now().strftime("%Y%m%d")


def iso_date(value: str) -> str:
    c = compact_date(value)
    return f"{c[:4]}-{c[4:6]}-{c[6:8]}"


def default_source_range(run_date: str) -> tuple[str, str]:
    end = datetime.strptime(compact_date(run_date), "%Y%m%d").date()
    start = end - timedelta(days=1)
    return start.isoformat(), end.isoformat()


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def enforce_evidence_release_gate(business: dict, dry_run: bool) -> None:
    """Block formal Phase 3/4 publication when Phase 2 evidence is incomplete."""
    if bool(business.get("evidence_incomplete")) and not dry_run:
        raise GateError(
            "Phase 2 evidence is marked incomplete. Only --dry-run is allowed; "
            "formal archive, Dingpan upload, and DingTalk send are blocked."
        )


def write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def read_config_text(path: Path | None) -> str:
    if path is None or not path.is_file():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore")


def read_config_value(config_text: str, key: str) -> str:
    patterns = [
        rf"{re.escape(key)}:\s*\"([^\"]*)\"",
        rf"{re.escape(key)}\s*=\s*\"([^\"]*)\"",
        rf"{re.escape(key)}\s*[:=]\s*([^\s\r\n]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, config_text)
        if match:
            return match.group(1).strip().strip('"')
    return ""


def flatten_results(data: dict) -> list[dict]:
    if isinstance(data.get("results"), list):
        return [item for item in data["results"] if isinstance(item, dict)]

    rows: list[dict] = []
    for section in ["recommendations", "pending_review", "exclusions"]:
        value = data.get(section, {})
        if isinstance(value, dict):
            for tag, item in value.items():
                if isinstance(item, dict):
                    row = dict(item)
                    row.setdefault("tag", str(tag))
                    rows.append(row)
        elif isinstance(value, list):
            rows.extend([item for item in value if isinstance(item, dict)])
    return rows


def normalized_decision(item: dict) -> str:
    raw = str(item.get("decision") or item.get("final_decision") or "").strip()
    if raw in KEEP_TEXT:
        return "keep"
    if raw in REVIEW_TEXT:
        return "needs_review"
    if raw in REJECT_TEXT:
        return "reject"
    return raw


def display_decision(decision: str) -> str:
    return {
        "keep": "🔥 推荐候选",
        "needs_review": "📝 待人工复核",
        "reject": "🚫 排除",
    }.get(decision, decision)


def evidence_text(item: dict) -> str:
    evidence = item.get("evidence", "")
    if isinstance(evidence, list):
        parts = []
        for part in evidence:
            if isinstance(part, dict):
                parts.append(str(part.get("snippet") or part.get("text") or ""))
            else:
                parts.append(str(part))
        return "\n".join(p for p in parts if p).strip()
    return str(evidence or "").strip()


def first_nonempty(*values: Any, default: str = "—") -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return default


def split_region(region: str) -> tuple[str, str]:
    parts = [p for p in re.split(r"[-/／\s]+", region or "") if p]
    if not parts:
        return "广东", "—"
    if len(parts) == 1:
        return parts[0], "—"
    return parts[0], parts[1]


def infer_project_type(item: dict) -> str:
    decision = normalized_decision(item)
    raw_type = first_nonempty(item.get("project_type"), item.get("business_direction"), default="")
    text = f"{item.get('title', '')} {item.get('review_reason', '')} {item.get('reason', '')} {evidence_text(item)} {raw_type}"
    if decision == "reject":
        if any(word in text for word in ["监理", "评价", "评估", "检测", "监测", "审计", "咨询", "会务", "租赁"]):
            return "服务类 / 排除"
        if any(word in text for word in ["LED显示屏", "LED大屏", "显示系统"]):
            return "显示屏系统 / 排除"
        if any(word in text for word in ["EPS", "主机柜", "备件", "事故照明装置"]):
            return "备件/设备供货 / 排除"
        return "非当前业务范围"
    if any(word in text for word in ["光伏", "分布式光伏", "光伏发电"]):
        return "光伏工程 / 安装改造"
    if any(word in text for word in ["路灯", "道路照明", "太阳能路灯"]):
        return "路灯 / 道路照明"
    if any(word in text for word in ["亮化", "景观灯", "洗墙灯", "线条灯", "投光灯"]):
        return "景观亮化 / 户外灯具"
    if any(word in text for word in ["照明", "灯具", "LED"]):
        return "照明工程 / 灯具更换"
    if "充电桩" in text:
        return "充电桩 / 待确认"
    if raw_type and raw_type not in {"工程", "货物", "服务", "不确定", "初筛排除"}:
        return raw_type
    return "工程 / 待确认"

def humanize_reason(value: Any) -> str:
    text = first_nonempty(value, default="")
    attachment_gate = "\u9879\u76ee\u9700\u8981\u62db\u6807\u6587\u4ef6/\u91c7\u8d2d\u9700\u6c42/\u5de5\u7a0b\u91cf\u6e05\u5355\u4f5c\u4e3a\u8bc1\u636e\uff0c\u4f46\u9644\u4ef6\u6b63\u6587\u672a\u6210\u529f\u5b8c\u6574\u8bfb\u53d6\uff0c\u9700\u4eba\u5de5\u590d\u6838"
    technical_attachment = (
        "attachment_required=true" in text
        or "attachment_read_ok=false" in text
        or "valid tender/spec/BOQ" in text
        or "procurement-demand attachment text was not read" in text
        or "\u003f\u003f\u003f\u003f" in text
    )
    text = re.sub(
        r"\s*;?\s*attachment_required=true.*?(?:was not read|$)",
        "",
        text,
        flags=re.I,
    ).strip(" ;?")
    text = text.replace(
        "valid tender/spec/BOQ/procurement-demand attachment text was not read",
        "",
    ).strip(" ;?")
    text = text.replace("body_read_ok=false", "VIP\u6b63\u6587\u7f3a\u5931").strip(" ;?")
    text = text.replace(
        "vip_status=missing and no VIP body; only needs_review may proceed",
        "VIP\u6b63\u6587\u7f3a\u5931\uff0c\u9700\u4eba\u5de5\u590d\u6838",
    )
    if technical_attachment and "\u9644\u4ef6" not in text:
        text = f"{text}; {attachment_gate}" if text else attachment_gate
    return text or "\u8bc1\u636e\u4e0d\u8db3\uff0c\u9700\u4eba\u5de5\u590d\u6838"


def concise_message_reason(value: Any, maximum: int = 90) -> str:
    text = humanize_reason(value)
    if "初筛排除依据" in text:
        score = re.search(r"总分\s*(\d+)", text)
        text = f"初筛评分不足（总分{score.group(1)}），未达推荐阈值" if score else "初筛评分不足，未达推荐阈值"
    text = re.sub(r"\s+", " ", text)
    return text if len(text) <= maximum else text[:maximum - 1] + "…"


def attachment_evidence_summary(items: list[dict]) -> dict:
    def truthy(value: Any) -> bool:
        return bool(value)

    page_available = [
        item for item in items
        if truthy(item.get("page_attachment_available")) or truthy(item.get("attachment_available"))
    ]
    page_complete = [
        item for item in page_available
        if truthy(item.get("page_attachment_read_complete"))
    ]
    page_any = [
        item for item in page_available
        if truthy(item.get("page_attachment_read_any_ok")) or truthy(item.get("page_attachment_read_complete"))
    ]
    evidence_required = [
        item for item in items
        if truthy(item.get("evidence_required")) or truthy(item.get("attachment_required"))
    ]
    evidence_read_ok = [
        item for item in evidence_required
        if truthy(item.get("evidence_read_ok")) or truthy(item.get("attachment_read_ok"))
    ]
    evidence_gap = [item for item in evidence_required if item not in evidence_read_ok]
    return {
        "page_attachment_available_count": len(page_available),
        "page_attachment_read_complete_count": len(page_complete),
        "page_attachment_read_any_count": len(page_any),
        "page_attachment_partial_or_failed_count": max(0, len(page_available) - len(page_complete)),
        "evidence_required_count": len(evidence_required),
        "evidence_read_ok_count": len(evidence_read_ok),
        "evidence_gap_count": len(evidence_gap),
        # Compatibility keys for older downstream readers.
        "required_count": len(evidence_required),
        "read_ok_count": len(evidence_read_ok),
        "systemic_failure": bool(page_available) and not page_any,
    }


def build_row(item: dict, seq: int) -> dict:
    decision = normalized_decision(item)
    region = first_nonempty(item.get("region"), item.get("地区"), default="广东-—")
    province, city = split_region(region)
    reason = humanize_reason(first_nonempty(
        item.get("reason"),
        item.get("review_reason"),
        item.get("reject_reason"),
        item.get("manual_review_reason"),
        item.get("manual_action_required"),
    ))
    evidence = evidence_text(item)
    title = first_nonempty(item.get("title"), default="")
    level = first_nonempty(item.get("recommendation_level"), item.get("final_grade"), default="C")
    tag = first_nonempty(item.get("tag"), item.get("index"), default=str(seq))
    amount = first_nonempty(item.get("amount"), item.get("预算金额"), default="—")
    if item.get("amount_wan") and amount == "—":
        amount = f"{item['amount_wan']}万元"
    analysis = reason
    if evidence and evidence not in analysis:
        analysis = f"{reason}\n证据：{evidence}"
    return {
        "seq": seq,
        "tag": tag,
        "notice_type": first_nonempty(item.get("notice_type"), item.get("公告类型"), default="招标公告"),
        "title": title,
        "province": province,
        "city": city,
        "buyer": first_nonempty(item.get("buyer"), item.get("招标单位"), item.get("采购单位"), default="未获取"),
        "project_type": infer_project_type(item),
        "amount": amount,
        "publish_date": first_nonempty(item.get("publish_date"), item.get("发布日期"), default="—"),
        "deadline": first_nonempty(item.get("deadline"), item.get("截止日期"), default="—"),
        "duration": first_nonempty(item.get("duration"), item.get("工期"), default="—"),
        "agency": first_nonempty(item.get("agency"), item.get("招标代理"), default="未获取"),
        "payment": first_nonempty(item.get("payment"), item.get("付款方式"), default="需查阅招标文件"),
        "url": first_nonempty(item.get("url"), item.get("千里马链接"), default=""),
        "decision": decision,
        "decision_display": display_decision(decision),
        "level": level,
        "condition": f"{level}级；{reason}",
        "analysis": analysis,
        "score_detail": first_nonempty(item.get("score_detail"), item.get("评分明细"), default=f"{level}级复判"),
        "review_status": "⬜待业务确认",
        "second_decision": display_decision(decision),
        "second_reason": reason,
        "exclude_reason": humanize_reason(item.get("reject_reason")) if decision == "reject" else "",
        "manual_review_reason": humanize_reason(first_nonempty(item.get("manual_review_reason"), item.get("manual_action_required"), default="")) if decision == "needs_review" else "",
        "evidence": evidence,
        "raw": item,
    }


def row_search_text(row: dict) -> str:
    raw = row.get("raw") or {}
    raw_evidence = raw.get("evidence", "")
    if isinstance(raw_evidence, (list, dict)):
        raw_evidence = json.dumps(raw_evidence, ensure_ascii=False)
    fields = [
        row.get("title", ""),
        row.get("project_type", ""),
        row.get("second_reason", ""),
        row.get("manual_review_reason", ""),
        row.get("analysis", ""),
        row.get("evidence", ""),
        raw.get("business_direction", ""),
        raw.get("judgment_reason", ""),
        raw.get("manual_review_reason", ""),
        raw.get("reject_reason", ""),
        raw.get("reason", ""),
        raw_evidence,
    ]
    return "\n".join(str(value) for value in fields if value)


def collect_business_anomalies(rows: list[dict]) -> list[dict]:
    anomalies: list[dict] = []

    def add(row: dict, severity: str, issue: str, required_action: str) -> None:
        anomalies.append({
            "severity": severity,
            "tag": row.get("tag", ""),
            "decision": row.get("decision", ""),
            "level": row.get("level", ""),
            "title": row.get("title", ""),
            "issue": issue,
            "required_action": required_action,
            "url": row.get("url", ""),
        })

    hard_service_words = ["监理", "评估", "检测", "监测", "审计", "咨询", "租赁", "安评", "评价", "设计"]
    broad_road_words = ["道路安全提升", "道路及附属设施", "道路整修", "市政道路", "路面", "硬底化", "拓宽"]
    core_words = ["路灯", "太阳能路灯", "道路照明", "亮化", "景观灯", "照明工程", "分布式光伏", "光伏发电"]
    pv_evidence_words = ["施工工期", "工程预算", "承装", "承修", "承试", "安全生产许可证", "安装", "拆装", "光伏电缆", "PVC管", "抗风绳", "光伏组件洗板"]

    for row in rows:
        title = row.get("title", "")
        text = row_search_text(row)
        decision = row.get("decision", "")
        if decision == "reject" and "光伏" in text and any(word in text for word in pv_evidence_words):
            add(
                row,
                "review",
                "光伏项目被排除，但正文/证据出现施工工期、工程预算、资质或现场拆装信号",
                "人工复核：确认是否应改为待人工复核，不能只因电缆/配电词直接排除",
            )
        if decision == "reject" and any(word in title for word in ["路灯", "太阳能路灯", "道路照明", "亮化"]):
            add(
                row,
                "review",
                "标题含明确照明/路灯/亮化主业词但被排除",
                "人工复核：确认是否为监理/检测/设计/纯服务；否则不应直接排除",
            )
        if decision == "keep" and any(word in title for word in hard_service_words):
            add(
                row,
                "block",
                "推荐候选标题命中监理/评估/检测/租赁/咨询等硬排除服务词",
                "停止发送：必须改为排除或待人工复核后重新校验",
            )
        if decision == "keep" and any(word in title for word in broad_road_words) and not any(word in text for word in core_words):
            add(
                row,
                "block",
                "泛道路/土建标题进入推荐候选，但未发现明确照明或光伏主业证据",
                "停止发送：必须补正文证据或降为排除/待人工复核",
            )
        if decision == "keep" and not row.get("evidence"):
            add(
                row,
                "block",
                "推荐候选缺少证据文本",
                "停止发送：推荐候选必须有正文、附件或页面证据",
            )
    return anomalies


def business_anomaly_summary(anomalies: list[dict]) -> dict:
    return {
        "ok": not any(item.get("severity") == "block" for item in anomalies),
        "block_count": sum(1 for item in anomalies if item.get("severity") == "block"),
        "review_count": sum(1 for item in anomalies if item.get("severity") == "review"),
        "total": len(anomalies),
    }


def sort_rows(rows: list[dict]) -> list[dict]:
    decision_rank = {"keep": 0, "needs_review": 1, "reject": 2}
    level_rank = {"A": 0, "B": 1, "C": 2, "D": 3}
    return sorted(
        rows,
        key=lambda row: (
            decision_rank.get(row["decision"], 9),
            level_rank.get(str(row["level"])[:1], 9),
            row["seq"],
        ),
    )


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)


def unmerge_rows(ws, start_row: int = 2) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row:
            ws.unmerge_cells(str(merged_range))


def clear_sheet_rows(ws, start_row: int = 2) -> None:
    unmerge_rows(ws, start_row)
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.hyperlink = None
            cell.comment = None


def set_cell(ws, addr: str, value: Any) -> None:
    ws[addr] = value
    ws[addr].alignment = Alignment(wrap_text=True, vertical="top")


def write_table(ws, rows: list[dict], mapping: dict[str, str], start_row: int = 2) -> None:
    clear_sheet_rows(ws, start_row)
    max_col = max(ws.max_column, 31)
    for index, row in enumerate(rows, start_row):
        copy_row_style(ws, 2, index, max_col)
        for key, col in mapping.items():
            ws[f"{col}{index}"] = row.get(key, "")
        if row.get("url") and "url" in mapping:
            cell = ws[f"{mapping['url']}{index}"]
            cell.hyperlink = row["url"]
            cell.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[index].height = 48


def write_today_sections(ws, rows: list[dict]) -> None:
    clear_sheet_rows(ws, 2)
    max_col = max(ws.max_column, 30)
    current = 2
    sections = [
        ("🔥 推荐候选（需业务确认）", [r for r in rows if r["decision"] == "keep"]),
        ("📝 待人工复核", [r for r in rows if r["decision"] == "needs_review"]),
        ("🚫 排除摘要", [r for r in rows if r["decision"] == "reject"]),
    ]
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    for title, section_rows in sections:
        copy_row_style(ws, 2, current, max_col)
        ws[f"A{current}"] = title
        ws[f"A{current}"].font = Font(bold=True, color="1F4E79")
        ws[f"A{current}"].fill = section_fill
        current += 1
        if not section_rows:
            copy_row_style(ws, 2, current, max_col)
            ws[f"A{current}"] = "无"
            current += 1
            continue
        for seq, row in enumerate(section_rows, 1):
            copy_row_style(ws, 2, current, max_col)
            row_copy = dict(row)
            row_copy["seq"] = seq
            for key, col in HEADER_TODAY.items():
                ws[f"{col}{current}"] = row_copy.get(key, "")
            if row.get("url"):
                cell = ws[f"{HEADER_TODAY['url']}{current}"]
                cell.hyperlink = row["url"]
                cell.font = Font(color="0563C1", underline="single")
            ws.row_dimensions[current].height = 54
            current += 1


def write_stats(ws, rows: list[dict], run_date: str, source_start: str, source_end: str) -> None:
    keep_count = sum(1 for row in rows if row["decision"] == "keep")
    review_count = sum(1 for row in rows if row["decision"] == "needs_review")
    reject_count = sum(1 for row in rows if row["decision"] == "reject")
    ws["C2"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["G2"] = f"广东 · 千里马招标网 · {source_start} 至 {source_end}"
    ws["A6"] = "总复判"
    ws["B6"] = len(rows)
    ws["C6"] = "100%"
    ws["A7"] = "推荐候选"
    ws["B7"] = keep_count
    ws["C7"] = f"{(keep_count / len(rows) * 100):.1f}%" if rows else "0%"
    ws["A8"] = "待人工复核"
    ws["B8"] = review_count
    ws["C8"] = f"{(review_count / len(rows) * 100):.1f}%" if rows else "0%"
    ws["A9"] = "排除"
    ws["B9"] = reject_count
    ws["C9"] = f"{(reject_count / len(rows) * 100):.1f}%" if rows else "0%"
    ws["A22"] = "🔥 推荐候选"
    ws["E22"] = keep_count
    ws["A23"] = "📝 待人工复核"
    ws["E23"] = review_count
    ws["A24"] = "🚫 排除"
    ws["E24"] = reject_count
    ws["A17"] = "正选词：路灯、照明、亮化、LED、光伏、储能、充电桩"
    ws["A18"] = "硬排除：检验监测、环评、纯设备、租赁、设计、监理、检测、合同/中标结果"
    ws["A19"] = "三级分类：推荐候选 / 待人工复核 / 排除"


def clone_project_sheets(wb, needed: int):
    project_sheets = [ws for ws in wb.worksheets if ws.title.startswith("项目")]
    if not project_sheets:
        raise GateError("Template has no project detail sheet.")
    template_ws = project_sheets[-1]
    for ws in list(project_sheets):
        if ws != template_ws:
            del wb[ws.title]
    template_ws.title = "项目1"
    result = [template_ws]
    for i in range(2, max(1, needed) + 1):
        copied = wb.copy_worksheet(template_ws)
        copied.title = f"项目{i}"
        result.append(copied)
    return result


def write_project_sheet(ws, row: dict) -> None:
    set_cell(ws, "A1", f"{row['decision_display']} - {row['title']}")
    set_cell(ws, "B4", row["notice_type"])
    set_cell(ws, "B5", f"{row['province']}-{row['city']}")
    set_cell(ws, "B6", row["publish_date"])
    set_cell(ws, "B7", row["deadline"])
    set_cell(ws, "B8", row["amount"])
    set_cell(ws, "B9", row["tag"])
    set_cell(ws, "B10", row["project_type"])
    set_cell(ws, "B11", f"{row['level']}级；{row['decision_display']}")
    set_cell(ws, "B12", ", ".join(row["raw"].get("scope_hits") or row["raw"].get("product_hits") or []) or "正文/标题综合判断")
    set_cell(ws, "B13", "需查阅招标文件")
    set_cell(ws, "B14", row["duration"])
    set_cell(ws, "B15", "需结合招标文件确认资质要求")
    set_cell(ws, "A17", row["url"])
    if row["url"]:
        ws["A17"].hyperlink = row["url"]
        ws["A17"].font = Font(color="0563C1", underline="single")
    set_cell(ws, "A20", row["analysis"])
    set_cell(ws, "C23", "（🔥 推荐候选 / 📝 待人工复核 / 🚫 排除）")
    set_cell(ws, "C24", row["second_reason"])
    set_cell(ws, "A28", datetime.now().strftime("%Y-%m-%d"))
    set_cell(ws, "B28", row["decision_display"])
    set_cell(ws, "C28", f"{row['title']}（当前）")
    set_cell(ws, "A33", row["evidence"][:4000] if row["evidence"] else row["analysis"][:4000])
    set_cell(ws, "A53", "附件需人工确认" if row["decision"] == "needs_review" else "无或已在正文中确认")
    set_cell(ws, "B56", "未获取")
    set_cell(ws, "B57", "未获取")


def add_raw_export_sheet(wb, raw_export: Path | None) -> None:
    title = "📦 原始导出全部数据"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    if not raw_export:
        ws["A1"] = "未提供原始导出 Excel。正式自动化应提供 raw_export_path 以保留全部来源数据。"
        return
    if not raw_export.exists():
        ws["A1"] = f"原始导出 Excel 不存在：{raw_export}"
        return
    raw_wb = load_workbook(raw_export, read_only=True, data_only=True)
    raw_ws = raw_wb.worksheets[0]
    for row in raw_ws.iter_rows(values_only=True):
        ws.append(list(row))


def versioned_path(directory: Path, stem: str, suffix: str) -> Path:
    directory.mkdir(parents=True, exist_ok=True)
    candidate = directory / f"{stem}{suffix}"
    if not candidate.exists():
        return candidate
    version = 1
    while True:
        version += 1
        candidate = directory / f"{stem}_v{version:02d}{suffix}"
        if not candidate.exists():
            return candidate


def get_next_version(base_dir: Path, date_str: str) -> int:
    """检测当天版本号，返回下一个可用版本号（v1.0 -> v2.0 -> v3.0）"""
    base_dir = Path(base_dir)
    if not base_dir.exists():
        return 1
    pattern = re.compile(rf"^{re.escape(date_str)}标讯筛选记录_v(\d+)\.0$")
    max_version = 0
    for item in base_dir.iterdir():
        if item.is_dir():
            match = pattern.match(item.name)
            if match:
                version = int(match.group(1))
                max_version = max(max_version, version)
    return max_version + 1


def archive_screening_records(
    run_dir: Path,
    local_base: Path,
    share_base: Path,
    run_date: str,
    report_path: Path,
    run_id: str,
    dry_run: bool
) -> dict:
    """归档筛选过程文件到本地和共享盘的筛选记录文件夹"""
    result = {
        "local_status": "dry_run" if dry_run else "pending",
        "share_status": "dry_run" if dry_run else "pending",
        "local_folder": "",
        "share_folder": "",
        "version": 0,
        "archived_items": [],
    }
    
    if dry_run:
        result["local_folder"] = str(local_base / f"{run_date}标讯筛选记录_v1.0")
        result["share_folder"] = str(share_base / f"{run_date}标讯筛选记录_v1.0")
        result["version"] = 1
        return result
    
    # 获取下一个版本号
    local_version = get_next_version(local_base, run_date)
    share_version = get_next_version(share_base, run_date)
    # 使用较大的版本号保持一致
    version = max(local_version, share_version)
    result["version"] = version
    
    folder_name = f"{run_date}标讯筛选记录_v{version}.0"
    local_folder = local_base / folder_name
    share_folder = share_base / folder_name
    
    # 创建本地文件夹
    local_folder.mkdir(parents=True, exist_ok=True)
    result["local_folder"] = str(local_folder)
    result["local_status"] = "ok"
    
    # 归档内容列表
    archive_items = [
        "02_screening",
        "03_body",
        "04_attachment",
        "05_rejudge",
        "07_archive_push",
        "99_logs",
    ]
    
    # 复制运行目录中的子文件夹
    for item_name in archive_items:
        src = run_dir / item_name
        dst = local_folder / item_name
        if src.exists() and src.is_dir():
            if dst.exists():
                shutil.rmtree(dst)
            shutil.copytree(src, dst)
            result["archived_items"].append(item_name)
    
    # 复制 manifest 文件
    for manifest_name in [f"pipeline_manifest_{run_id}.json", f"run_manifest_{run_id}.json"]:
        src = run_dir / manifest_name
        if src.exists():
            shutil.copy2(src, local_folder / manifest_name)
            result["archived_items"].append(manifest_name)
    
    # 复制最终 Excel 到根目录
    if report_path.exists():
        shutil.copy2(report_path, local_folder / report_path.name)
        result["archived_items"].append(report_path.name)
    
    # 复制到共享盘
    if str(share_base).startswith(r"\\"):
        if not share_base.exists():
            raise GateError(f"共享盘路径不可达：{share_base}")
        share_folder.mkdir(parents=True, exist_ok=True)
        result["share_folder"] = str(share_folder)
        
        # 复制所有内容到共享盘
        for item in local_folder.iterdir():
            dst = share_folder / item.name
            if item.is_dir():
                if dst.exists():
                    shutil.rmtree(dst)
                shutil.copytree(item, dst)
            else:
                shutil.copy2(item, dst)
        result["share_status"] = "ok"
    else:
        raise GateError(f"共享盘路径必须是 UNC 路径：{share_base}")
    
    return result


def run_subprocess(args: list[str], timeout: int = 120) -> subprocess.CompletedProcess:
    env = os.environ.copy()
    env.pop("NODE_OPTIONS", None)
    for key in ["HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "GIT_HTTP_PROXY", "GIT_HTTPS_PROXY"]:
        env[key] = ""
    return subprocess.run(args, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=timeout, env=env)


def run_python(script: Path, args: list[str], timeout: int = 120) -> dict:
    cmd = [sys.executable, str(script)] + args
    result = run_subprocess(cmd, timeout=timeout)
    payload = {
        "command": cmd,
        "returncode": result.returncode,
        "stdout": result.stdout,
        "stderr": result.stderr,
    }
    if result.returncode != 0:
        raise GateError(f"Command failed: {' '.join(cmd)}\n{result.stdout}\n{result.stderr}")
    return payload


def find_dws_command() -> str | None:
    candidates = [
        os.environ.get("DWS_CLI"),
        shutil.which("dws"),
        shutil.which("dws.cmd"),
        str(Path.home() / "AppData" / "Roaming" / "npm" / "dws.cmd"),
    ]
    candidates.extend(
        str(path)
        for path in sorted(
            (Path.home() / ".workbuddy" / "binaries" / "node" / "versions").glob("*/dws.cmd"),
            reverse=True,
        )
    )
    seen = set()
    for candidate in candidates:
        if not candidate:
            continue
        resolved = str(Path(candidate))
        if resolved.lower() in seen or not Path(resolved).exists():
            continue
        seen.add(resolved.lower())
        try:
            probe = run_subprocess([resolved, "--version"], timeout=20)
        except Exception:
            continue
        if probe.returncode == 0 and (probe.stdout.strip() or probe.stderr.strip()):
            return resolved
    return None


def require_dws_auth(dws: str) -> None:
    """Require a usable DWS session. status 可能 authenticated=true 但 access token 已过期。"""
    probe = run_subprocess([dws, "auth", "status", "--format", "json"], timeout=60)
    if probe.returncode != 0:
        raise GateError(
            "DWS authorization is unavailable. Run dws auth login in an interactive session, "
            "then rerun Phase 4. DingTalk must not be sent without a Dingpan doc_url."
        )
    try:
        payload = json.loads(probe.stdout or "{}")
    except Exception:
        payload = {}
    # 兼容字段：token_valid / authenticated；expires_at 过期则要求 re-login
    token_valid = payload.get("token_valid")
    authenticated = payload.get("authenticated")
    if token_valid is False or authenticated is False:
        raise GateError(
            "DWS access token is invalid/expired. Run `dws auth login` (refresh), "
            "then rerun Phase 4. DingTalk must not be sent without a Dingpan doc_url."
        )
    expires_at = str(payload.get("expires_at") or "").strip()
    if expires_at:
        try:
            # e.g. 2026-07-06T10:31:40.7005318+08:00
            exp = datetime.fromisoformat(expires_at)
            if exp.tzinfo is not None:
                now = datetime.now(exp.tzinfo)
            else:
                now = datetime.now()
            if now >= exp:
                raise GateError(
                    f"DWS access token expired at {expires_at}. Run `dws auth login`, "
                    "then rerun Phase 4."
                )
        except GateError:
            raise
        except Exception:
            pass


def parse_upload_doc_url(upload_data: dict) -> tuple[str, str]:
    text = json.dumps(upload_data, ensure_ascii=False)
    url_match = re.search(r"https://alidocs\.dingtalk\.com/i/nodes/[A-Za-z0-9_-]+", text)
    if url_match:
        url = url_match.group(0)
        return url, url.rsplit("/", 1)[-1]
    id_match = re.search(r'"(?:fileId|dentryUuid|nodeId|id)"\s*:\s*"([A-Za-z0-9_-]{16,})"', text)
    if id_match:
        file_id = id_match.group(1)
        return f"https://alidocs.dingtalk.com/i/nodes/{file_id}", file_id
    return "", ""


def upload_to_dingpan(file_path: Path, config_text: str, dry_run: bool) -> dict:
    space_id = read_config_value(config_text, "dingtalk_drive_space_id")
    folder_id = read_config_value(config_text, "dingtalk_drive_parent_id_for_dws")
    target_url = read_config_value(config_text, "dingtalk_drive_share_url")
    result = {
        "success": False,
        "status": "dry_run" if dry_run else "pending",
        "target_configured": bool(space_id and folder_id),
        "share_url_configured": bool(target_url),
        "doc_url": "",
        "file_id": "",
        "file_name": file_path.name,
        "reason": "",
    }
    if dry_run:
        result["reason"] = "dry-run: skipped Dingpan upload"
        return result
    if not space_id or not folder_id:
        raise GateError("Dingpan config missing dingtalk_drive_space_id or dingtalk_drive_parent_id_for_dws.")
    dws = find_dws_command()
    if not dws:
        raise GateError("dws command not found. Install/repair dingtalk-workspace-cli before Phase 4.")
    require_dws_auth(dws)
    cmd = [
        dws,
        "drive",
        "upload",
        "--file",
        str(file_path),
        "--file-name",
        file_path.name,
        "--space-id",
        space_id,
        "--folder",
        folder_id,
        "--format",
        "json",
    ]
    run = run_subprocess(cmd, timeout=240)
    safe_stdout = (run.stdout or "").replace(space_id, "<redacted>").replace(folder_id, "<redacted>")
    safe_stderr = (run.stderr or "").replace(space_id, "<redacted>").replace(folder_id, "<redacted>")
    result["command_name"] = "dws drive upload"
    result["stdout"] = safe_stdout
    result["stderr"] = safe_stderr
    if run.returncode != 0:
        raise GateError(f"Dingpan upload failed:\n{safe_stdout}\n{safe_stderr}")
    try:
        upload_data = json.loads(run.stdout or run.stderr)
    except Exception:
        upload_data = {"raw": run.stdout or run.stderr}
    doc_url, file_id = parse_upload_doc_url(upload_data)
    if not doc_url:
        raise GateError(f"Dingpan upload did not return doc_url/file id: {json.dumps(upload_data, ensure_ascii=False)[:1000]}")
    result.update({"success": True, "status": "ok", "doc_url": doc_url, "file_id": file_id})
    return result


def archive_report(
    report_path: Path,
    run_date: str,
    local_dir: Path,
    share_dir: Path,
    dry_run: bool,
    run_dir: Path | None = None,
    run_id: str = "",
) -> dict:
    """归档日报和筛选记录到文件夹结构"""
    result = {
        "local_status": "dry_run" if dry_run else "pending",
        "share_status": "dry_run" if dry_run else "pending",
        "local_folder": "",
        "share_folder": "",
        "local_file": "",  # 兼容旧字段
        "share_file": "",  # 兼容旧字段
        "version": 0,
        "archived_items": [],
    }
    
    if dry_run:
        folder_name = f"{run_date}标讯筛选记录_v1.0"
        result["local_folder"] = str(local_dir / folder_name)
        result["share_folder"] = str(share_dir / folder_name)
        result["local_file"] = str(local_dir / folder_name / report_path.name)
        result["share_file"] = str(share_dir / folder_name / report_path.name)
        result["version"] = 1
        return result
    
    # 使用新的筛选记录归档功能
    if run_dir and run_id:
        screening_result = archive_screening_records(
            run_dir, local_dir, share_dir, run_date, report_path, run_id, dry_run
        )
        result.update({
            "local_status": screening_result["local_status"],
            "share_status": screening_result["share_status"],
            "local_folder": screening_result["local_folder"],
            "share_folder": screening_result["share_folder"],
            "version": screening_result["version"],
            "archived_items": screening_result["archived_items"],
        })
        # 兼容旧字段
        result["local_file"] = str(Path(screening_result["local_folder"]) / report_path.name)
        result["share_file"] = str(Path(screening_result["share_folder"]) / report_path.name)
    else:
        # 回退到旧逻辑：只归档 Excel
        local_dir.mkdir(parents=True, exist_ok=True)
        local_file = versioned_path(local_dir, f"照明招标线索日报_广东_{run_date}_v8.5", ".xlsx")
        shutil.copy2(report_path, local_file)
        result["local_file"] = str(local_file)
        result["local_status"] = "ok"
        
        if str(share_dir).startswith(r"\\") and share_dir.exists():
            share_file = versioned_path(share_dir, f"照明招标线索日报_广东_{run_date}_v1.0", ".xlsx")
            shutil.copy2(report_path, share_file)
            result["share_file"] = str(share_file)
            result["share_status"] = "ok"
        else:
            raise GateError(f"共享盘路径不可达：{share_dir}")
    
    return result


def bid_label(url: str) -> str:
    match = re.search(r"bid-(\d+)", url or "")
    return f"bid-{match.group(1)}.html" if match else "千里马链接"


def build_message(rows: list[dict], run_date: str, source_start: str, source_end: str, raw_count: str, archive: dict, dingpan: dict) -> str:
    keep = [r for r in rows if r["decision"] == "keep"]
    review_rows = [r for r in rows if r["decision"] == "needs_review"]
    rejects = [r for r in rows if r["decision"] == "reject"]
    lines = [
        f"## {iso_date(run_date)} 招投标日报 · 🦐 QClaw标神",
        "",
        f"**筛选结果**：🔥推荐候选{len(keep)}条 | 📝待人工复核{len(review_rows)}条 | 🚫排除{len(rejects)}条",
        "",
        f"**数据源**：千里马导出_广东_{source_start}至{source_end}（原始{raw_count}条 → 复判{len(rows)}条）",
        "",
        "---",
        "",
        "### 🔥 推荐候选（需业务确认）",
        "",
    ]
    if keep:
        for index, row in enumerate(keep, 1):
            lines.append(f"**{index}. [{row['tag']}] {row['title']}** [{row['level']}级]")
            lines.append(f"> {row['province']}-{row['city']} | 预算：{row['amount']} | {row['second_reason']}")
            if row["url"]:
                lines.append(f"> 🔗 千里马链接：[{bid_label(row['url'])}]({row['url']})")
            lines.append("")
    else:
        lines.extend(["- 无", ""])

    lines.extend(["---", "", f"### 📝 待人工复核（共{len(review_rows)}条，仅显示前10条）", ""])
    if review_rows:
        for index, row in enumerate(review_rows[:10], 1):
            lines.append(f"{index}. **[{row['tag']}] {row['title']}** | {row['amount']} | {row['province']}-{row['city']}")
            lines.append(f"> {concise_message_reason(row['manual_review_reason'] or row['second_reason'])}")
            if row["url"]:
                lines.append(f"> 🔗 千里马链接：[{bid_label(row['url'])}]({row['url']})")
            lines.append("")
        if len(review_rows) > 10:
            lines.append(f"*... 还有{len(review_rows) - 10}条，详见钉钉文档*\n")
    else:
        lines.extend(["- 无", ""])

    lines.extend(["---", "", "### 🚫 排除摘要", ""])
    if rejects:
        for row in rejects[:8]:
            reason = concise_message_reason(
                first_nonempty(
                    row.get("second_reason"),
                    (row.get("raw") or {}).get("reason"),
                    (row.get("raw") or {}).get("reject_reason"),
                    row.get("exclude_reason"),
                    default="",
                ),
                60,
            )
            suffix = f"（{reason}）" if reason else ""
            lines.append(f"- [{row['tag']}] {row['title']}{suffix}")
    else:
        lines.append("- 无")

    doc_url = str(dingpan.get("doc_url") or "").strip()
    document_line = (
        f"• 📎 **钉钉文档**：[照明招标线索日报_广东_{run_date}_v8.5.xlsx]({doc_url})"
        if doc_url
        else "• 📎 **钉钉文档**：正式运行上传后生成链接"
    )
    
    # 使用新的文件夹路径格式
    local_folder = archive.get("local_folder", "")
    share_folder = archive.get("share_folder", "")
    version = archive.get("version", 1)
    
    lines.extend([
        "",
        "---",
        "",
        "### 📁 文件位置",
        "",
        document_line,
        "",
        f"• 💻 **本地存档**：`{local_folder}`",
        "",
        f"• 🗂️ **共享盘**：`{share_folder}`",
    ])
    return "\n".join(lines)


def validate_message_for_send(text: str, dingpan: dict, archive: dict) -> None:
    if not dingpan.get("doc_url"):
        raise GateError("DingTalk message blocked: missing Dingpan doc_url.")
    if archive.get("local_status") != "ok" or archive.get("share_status") != "ok":
        raise GateError("DingTalk message blocked: archive gate not ok.")
    for token in BAD_MESSAGE_TOKENS:
        if token in text:
            raise GateError(f"DingTalk message blocked: bad token found: {token}")


def append_project_run_log(run_date: str, manifest_out: dict, anomaly_summary: dict) -> None:
    log_path = SOP_ROOT / f"项目运行日志_{run_date}.md"
    lines = [
        "",
        f"## 有效运行记录 {manifest_out.get('run_id', '')}",
        "",
        "- 状态：Phase 2/3/4 校验通过后写入；以本条 run_manifest 为有效结果。",
        f"- 复判统计：推荐候选 {manifest_out.get('counts', {}).get('keep', 0)}；待人工复核 {manifest_out.get('counts', {}).get('needs_review', 0)}；排除 {manifest_out.get('counts', {}).get('reject', 0)}。",
        f"- Phase 2 校验：{manifest_out.get('phase2_validation', {}).get('ok')}; Excel 校验：{manifest_out.get('excel_validation', {}).get('ok')}; 业务异常复核：{anomaly_summary}.",
        f"- Excel：{manifest_out.get('report_xlsx', '')}",
        f"- 本地归档：{manifest_out.get('local_archive', '')}",
        f"- 共享盘：{manifest_out.get('share_archive', '')}",
        f"- 钉盘：{manifest_out.get('dingpan_doc_url', '')}",
        f"- 钉钉结果文件：{manifest_out.get('dingtalk_result_json', '')}",
        "",
        "说明：如果同日存在更早失败或无效 run，应以后续通过校验并成功归档推送的本记录为准。",
    ]
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write("\n".join(lines) + "\n")


def safe_append_project_run_log(run_date: str, manifest_out: dict, anomaly_summary: dict, fallback_dir: Path) -> None:
    try:
        append_project_run_log(run_date, manifest_out, anomaly_summary)
        return
    except Exception as exc:
        fallback_dir.mkdir(parents=True, exist_ok=True)
        fallback_path = fallback_dir / f"project_run_log_write_failed_{manifest_out.get('run_id', '')}.json"
        write_json(
            fallback_path,
            {
                "ok": False,
                "target": str(SOP_ROOT / f"项目运行日志_{run_date}.md"),
                "error": str(exc),
                "note": "Project root log append failed after report generation; this fallback does not block the pipeline.",
            },
        )
        print(f"WARNING: project run log append failed; wrote fallback: {fallback_path}", file=sys.stderr)
def _send_single_webhook(webhook: str, text: str, title: str = "照明招标线索日报") -> dict:
    payload = {
        "msgtype": "markdown",
        "markdown": {
            "title": title,
            "text": text,
        },
    }
    request = urllib.request.Request(
        webhook,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        body = response.read().decode("utf-8", errors="replace")
    return {"webhook": webhook, "response": body}


def send_dingtalk_webhook(text: str, config_text: str, dry_run: bool) -> dict:
    webhook = read_config_value(config_text, "dingtalk_webhook_url")
    webhook_backup = read_config_value(config_text, "dingtalk_webhook_url_backup")
    result = {
        "status": "dry_run" if dry_run else "pending",
        "success": False,
        "webhook_configured": bool(webhook),
        "backup_webhook_configured": bool(webhook_backup),
        # S3：结构化 backup 结果，便于日志检索（从不回显完整 URL）
        "backup_webhook": {
            "success": False,
            "skipped_reason": "not_attempted",
            "error": "",
        },
    }
    if dry_run:
        result["reason"] = "dry-run: skipped DingTalk webhook"
        result["backup_webhook"]["skipped_reason"] = "dry_run"
        return result
    if not webhook:
        raise GateError("DingTalk webhook URL is missing.")
    # 主 webhook 发送
    try:
        main_result = _send_single_webhook(webhook, text)
        result.update({"status": "ok", "success": True, "response": main_result["response"]})
    except Exception as e:
        result.update({"status": "failed", "success": False, "error": str(e)})
        result["backup_webhook"]["skipped_reason"] = "primary_failed"
        # 主 webhook 失败，不抄送到备用 webhook
        return result
    # 主 webhook 成功，抄送到备用 webhook（内容相同）
    if not webhook_backup:
        result["backup_webhook"] = {
            "success": False,
            "skipped_reason": "backup_not_configured",
            "error": "",
        }
        return result
    try:
        backup_result = _send_single_webhook(webhook_backup, text)
        result["backup_webhook"] = {
            "success": True,
            "skipped_reason": "",
            "error": "",
            "response": backup_result["response"],
        }
    except Exception as e:
        # 默认 backup 失败不否掉 full_run；仅 WARNING
        result["backup_webhook"] = {
            "success": False,
            "skipped_reason": "",
            "error": str(e),
        }
        print(f"WARNING: backup webhook send failed: {e}", file=sys.stderr)
    return result


def load_manifest(path: Path | None) -> dict:
    if not path:
        return {}
    if not path.exists():
        raise FileNotFoundError(path)
    return read_json(path)


def manifest_value(manifest: dict, args, key: str, default: Any = None) -> Any:
    value = getattr(args, key, None)
    if value not in (None, ""):
        return value
    return manifest.get(key) or default


def main() -> int:
    parser = argparse.ArgumentParser(description="Formal Phase 3/4 report archive push script v06. Uses report_builder_20260615_v1 and the v05 authority template.")
    parser.add_argument("--manifest", help="run_manifest JSON path")
    parser.add_argument("--business-json", dest="business_json")
    parser.add_argument("--vip-json", dest="vip_json")
    parser.add_argument("--raw-export", dest="raw_export")
    parser.add_argument("--screening-json", dest="screening_json")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE))
    parser.add_argument("--config", default=DEFAULT_CONFIG)
    parser.add_argument("--run-date", dest="run_date")
    parser.add_argument("--source-start-date", dest="source_start_date")
    parser.add_argument("--source-end-date", dest="source_end_date")
    parser.add_argument("--run-id", dest="run_id")
    parser.add_argument("--run-dir", dest="run_dir")
    parser.add_argument(
        "--local-archive-dir",
        dest="local_archive_dir",
        default=DEFAULT_LOCAL_ARCHIVE_DIR,
        help="本地归档目录；也可用 QLM_LOCAL_ARCHIVE 或外部配置 local_archive_dir",
    )
    parser.add_argument(
        "--share-archive-dir",
        dest="share_archive_dir",
        default=DEFAULT_SHARE_ARCHIVE_DIR,
        help="共享盘归档目录；也可用 QLM_SHARE_ARCHIVE 或外部配置 share_archive_dir",
    )
    parser.add_argument("--dry-run", action="store_true", help="Generate report and message but skip archive/upload/send gates.")
    parser.add_argument("--no-send", action="store_true", help="Archive and upload, but do not send DingTalk.")
    args = parser.parse_args()

    manifest_path = Path(args.manifest) if args.manifest else None
    manifest = load_manifest(manifest_path)
    run_date = compact_date(manifest_value(manifest, args, "run_date"))
    source_start, source_end = default_source_range(run_date)
    source_start = manifest_value(manifest, args, "source_start_date", source_start)
    source_end = manifest_value(manifest, args, "source_end_date", source_end)
    run_id = manifest_value(manifest, args, "run_id", f"{run_date}_{datetime.now().strftime('%H%M%S')}")
    run_dir = Path(manifest_value(manifest, args, "run_dir", str(DEFAULT_RUNS_DIR / run_id)))
    business_path = Path(manifest_value(manifest, args, "business_json"))
    vip_path_value = manifest_value(manifest, args, "vip_json", "")
    vip_path = Path(vip_path_value) if vip_path_value else None
    raw_export_value = manifest_value(manifest, args, "raw_export", "")
    raw_export = Path(raw_export_value) if raw_export_value else None
    screening_value = manifest_value(manifest, args, "screening_json", "")
    screening_path = Path(screening_value) if screening_value else None
    if not screening_path:
        cands = sorted((SOP_ROOT / "output" / "v2_4").glob("标讯筛选结果_*.json"))
        screening_path = cands[-1] if cands else None
    template = Path(args.template)
    config_path = Path(args.config).expanduser() if args.config else None

    if not business_path.exists():
        raise FileNotFoundError(f"business-json not found: {business_path}")
    if not template.exists():
        raise FileNotFoundError(f"template not found: {template}")

    report_dir = run_dir / "06_report"
    push_dir = run_dir / "07_archive_push"
    log_dir = run_dir / "99_logs"
    for directory in [report_dir, push_dir, log_dir]:
        directory.mkdir(parents=True, exist_ok=True)

    phase2_validation_path = log_dir / f"phase2_validation_{run_id}.json"
    validator_args = ["--business-json", str(business_path), "--output", str(phase2_validation_path)]
    if vip_path:
        validator_args += ["--vip-json", str(vip_path)]
    run_python(PHASE2_VALIDATOR, validator_args)
    phase2_validation = read_json(phase2_validation_path)

    business = read_json(business_path)
    enforce_evidence_release_gate(business, args.dry_run)
    business_items = flatten_results(business)
    attachment_evidence = attachment_evidence_summary(business_items)
    if attachment_evidence["systemic_failure"] and not args.dry_run:
        raise GateError(
            "附件证据通道整体失败：存在必读附件，但成功读取数为0。仅允许 dry-run 草稿，禁止正式归档和钉钉推送。"
        )
    rows = [build_row(item, index) for index, item in enumerate(business_items, 1)]
    rows = sort_rows(rows)
    if not rows:
        raise GateError("No rows after Phase 2 normalization.")

    business_anomalies = collect_business_anomalies(rows)
    business_anomaly_check = business_anomaly_summary(business_anomalies)
    business_anomaly_path = log_dir / f"business_anomaly_review_{run_id}.json"
    write_json(business_anomaly_path, {"summary": business_anomaly_check, "anomalies": business_anomalies})
    if not business_anomaly_check["ok"]:
        raise GateError(f"Business anomaly gate blocked send: {business_anomaly_check}")

    report_path = report_dir / f"照明招标线索日报_广东_{run_date}_v8.5.xlsx"
    if not raw_export or not Path(raw_export).exists():
        raise GateError("缺少原始导出 raw_export：新版『全部标讯』需要它做底表与字段关联。")
    if not screening_path or not Path(screening_path).exists():
        raise GateError(f"缺少标讯筛选结果 JSON（评分明细来源）：{screening_path}")
    build_workbook(str(template), str(raw_export), str(screening_path), str(business_path),
                   str(vip_path) if vip_path else "", str(report_path),
                   run_date, source_start, source_end)

    excel_validation_path = log_dir / f"excel_validation_{run_id}.json"
    excel_validation = run_python(
        EXCEL_VALIDATOR,
        ["--template", str(template), "--workbook", str(report_path), "--business-json", str(business_path)],
    )
    excel_validation_path.write_text(excel_validation["stdout"], encoding="utf-8")
    excel_validation_summary = json.loads(excel_validation["stdout"])

    config_text = read_config_text(config_path)
    local_archive_value = args.local_archive_dir or read_config_value(config_text, "local_archive_dir")
    share_archive_value = args.share_archive_dir or read_config_value(config_text, "share_archive_dir")
    if args.dry_run:
        local_archive_value = local_archive_value or str(run_dir / "dry-run-local-archive")
        share_archive_value = share_archive_value or str(run_dir / "dry-run-share-archive")
    if not local_archive_value or not share_archive_value:
        raise GateError("Archive paths are not configured. Set QLM_LOCAL_ARCHIVE and QLM_SHARE_ARCHIVE, pass archive arguments, or add local_archive_dir/share_archive_dir to the external config.")
    archive = archive_report(
        report_path,
        run_date,
        Path(local_archive_value),
        Path(share_archive_value),
        args.dry_run,
        run_dir=run_dir,
        run_id=run_id,
    )
    archive_path = push_dir / f"归档结果_广东_{run_date}_{run_id}.json"
    write_json(archive_path, archive)

    dingpan = upload_to_dingpan(report_path, config_text, args.dry_run)
    dingpan_path = push_dir / f"钉盘上传结果_广东_{run_date}_{run_id}.json"
    write_json(dingpan_path, dingpan)

    screening_payload = read_json(screening_path) if screening_path and Path(screening_path).exists() else {}
    screening_meta = screening_payload.get("meta", {}) if isinstance(screening_payload, dict) else {}
    raw_count = str(
        manifest.get("raw_export_count")
        or screening_meta.get("total_raw")
        or business.get("summary", {}).get("total_raw")
        or "—"
    )
    message_text = build_message(rows, run_date, source_start, source_end, raw_count, archive, dingpan)
    message_path = push_dir / f"钉钉消息草稿_广东_{run_date}_{run_id}.md"
    message_path.write_text(message_text, encoding="utf-8")

    counts = {
        "total": len(rows),
        "keep": sum(1 for r in rows if r["decision"] == "keep"),
        "needs_review": sum(1 for r in rows if r["decision"] == "needs_review"),
        "reject": sum(1 for r in rows if r["decision"] == "reject"),
    }

    dingtalk_result = {"status": "not_sent", "success": False, "reason": "no-send requested"}
    if args.dry_run:
        dingtalk_result = {
            "status": "dry_run",
            "success": False,
            "reason": "dry-run: message draft generated; no Dingpan upload or DingTalk send was attempted",
        }
    elif not args.no_send:
        validate_message_for_send(message_text, dingpan, archive)
        dingtalk_result = send_dingtalk_webhook(message_text, config_text, False)
    dingtalk_path = push_dir / f"钉钉发送结果_广东_{run_date}_{run_id}.json"
    write_json(dingtalk_path, dingtalk_result)

    manifest_out = {
        **manifest,
        "run_id": run_id,
        "run_date": run_date,
        "source_start_date": source_start,
        "source_end_date": source_end,
        "run_dir": str(run_dir),
        "business_json": str(business_path),
        "vip_json": str(vip_path) if vip_path else "",
        "raw_export": str(raw_export) if raw_export else "",
        "counts": counts,
        "raw_export_count": raw_count,
        "attachment_evidence": attachment_evidence,
        "phase2_validation": {
            "ok": phase2_validation.get("ok"),
            "rows": phase2_validation.get("rows"),
            "errors": phase2_validation.get("errors", []),
            "warnings": phase2_validation.get("warnings", []),
        },
        "excel_validation": {
            "ok": excel_validation_summary.get("ok"),
            "project_sheet_count": excel_validation_summary.get("project_sheet_count"),
            "errors": excel_validation_summary.get("errors", []),
            "warnings": excel_validation_summary.get("warnings", []),
        },
        "business_anomaly_review": business_anomaly_check,
        "report_xlsx": str(report_path),
        "phase2_validation_json": str(phase2_validation_path),
        "excel_validation_json": str(excel_validation_path),
        "business_anomaly_review_json": str(business_anomaly_path),
        "archive_json": str(archive_path),
        "dingpan_json": str(dingpan_path),
        "dingtalk_message_md": str(message_path),
        "dingtalk_result_json": str(dingtalk_path),
        "local_archive": archive.get("local_folder", archive.get("local_file", "")),
        "share_archive": archive.get("share_folder", archive.get("share_file", "")),
        "dingpan_doc_url": dingpan.get("doc_url", ""),
        "dry_run": args.dry_run,
        "no_send": args.no_send,
        # full_run_success：仅「非 dry-run 且非 no_send 且归档+钉盘+钉钉全成功」
        "full_run_success": bool(
            not args.dry_run
            and not args.no_send
            and archive.get("local_status") == "ok"
            and archive.get("share_status") == "ok"
            and dingpan.get("success")
            and dingtalk_result.get("success")
        ),
        # P2-01：阶段成功与「生产全链路绿灯」拆开，避免 dry-run 被误读为失败
        "phase34_ok": True,
        "report_build_ok": bool(report_path and Path(report_path).exists()),
        "success_semantics": {
            "ok": "本脚本无异常并完成约定产物（含 dry-run 草稿/日报）",
            "phase34_ok": "Phase3/4 报告生成流程完成",
            "full_run_success": "非dry-run且归档+钉盘+钉钉均成功；dry-run 时必为 false（预期）",
        },
    }
    manifest_out_path = run_dir / f"run_manifest_{run_id}.json"
    write_json(manifest_out_path, manifest_out)
    safe_append_project_run_log(run_date, manifest_out, business_anomaly_check, log_dir)

    phase_status = (
        "dry_run_complete" if args.dry_run
        else ("archive_only_complete" if args.no_send else "full_run_complete")
    )
    full_run_success = bool(manifest_out.get("full_run_success"))
    result = {
        "ok": True,
        "status": phase_status,
        "phase34_ok": True,
        "report_build_ok": bool(manifest_out.get("report_build_ok")),
        "full_run_success": full_run_success,
        "run_id": run_id,
        "dry_run": args.dry_run,
        "no_send": args.no_send,
        "success_semantics": manifest_out.get("success_semantics"),
        "note": (
            "dry_run=true 时 full_run_success=false 为预期，请看 ok/phase34_ok/status=dry_run_complete"
            if args.dry_run
            else (
                "no_send=true 时 full_run_success=false 为预期（未发钉钉），请看 ok/status=archive_only_complete"
                if args.no_send
                else "full_run_success=true 表示归档+钉盘+钉钉均成功"
            )
        ),
        "manifest": str(manifest_out_path),
        "report": str(report_path),
        "message": str(message_path),
        "archive": archive,
        "dingpan": {k: v for k, v in dingpan.items() if k not in {"stdout", "stderr", "raw"}},
        "dingtalk": dingtalk_result,
        "counts": counts,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False, indent=2), file=sys.stderr)
        raise


